from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from contextlib import asynccontextmanager
import os
import logging
import json
import re
import io
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import asyncio
from serpapi import GoogleSearch
import httpx
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urljoin, urlparse, parse_qs, urlencode, urlunparse
import random as rand_module
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import zipfile

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging (must be early for other setup code)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection (optional)
try:
    mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
    client = AsyncIOMotorClient(mongo_url)
    db = client[os.environ.get('DB_NAME', 'pricenexus')]
    MONGODB_AVAILABLE = True
    logger.info(f"MongoDB connected: {mongo_url}")
except Exception as e:
    MONGODB_AVAILABLE = False
    db = None
    logger.warning(f"MongoDB not available: {e}. Search history will not be saved.")

# SerpAPI configuration
SERPAPI_API_KEY = os.environ.get('SERPAPI_API_KEY', '')

# RapidAPI configuration (FREE - 500+ requests/month across multiple product APIs)
RAPIDAPI_KEY = os.environ.get('RAPIDAPI_KEY', '')

# Foursquare API configuration (disabled - requires payment)
# GOOGLE_PLACES_API_KEY = os.environ.get('GOOGLE_PLACES_API_KEY', '')

# Foursquare Places API configuration (FREE - 100k calls/month)
FOURSQUARE_API_KEY = os.environ.get('FOURSQUARE_API_KEY', '')

# Lifespan context manager for startup/shutdown events
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: nothing specific needed
    yield
    # Shutdown: close MongoDB client if available
    if MONGODB_AVAILABLE and client:
        client.close()
        logger.info("MongoDB connection closed")

# Create the main app without a prefix
app = FastAPI(lifespan=lifespan)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ================== MODELS ==================
class SearchRequest(BaseModel):
    query: str
    max_results: int = 50

class SearchResult(BaseModel):
    name: str
    price: float
    currency_symbol: str
    currency_code: str
    source: str
    source_url: str
    description: str
    rating: float
    availability: str
    unit: str
    last_updated: str
    image: str
    location: str

class DataSource(BaseModel):
    name: str
    url: str
    type: str
    description: str

class SearchResponse(BaseModel):
    success: bool
    query: str
    message: Optional[str] = None
    response: str
    results: List[Dict[str, Any]]
    results_count: int
    ai_model: str
    data_sources: List[Dict[str, Any]]
    # Advanced filter options extracted from AI
    available_filters: Optional[Dict[str, Any]] = None
    # Local stores from Foursquare API
    local_stores: Optional[List[Dict[str, Any]]] = None
    local_stores_city: Optional[str] = None

# ================== CURRENCY DATA ==================
CURRENCY_DATA = {
    "india": {"symbol": "₹", "rate": 1.0, "code": "INR"},
    "usa": {"symbol": "$", "rate": 0.012, "code": "USD"},
    "uk": {"symbol": "£", "rate": 0.0095, "code": "GBP"},
    "uae": {"symbol": "AED", "rate": 0.044, "code": "AED"},
    "europe": {"symbol": "€", "rate": 0.011, "code": "EUR"},
    "japan": {"symbol": "¥", "rate": 1.8, "code": "JPY"},
    "australia": {"symbol": "A$", "rate": 0.018, "code": "AUD"},
    "canada": {"symbol": "C$", "rate": 0.016, "code": "CAD"},
    "global": {"symbol": "$", "rate": 0.012, "code": "USD"}
}

# ================== LOCATION CITIES DATABASE ==================
CITIES_DB = {
    "mumbai": {"city": "Mumbai", "state": "Maharashtra", "country": "india"},
    "delhi": {"city": "Delhi", "state": "Delhi NCR", "country": "india"},
    "bangalore": {"city": "Bangalore", "state": "Karnataka", "country": "india"},
    "bengaluru": {"city": "Bangalore", "state": "Karnataka", "country": "india"},
    "chennai": {"city": "Chennai", "state": "Tamil Nadu", "country": "india"},
    "hyderabad": {"city": "Hyderabad", "state": "Telangana", "country": "india"},
    "kolkata": {"city": "Kolkata", "state": "West Bengal", "country": "india"},
    "pune": {"city": "Pune", "state": "Maharashtra", "country": "india"},
    "ahmedabad": {"city": "Ahmedabad", "state": "Gujarat", "country": "india"},
    "new york": {"city": "New York", "state": "New York", "country": "usa"},
    "los angeles": {"city": "Los Angeles", "state": "California", "country": "usa"},
    "chicago": {"city": "Chicago", "state": "Illinois", "country": "usa"},
    "houston": {"city": "Houston", "state": "Texas", "country": "usa"},
    "san francisco": {"city": "San Francisco", "state": "California", "country": "usa"},
    "london": {"city": "London", "state": "England", "country": "uk"},
    "manchester": {"city": "Manchester", "state": "England", "country": "uk"},
    "birmingham": {"city": "Birmingham", "state": "England", "country": "uk"},
    "dubai": {"city": "Dubai", "state": "Dubai", "country": "uae"},
    "abu dhabi": {"city": "Abu Dhabi", "state": "Abu Dhabi", "country": "uae"},
    "tokyo": {"city": "Tokyo", "state": "Tokyo", "country": "japan"},
    "sydney": {"city": "Sydney", "state": "NSW", "country": "australia"},
    "melbourne": {"city": "Melbourne", "state": "Victoria", "country": "australia"},
    "toronto": {"city": "Toronto", "state": "Ontario", "country": "canada"},
    "vancouver": {"city": "Vancouver", "state": "BC", "country": "canada"},
    "paris": {"city": "Paris", "state": "Île-de-France", "country": "europe"},
    "berlin": {"city": "Berlin", "state": "Berlin", "country": "europe"},
}

COUNTRY_KEYWORDS = {
    "india": "india",
    "indian": "india",
    "usa": "usa",
    "united states": "usa",
    "america": "usa",
    "american": "usa",
    "uk": "uk",
    "united kingdom": "uk",
    "britain": "uk",
    "british": "uk",
    "england": "uk",
    "uae": "uae",
    "dubai": "uae",
    "emirates": "uae",
    "japan": "japan",
    "japanese": "japan",
    "australia": "australia",
    "australian": "australia",
    "canada": "canada",
    "canadian": "canada",
    "europe": "europe",
    "european": "europe",
    "germany": "europe",
    "france": "europe",
}

# ================== AI INTEGRATION ==================
# Using fallback mode for AI features (emergentintegrations not available)
EMERGENT_AVAILABLE = False

# Define stub classes for AI functionality
class LlmChat:
    def __init__(self, *args, **kwargs):
        pass
    def with_model(self, *args, **kwargs):
        return self
    async def send_message(self, *args, **kwargs):
        return ""

class UserMessage:
    def __init__(self, text=""):
        self.text = text

# ================== LOCATION AND CURRENCY ==================

def extract_location(query: str) -> Dict[str, str]:
    """Extract location from query"""
    query_lower = query.lower()

    
    # Check for country keywords first
    for keyword, country in COUNTRY_KEYWORDS.items():
        if keyword in query_lower:
            return {"city": "Various Cities", "state": "Nationwide", "country": country}
    
    # Check for specific cities
    for city_key, city_data in CITIES_DB.items():
        if city_key in query_lower:
            return city_data
    
    # Default to global
    return {"city": "Global", "state": "International", "country": "global"}

def get_currency_info(country: str) -> Dict[str, Any]:
    """Get currency info for country"""
    return CURRENCY_DATA.get(country.lower(), CURRENCY_DATA["global"])

# Cache for dynamic marketplaces
_marketplace_cache = {}

async def discover_marketplaces_with_ai(product_name: str, category: str, country: str, source_type: str) -> List[Dict[str, str]]:
    """Use AI to dynamically discover relevant marketplaces for a specific product"""
    cache_key = f"{product_name}_{category}_{country}_{source_type}"
    
    if cache_key in _marketplace_cache:
        return _marketplace_cache[cache_key]
    
    try:
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key or not EMERGENT_AVAILABLE:
            return get_fallback_marketplaces(country, source_type)
        
        source_descriptions = {
            "global_suppliers": "B2B wholesale suppliers, international trade platforms, manufacturer directories",
            "local_markets": "local retail stores, regional dealers, neighborhood shops, local business directories",
            "online_marketplaces": "e-commerce websites, online retail stores, digital shopping platforms"
        }
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"marketplace-discovery-{uuid.uuid4()}",
            system_message="You are a marketplace expert. Return ONLY valid JSON with no extra text."
        )
        chat.with_model("openai", "gpt-4o")
        
        prompt = f"""Find REAL marketplaces where "{product_name}" (category: {category}) is sold in {country.upper()}.
        
Source type: {source_type} ({source_descriptions.get(source_type, 'various marketplaces')})

Return a JSON array with 4-6 REAL, EXISTING marketplaces. Each must have:
- "name": Real marketplace/store name (must actually exist)
- "url": Real search URL pattern (use actual website search URLs)

Example format:
[
    {{"name": "StoreName", "url": "https://www.storename.com/search?q="}}
]

IMPORTANT:
- Only include REAL marketplaces that actually sell {product_name}
- Use actual working search URLs for those websites
- Be specific to the product category - don't use generic marketplaces unless they're relevant
- For {source_type}, focus on {source_descriptions.get(source_type, 'relevant platforms')}
- Consider {country.upper()} regional marketplaces

Return ONLY the JSON array, no other text."""

        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        # Parse JSON response
        response_text = response.strip()
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(lines[1:-1] if lines[-1] == "```" else lines[1:])
        
        marketplaces = json.loads(response_text)
        
        if isinstance(marketplaces, list) and len(marketplaces) > 0:
            # Validate structure
            valid_marketplaces = []
            for mp in marketplaces:
                if isinstance(mp, dict) and "name" in mp and "url" in mp:
                    valid_marketplaces.append(mp)
            
            if valid_marketplaces:
                _marketplace_cache[cache_key] = valid_marketplaces
                return valid_marketplaces
        
        return get_fallback_marketplaces(country, source_type)
        
    except Exception as e:
        logger.error(f"AI marketplace discovery error: {e}")
        return get_fallback_marketplaces(country, source_type)

def get_fallback_marketplaces(country: str, source_type: str) -> List[Dict[str, str]]:
    """Fallback marketplaces when AI discovery fails"""
    fallback = {
        "global_suppliers": [
            {"name": "Alibaba", "url": "https://www.alibaba.com/trade/search?SearchText="},
            {"name": "Global Sources", "url": "https://www.globalsources.com/searchList/products?search="},
            {"name": "Made-in-China", "url": "https://www.made-in-china.com/products-search/hot-china-products/"},
        ],
        "local_markets": [
            {"name": "Google Local", "url": "https://www.google.com/search?q="},
            {"name": "Yelp", "url": "https://www.yelp.com/search?find_desc="},
        ],
        "online_marketplaces": [
            {"name": "Amazon", "url": "https://www.amazon.com/s?k="},
            {"name": "eBay", "url": "https://www.ebay.com/sch/i.html?_nkw="},
        ]
    }
    return fallback.get(source_type, fallback["online_marketplaces"])

# ================== REMOVED: VENDOR GENERATION FUNCTIONS ==================
# The functions generate_vendor_details() and generate_vendor_for_real_source() 
# have been removed because they were never called and contained 380+ lines of 
# fake/random vendor contact data (phone, email, address).
# Real vendor data (name + website URL) is extracted directly from search results.

# ================== DIRECT VENDOR LINKS ==================
def get_direct_vendor_link(source_name: str, product_name: str) -> str:
    """
    Generate direct search link to vendor's website instead of Google redirect.
    Maps vendor names to their actual website search URLs.
    """
    source_lower = source_name.lower().strip()
    encoded_product = product_name.replace(" ", "+")
    
    # Comprehensive vendor URL mapping
    vendor_urls = {
        # Indian E-commerce
        "amazon": f"https://www.amazon.in/s?k={encoded_product}",
        "amazon.in": f"https://www.amazon.in/s?k={encoded_product}",
        "amazon india": f"https://www.amazon.in/s?k={encoded_product}",
        "flipkart": f"https://www.flipkart.com/search?q={encoded_product}",
        "myntra": f"https://www.myntra.com/{encoded_product}",
        "ajio": f"https://www.ajio.com/search/?text={encoded_product}",
        "tata cliq": f"https://www.tatacliq.com/search/?searchCategory=all&text={encoded_product}",
        "snapdeal": f"https://www.snapdeal.com/search?keyword={encoded_product}",
        "meesho": f"https://www.meesho.com/search?q={encoded_product}",
        "jiomart": f"https://www.jiomart.com/search/{encoded_product}",
        "jiomart electronics": f"https://www.jiomart.com/search/{encoded_product}",
        "reliance digital": f"https://www.reliancedigital.in/search?q={encoded_product}",
        "croma": f"https://www.croma.com/searchB?q={encoded_product}",
        "vijay sales": f"https://www.vijaysales.com/search/{encoded_product}",
        "poorvika": f"https://www.poorvikamobile.com/catalogsearch/result/?q={encoded_product}",
        "sangeetha": f"https://www.sangeethamobiles.com/catalogsearch/result/?q={encoded_product}",
        "sangeetha mobiles": f"https://www.sangeethamobiles.com/catalogsearch/result/?q={encoded_product}",
        
        # Mobile/Electronics specific India
        "cashify": f"https://www.cashify.in/buy-refurbished-mobiles?q={encoded_product}",
        "2gud": f"https://www.2gud.com/search?q={encoded_product}",
        "iplanet": f"https://www.iplanet.in/catalogsearch/result/?q={encoded_product}",
        "imagine": f"https://www.imagineonline.store/catalogsearch/result/?q={encoded_product}",
        "imagine apple": f"https://www.imagineonline.store/catalogsearch/result/?q={encoded_product}",
        "imagine apple premium reseller": f"https://www.imagineonline.store/catalogsearch/result/?q={encoded_product}",
        "apple store": f"https://www.apple.com/in/shop/buy-iphone",
        "apple": f"https://www.apple.com/in/shop/buy-iphone",
        "mi": f"https://www.mi.com/in/search?keyword={encoded_product}",
        "mi.com": f"https://www.mi.com/in/search?keyword={encoded_product}",
        "samsung": f"https://www.samsung.com/in/search/?searchvalue={encoded_product}",
        "oneplus": f"https://www.oneplus.in/search?keyword={encoded_product}",
        "dell": f"https://www.dell.com/en-in/search/{encoded_product}",
        "dell india": f"https://www.dell.com/en-in/search/{encoded_product}",
        "hp": f"https://www.hp.com/in-en/shop/search?q={encoded_product}",
        "hp india": f"https://www.hp.com/in-en/shop/search?q={encoded_product}",
        "lenovo": f"https://www.lenovo.com/in/en/search?text={encoded_product}",
        "asus": f"https://www.asus.com/in/searchresult?searchType=products&searchKey={encoded_product}",
        
        # B2B India - Directory Sites
        "indiamart": f"https://dir.indiamart.com/search.mp?ss={encoded_product}",
        "tradeindia": f"https://www.tradeindia.com/search.html?search_query={encoded_product}",
        "exportersindia": f"https://www.exportersindia.com/search.htm?search={encoded_product}",
        "justdial": f"https://www.justdial.com/search?q={encoded_product}",
        "jd": f"https://www.justdial.com/search?q={encoded_product}",
        "sulekha": f"https://www.sulekha.com/search?q={encoded_product}",
        "aajjo": f"https://www.aajjo.com/search?q={encoded_product}",
        "aajjo.com": f"https://www.aajjo.com/search?q={encoded_product}",
        "go4worldbusiness": f"https://www.go4worldbusiness.com/search/?keyword={encoded_product}",
        "exportershub": f"https://www.exportershub.com/search?q={encoded_product}",
        "dial4trade": f"https://www.dial4trade.com/search?q={encoded_product}",
        "infoisinfo": f"https://www.infoisinfo.co.in/search/{encoded_product}",
        "ovantica": f"https://www.ovantica.com/search?q={encoded_product}",
        "ovantica.com": f"https://www.ovantica.com/search?q={encoded_product}",
        
        # Construction/Building Materials India
        "buildmart": f"https://www.buildmart.in/search?q={encoded_product}",
        "materialtree": f"https://www.materialtree.com/search?q={encoded_product}",
        "buildingmaterialsdirect": f"https://www.buildingmaterialsdirect.co.in/search?q={encoded_product}",
        
        # US E-commerce
        "amazon.com": f"https://www.amazon.com/s?k={encoded_product}",
        "amazon us": f"https://www.amazon.com/s?k={encoded_product}",
        "walmart": f"https://www.walmart.com/search?q={encoded_product}",
        "target": f"https://www.target.com/s?searchTerm={encoded_product}",
        "best buy": f"https://www.bestbuy.com/site/searchpage.jsp?st={encoded_product}",
        "bestbuy": f"https://www.bestbuy.com/site/searchpage.jsp?st={encoded_product}",
        "newegg": f"https://www.newegg.com/p/pl?d={encoded_product}",
        "ebay": f"https://www.ebay.com/sch/i.html?_nkw={encoded_product}",
        "costco": f"https://www.costco.com/CatalogSearch?keyword={encoded_product}",
        
        # Home Improvement
        "home depot": f"https://www.homedepot.com/s/{encoded_product}",
        "homedepot": f"https://www.homedepot.com/s/{encoded_product}",
        "lowe's": f"https://www.lowes.com/search?searchTerm={encoded_product}",
        "lowes": f"https://www.lowes.com/search?searchTerm={encoded_product}",
        "ace hardware": f"https://www.acehardware.com/search?query={encoded_product}",
        "menards": f"https://www.menards.com/main/search.html?search={encoded_product}",
        
        # UK E-commerce
        "amazon.co.uk": f"https://www.amazon.co.uk/s?k={encoded_product}",
        "amazon uk": f"https://www.amazon.co.uk/s?k={encoded_product}",
        "argos": f"https://www.argos.co.uk/search/{encoded_product}",
        "currys": f"https://www.currys.co.uk/search?q={encoded_product}",
        "john lewis": f"https://www.johnlewis.com/search?search-term={encoded_product}",
        
        # UAE E-commerce
        "noon": f"https://www.noon.com/uae-en/search/?q={encoded_product}",
        "amazon.ae": f"https://www.amazon.ae/s?k={encoded_product}",
        "sharaf dg": f"https://uae.sharafdg.com/search/?q={encoded_product}",
        "lulu": f"https://www.luluhypermarket.com/en-ae/search?q={encoded_product}",
        
        # Global
        "alibaba": f"https://www.alibaba.com/trade/search?SearchText={encoded_product}",
        "aliexpress": f"https://www.aliexpress.com/wholesale?SearchText={encoded_product}",
    }
    
    # Try exact match first
    if source_lower in vendor_urls:
        return vendor_urls[source_lower]
    
    # Try partial match - but be more strict to avoid false positives
    # Only match if the vendor key is a significant part of the source name
    for vendor_key, url in vendor_urls.items():
        # Skip very short keys to avoid false matches
        if len(vendor_key) < 4:
            continue
        # Check if vendor key is contained in source (but not vice versa to avoid "mi" matching everything)
        if vendor_key in source_lower:
            return url
            return url
    
    # Return None to signal we should use the original SerpAPI link
    return None

# ================== GOOGLE PLACES API - LOCAL STORES ==================
# City coordinates for location-based searches
CITY_COORDINATES = {
    # India - Major Cities
    "bangalore": {"lat": 12.9716, "lng": 77.5946, "country": "India", "name": "Bengaluru"},
    "bengaluru": {"lat": 12.9716, "lng": 77.5946, "country": "India", "name": "Bengaluru"},
    "banglore": {"lat": 12.9716, "lng": 77.5946, "country": "India", "name": "Bengaluru"},
    "mumbai": {"lat": 19.0760, "lng": 72.8777, "country": "India", "name": "Mumbai"},
    "bombay": {"lat": 19.0760, "lng": 72.8777, "country": "India", "name": "Mumbai"},
    "delhi": {"lat": 28.6139, "lng": 77.2090, "country": "India", "name": "Delhi"},
    "new delhi": {"lat": 28.6139, "lng": 77.2090, "country": "India", "name": "Delhi"},
    "chennai": {"lat": 13.0827, "lng": 80.2707, "country": "India", "name": "Chennai"},
    "madras": {"lat": 13.0827, "lng": 80.2707, "country": "India", "name": "Chennai"},
    "hyderabad": {"lat": 17.3850, "lng": 78.4867, "country": "India", "name": "Hyderabad"},
    "kolkata": {"lat": 22.5726, "lng": 88.3639, "country": "India", "name": "Kolkata"},
    "calcutta": {"lat": 22.5726, "lng": 88.3639, "country": "India", "name": "Kolkata"},
    "pune": {"lat": 18.5204, "lng": 73.8567, "country": "India", "name": "Pune"},
    "ahmedabad": {"lat": 23.0225, "lng": 72.5714, "country": "India", "name": "Ahmedabad"},
    "jaipur": {"lat": 26.9124, "lng": 75.7873, "country": "India", "name": "Jaipur"},
    "lucknow": {"lat": 26.8467, "lng": 80.9462, "country": "India", "name": "Lucknow"},
    "surat": {"lat": 21.1702, "lng": 72.8311, "country": "India", "name": "Surat"},
    "kochi": {"lat": 9.9312, "lng": 76.2673, "country": "India", "name": "Kochi"},
    "cochin": {"lat": 9.9312, "lng": 76.2673, "country": "India", "name": "Kochi"},
    "chandigarh": {"lat": 30.7333, "lng": 76.7794, "country": "India", "name": "Chandigarh"},
    "indore": {"lat": 22.7196, "lng": 75.8577, "country": "India", "name": "Indore"},
    "nagpur": {"lat": 21.1458, "lng": 79.0882, "country": "India", "name": "Nagpur"},
    "bhopal": {"lat": 23.2599, "lng": 77.4126, "country": "India", "name": "Bhopal"},
    "coimbatore": {"lat": 11.0168, "lng": 76.9558, "country": "India", "name": "Coimbatore"},
    "visakhapatnam": {"lat": 17.6868, "lng": 83.2185, "country": "India", "name": "Visakhapatnam"},
    "vizag": {"lat": 17.6868, "lng": 83.2185, "country": "India", "name": "Visakhapatnam"},
    "patna": {"lat": 25.5941, "lng": 85.1376, "country": "India", "name": "Patna"},
    "vadodara": {"lat": 22.3072, "lng": 73.1812, "country": "India", "name": "Vadodara"},
    "goa": {"lat": 15.2993, "lng": 74.1240, "country": "India", "name": "Goa"},
    "india": {"lat": 20.5937, "lng": 78.9629, "country": "India", "name": "India"},
    
    # USA - Major Cities & States
    "new york": {"lat": 40.7128, "lng": -74.0060, "country": "USA", "name": "New York"},
    "nyc": {"lat": 40.7128, "lng": -74.0060, "country": "USA", "name": "New York"},
    "los angeles": {"lat": 34.0522, "lng": -118.2437, "country": "USA", "name": "Los Angeles"},
    "la": {"lat": 34.0522, "lng": -118.2437, "country": "USA", "name": "Los Angeles"},
    "chicago": {"lat": 41.8781, "lng": -87.6298, "country": "USA", "name": "Chicago"},
    "san francisco": {"lat": 37.7749, "lng": -122.4194, "country": "USA", "name": "San Francisco"},
    "sf": {"lat": 37.7749, "lng": -122.4194, "country": "USA", "name": "San Francisco"},
    "seattle": {"lat": 47.6062, "lng": -122.3321, "country": "USA", "name": "Seattle"},
    "boston": {"lat": 42.3601, "lng": -71.0589, "country": "USA", "name": "Boston"},
    "houston": {"lat": 29.7604, "lng": -95.3698, "country": "USA", "name": "Houston"},
    "dallas": {"lat": 32.7767, "lng": -96.7970, "country": "USA", "name": "Dallas"},
    "austin": {"lat": 30.2672, "lng": -97.7431, "country": "USA", "name": "Austin"},
    "miami": {"lat": 25.7617, "lng": -80.1918, "country": "USA", "name": "Miami"},
    "denver": {"lat": 39.7392, "lng": -104.9903, "country": "USA", "name": "Denver"},
    "phoenix": {"lat": 33.4484, "lng": -112.0740, "country": "USA", "name": "Phoenix"},
    "san diego": {"lat": 32.7157, "lng": -117.1611, "country": "USA", "name": "San Diego"},
    "san jose": {"lat": 37.3382, "lng": -121.8863, "country": "USA", "name": "San Jose"},
    "california": {"lat": 36.7783, "lng": -119.4179, "country": "USA", "name": "California"},
    "texas": {"lat": 31.9686, "lng": -99.9018, "country": "USA", "name": "Texas"},
    "florida": {"lat": 27.6648, "lng": -81.5158, "country": "USA", "name": "Florida"},
    "usa": {"lat": 37.0902, "lng": -95.7129, "country": "USA", "name": "USA"},
    "america": {"lat": 37.0902, "lng": -95.7129, "country": "USA", "name": "USA"},
    
    # UK
    "london": {"lat": 51.5074, "lng": -0.1278, "country": "UK", "name": "London"},
    "manchester": {"lat": 53.4808, "lng": -2.2426, "country": "UK", "name": "Manchester"},
    "birmingham": {"lat": 52.4862, "lng": -1.8904, "country": "UK", "name": "Birmingham"},
    "liverpool": {"lat": 53.4084, "lng": -2.9916, "country": "UK", "name": "Liverpool"},
    "edinburgh": {"lat": 55.9533, "lng": -3.1883, "country": "UK", "name": "Edinburgh"},
    "glasgow": {"lat": 55.8642, "lng": -4.2518, "country": "UK", "name": "Glasgow"},
    "uk": {"lat": 55.3781, "lng": -3.4360, "country": "UK", "name": "UK"},
    "england": {"lat": 52.3555, "lng": -1.1743, "country": "UK", "name": "England"},
    "britain": {"lat": 55.3781, "lng": -3.4360, "country": "UK", "name": "UK"},
    
    # UAE & Middle East
    "dubai": {"lat": 25.2048, "lng": 55.2708, "country": "UAE", "name": "Dubai"},
    "abu dhabi": {"lat": 24.4539, "lng": 54.3773, "country": "UAE", "name": "Abu Dhabi"},
    "sharjah": {"lat": 25.3463, "lng": 55.4209, "country": "UAE", "name": "Sharjah"},
    "uae": {"lat": 23.4241, "lng": 53.8478, "country": "UAE", "name": "UAE"},
    "saudi": {"lat": 23.8859, "lng": 45.0792, "country": "Saudi Arabia", "name": "Saudi Arabia"},
    "riyadh": {"lat": 24.7136, "lng": 46.6753, "country": "Saudi Arabia", "name": "Riyadh"},
    "qatar": {"lat": 25.3548, "lng": 51.1839, "country": "Qatar", "name": "Qatar"},
    "doha": {"lat": 25.2854, "lng": 51.5310, "country": "Qatar", "name": "Doha"},
    "kuwait": {"lat": 29.3759, "lng": 47.9774, "country": "Kuwait", "name": "Kuwait"},
    "bahrain": {"lat": 26.0667, "lng": 50.5577, "country": "Bahrain", "name": "Bahrain"},
    
    # Asia Pacific
    "tokyo": {"lat": 35.6762, "lng": 139.6503, "country": "Japan", "name": "Tokyo"},
    "osaka": {"lat": 34.6937, "lng": 135.5023, "country": "Japan", "name": "Osaka"},
    "japan": {"lat": 36.2048, "lng": 138.2529, "country": "Japan", "name": "Japan"},
    "singapore": {"lat": 1.3521, "lng": 103.8198, "country": "Singapore", "name": "Singapore"},
    "hong kong": {"lat": 22.3193, "lng": 114.1694, "country": "Hong Kong", "name": "Hong Kong"},
    "hongkong": {"lat": 22.3193, "lng": 114.1694, "country": "Hong Kong", "name": "Hong Kong"},
    "seoul": {"lat": 37.5665, "lng": 126.9780, "country": "South Korea", "name": "Seoul"},
    "korea": {"lat": 35.9078, "lng": 127.7669, "country": "South Korea", "name": "South Korea"},
    "bangkok": {"lat": 13.7563, "lng": 100.5018, "country": "Thailand", "name": "Bangkok"},
    "thailand": {"lat": 15.8700, "lng": 100.9925, "country": "Thailand", "name": "Thailand"},
    "kuala lumpur": {"lat": 3.1390, "lng": 101.6869, "country": "Malaysia", "name": "Kuala Lumpur"},
    "kl": {"lat": 3.1390, "lng": 101.6869, "country": "Malaysia", "name": "Kuala Lumpur"},
    "malaysia": {"lat": 4.2105, "lng": 101.9758, "country": "Malaysia", "name": "Malaysia"},
    "jakarta": {"lat": -6.2088, "lng": 106.8456, "country": "Indonesia", "name": "Jakarta"},
    "indonesia": {"lat": -0.7893, "lng": 113.9213, "country": "Indonesia", "name": "Indonesia"},
    "manila": {"lat": 14.5995, "lng": 120.9842, "country": "Philippines", "name": "Manila"},
    "philippines": {"lat": 12.8797, "lng": 121.7740, "country": "Philippines", "name": "Philippines"},
    "vietnam": {"lat": 14.0583, "lng": 108.2772, "country": "Vietnam", "name": "Vietnam"},
    "hanoi": {"lat": 21.0285, "lng": 105.8542, "country": "Vietnam", "name": "Hanoi"},
    "ho chi minh": {"lat": 10.8231, "lng": 106.6297, "country": "Vietnam", "name": "Ho Chi Minh City"},
    
    # Australia & New Zealand
    "sydney": {"lat": -33.8688, "lng": 151.2093, "country": "Australia", "name": "Sydney"},
    "melbourne": {"lat": -37.8136, "lng": 144.9631, "country": "Australia", "name": "Melbourne"},
    "brisbane": {"lat": -27.4698, "lng": 153.0251, "country": "Australia", "name": "Brisbane"},
    "perth": {"lat": -31.9505, "lng": 115.8605, "country": "Australia", "name": "Perth"},
    "australia": {"lat": -25.2744, "lng": 133.7751, "country": "Australia", "name": "Australia"},
    "auckland": {"lat": -36.8485, "lng": 174.7633, "country": "New Zealand", "name": "Auckland"},
    "new zealand": {"lat": -40.9006, "lng": 174.8860, "country": "New Zealand", "name": "New Zealand"},
    
    # Canada
    "toronto": {"lat": 43.6532, "lng": -79.3832, "country": "Canada", "name": "Toronto"},
    "vancouver": {"lat": 49.2827, "lng": -123.1207, "country": "Canada", "name": "Vancouver"},
    "montreal": {"lat": 45.5017, "lng": -73.5673, "country": "Canada", "name": "Montreal"},
    "calgary": {"lat": 51.0447, "lng": -114.0719, "country": "Canada", "name": "Calgary"},
    "ottawa": {"lat": 45.4215, "lng": -75.6972, "country": "Canada", "name": "Ottawa"},
    "canada": {"lat": 56.1304, "lng": -106.3468, "country": "Canada", "name": "Canada"},
    
    # Europe
    "paris": {"lat": 48.8566, "lng": 2.3522, "country": "France", "name": "Paris"},
    "france": {"lat": 46.2276, "lng": 2.2137, "country": "France", "name": "France"},
    "berlin": {"lat": 52.5200, "lng": 13.4050, "country": "Germany", "name": "Berlin"},
    "munich": {"lat": 48.1351, "lng": 11.5820, "country": "Germany", "name": "Munich"},
    "frankfurt": {"lat": 50.1109, "lng": 8.6821, "country": "Germany", "name": "Frankfurt"},
    "germany": {"lat": 51.1657, "lng": 10.4515, "country": "Germany", "name": "Germany"},
    "amsterdam": {"lat": 52.3676, "lng": 4.9041, "country": "Netherlands", "name": "Amsterdam"},
    "netherlands": {"lat": 52.1326, "lng": 5.2913, "country": "Netherlands", "name": "Netherlands"},
    "rome": {"lat": 41.9028, "lng": 12.4964, "country": "Italy", "name": "Rome"},
    "milan": {"lat": 45.4642, "lng": 9.1900, "country": "Italy", "name": "Milan"},
    "italy": {"lat": 41.8719, "lng": 12.5674, "country": "Italy", "name": "Italy"},
    "madrid": {"lat": 40.4168, "lng": -3.7038, "country": "Spain", "name": "Madrid"},
    "barcelona": {"lat": 41.3851, "lng": 2.1734, "country": "Spain", "name": "Barcelona"},
    "spain": {"lat": 40.4637, "lng": -3.7492, "country": "Spain", "name": "Spain"},
    "moscow": {"lat": 55.7558, "lng": 37.6173, "country": "Russia", "name": "Moscow"},
    "russia": {"lat": 61.5240, "lng": 105.3188, "country": "Russia", "name": "Russia"},
    "zurich": {"lat": 47.3769, "lng": 8.5417, "country": "Switzerland", "name": "Zurich"},
    "switzerland": {"lat": 46.8182, "lng": 8.2275, "country": "Switzerland", "name": "Switzerland"},
    "vienna": {"lat": 48.2082, "lng": 16.3738, "country": "Austria", "name": "Vienna"},
    "austria": {"lat": 47.5162, "lng": 14.5501, "country": "Austria", "name": "Austria"},
    "stockholm": {"lat": 59.3293, "lng": 18.0686, "country": "Sweden", "name": "Stockholm"},
    "sweden": {"lat": 60.1282, "lng": 18.6435, "country": "Sweden", "name": "Sweden"},
    "dublin": {"lat": 53.3498, "lng": -6.2603, "country": "Ireland", "name": "Dublin"},
    "ireland": {"lat": 53.1424, "lng": -7.6921, "country": "Ireland", "name": "Ireland"},
    
    # China
    "beijing": {"lat": 39.9042, "lng": 116.4074, "country": "China", "name": "Beijing"},
    "shanghai": {"lat": 31.2304, "lng": 121.4737, "country": "China", "name": "Shanghai"},
    "shenzhen": {"lat": 22.5431, "lng": 114.0579, "country": "China", "name": "Shenzhen"},
    "guangzhou": {"lat": 23.1291, "lng": 113.2644, "country": "China", "name": "Guangzhou"},
    "china": {"lat": 35.8617, "lng": 104.1954, "country": "China", "name": "China"},
    
    # Africa
    "johannesburg": {"lat": -26.2041, "lng": 28.0473, "country": "South Africa", "name": "Johannesburg"},
    "cape town": {"lat": -33.9249, "lng": 18.4241, "country": "South Africa", "name": "Cape Town"},
    "south africa": {"lat": -30.5595, "lng": 22.9375, "country": "South Africa", "name": "South Africa"},
    "cairo": {"lat": 30.0444, "lng": 31.2357, "country": "Egypt", "name": "Cairo"},
    "egypt": {"lat": 26.8206, "lng": 30.8025, "country": "Egypt", "name": "Egypt"},
    "lagos": {"lat": 6.5244, "lng": 3.3792, "country": "Nigeria", "name": "Lagos"},
    "nigeria": {"lat": 9.0820, "lng": 8.6753, "country": "Nigeria", "name": "Nigeria"},
    "nairobi": {"lat": -1.2921, "lng": 36.8219, "country": "Kenya", "name": "Nairobi"},
    "kenya": {"lat": -0.0236, "lng": 37.9062, "country": "Kenya", "name": "Kenya"},
    
    # South America
    "sao paulo": {"lat": -23.5505, "lng": -46.6333, "country": "Brazil", "name": "São Paulo"},
    "rio de janeiro": {"lat": -22.9068, "lng": -43.1729, "country": "Brazil", "name": "Rio de Janeiro"},
    "brazil": {"lat": -14.2350, "lng": -51.9253, "country": "Brazil", "name": "Brazil"},
    "buenos aires": {"lat": -34.6037, "lng": -58.3816, "country": "Argentina", "name": "Buenos Aires"},
    "argentina": {"lat": -38.4161, "lng": -63.6167, "country": "Argentina", "name": "Argentina"},
    "mexico city": {"lat": 19.4326, "lng": -99.1332, "country": "Mexico", "name": "Mexico City"},
    "mexico": {"lat": 23.6345, "lng": -102.5528, "country": "Mexico", "name": "Mexico"},
}

def get_city_from_query(query: str) -> Optional[Dict]:
    """Extract city coordinates from search query - supports 150+ cities worldwide"""
    query_lower = query.lower()
    
    # Sort by length (longer names first) to match "new york" before "york"
    sorted_cities = sorted(CITY_COORDINATES.keys(), key=len, reverse=True)
    
    for city_name in sorted_cities:
        if city_name in query_lower:
            coords = CITY_COORDINATES[city_name]
            return {
                "name": coords.get("name", city_name.title()),
                "lat": coords["lat"],
                "lng": coords["lng"],
                "country": coords["country"]
            }
    return None

def get_store_type_from_query(query: str) -> str:
    """Determine store type from product query"""
    query_lower = query.lower()
    
    if any(word in query_lower for word in ["phone", "mobile", "iphone", "samsung", "xiaomi", "oneplus", "vivo", "oppo", "realme"]):
        return "mobile phone store"
    elif any(word in query_lower for word in ["laptop", "computer", "pc", "desktop", "macbook"]):
        return "computer store"
    elif any(word in query_lower for word in ["tv", "television", "led", "oled"]):
        return "electronics store"
    elif any(word in query_lower for word in ["camera", "dslr", "mirrorless"]):
        return "camera store"
    elif any(word in query_lower for word in ["headphone", "earphone", "earbuds", "speaker", "audio"]):
        return "electronics store"
    elif any(word in query_lower for word in ["watch", "smartwatch"]):
        return "watch store"
    elif any(word in query_lower for word in ["tile", "bathroom", "kitchen", "flooring", "ceramic"]):
        return "tile store"
    elif any(word in query_lower for word in ["furniture", "sofa", "bed", "table", "chair"]):
        return "furniture store"
    elif any(word in query_lower for word in ["cloth", "shirt", "pant", "dress", "fashion"]):
        return "clothing store"
    elif any(word in query_lower for word in ["shoe", "footwear", "sneaker", "sandal"]):
        return "shoe store"
    else:
        return "store"

async def search_local_stores_with_places_api(query: str, city: str = None, max_results: int = 30) -> List[Dict]:
    """
    Search for LOCAL businesses using OpenStreetMap Overpass API (FREE, No API Key Required).
    Returns stores, factory outlets, manufacturing units, wholesalers, and retail shops.
    """
    try:
        # Get city coordinates from query or provided city
        city_info = None
        if city:
            city_info = CITY_COORDINATES.get(city.lower())
        if not city_info:
            city_info = get_city_from_query(query)
        
        if not city_info:
            logger.info("No city found in query, skipping local store search")
            return []
        
        # Determine OSM categories from query
        shop_categories = get_osm_categories_extended(query)
        city_name = city_info.get("name", city).title()
        
        # Get OSM area name from city_info (already properly formatted)
        osm_area = city_info.get("name", city_name)
        
        # Get shop category and product keywords for dynamic search
        shop_regex = shop_categories.get('shop', '')
        product_keywords = shop_categories.get('keywords', [])
        
        logger.info(f"OpenStreetMap search: shop_regex={shop_regex}, keywords={product_keywords} in area '{osm_area}'")
        
        if not shop_regex and not product_keywords:
            logger.info(f"No search criteria found for query, skipping local store search")
            return []
        
        # Build dynamic Overpass query - search by shop type AND/OR name containing keywords
        # IMPORTANT: Only search for specific shop types OR shops with product keywords in name
        # DO NOT return all shops when no category is found
        
        query_parts = []
        
        # Only add shop type search if we have specific shop categories
        if shop_regex:
            query_parts.append(f'node["shop"~"{shop_regex}"](area.searchArea);')
            query_parts.append(f'way["shop"~"{shop_regex}"](area.searchArea);')
        
        # Search for stores with product keywords in their name (must have shop tag)
        # Use word boundary matching to avoid partial matches (e.g., "fan" matching "fancy")
        if product_keywords:
            # Build regex with word boundaries for each keyword
            # OpenStreetMap regex uses POSIX ERE, word boundary is \b but may not work
            # Instead, use pattern like: (^|[^a-z])keyword([^a-z]|$) for word boundary simulation
            # Simpler approach: search for exact word match with space/start/end boundaries
            keyword_patterns = []
            for kw in product_keywords:
                # Match keyword at start, end, or surrounded by non-letter characters
                # This prevents "fan" from matching "fancy" 
                keyword_patterns.append(f'(^|[^a-zA-Z]){kw}([^a-zA-Z]|$)')
            keyword_regex = '|'.join(product_keywords)  # Fallback to simple match
            
            # For better results, search for each keyword separately and combine
            # This ensures we find stores with actual product names
            query_parts.append(f'node["name"~"{keyword_regex}",i]["shop"](area.searchArea);')
            query_parts.append(f'way["name"~"{keyword_regex}",i]["shop"](area.searchArea);')
        
        # If no query parts, return empty
        if not query_parts:
            logger.info("No valid query parts generated, skipping local store search")
            return []
        
        overpass_query = f'''[out:json][timeout:25];
area["name"="{osm_area}"]->.searchArea;
(
  {chr(10).join(query_parts)}
);
out body {max_results * 2};'''
        
        logger.info(f"Overpass query for area '{osm_area}' - dynamic search")
        
        # List of Overpass API servers for fallback
        overpass_servers = [
            "https://overpass-api.de/api/interpreter",
            "https://overpass.kumi.systems/api/interpreter",
            "https://maps.mail.ru/osm/tools/overpass/api/interpreter"
        ]
        
        async with httpx.AsyncClient() as client:
            response = None
            for server in overpass_servers:
                try:
                    response = await client.post(
                        server,
                        data=overpass_query,
                        timeout=30.0
                    )
                    if response.status_code == 200:
                        logger.info(f"Overpass API success from {server}")
                        break
                    else:
                        logger.warning(f"Overpass API {server} returned {response.status_code}, trying next...")
                except Exception as e:
                    logger.warning(f"Overpass API {server} failed: {str(e)}, trying next...")
                    continue
            
            if not response or response.status_code != 200:
                logger.error(f"All Overpass API servers failed")
                return []
            
            data = response.json()
            
            local_stores = []
            seen_names = set()  # Avoid duplicates
            relevant_stores = []  # Stores matching keywords (priority)
            other_stores = []  # Other matching stores
            
            # Keywords to ALWAYS EXCLUDE - these are never relevant
            always_exclude = [
                "fruit market", "vegetable market", "flower market", "fish market", 
                "meat", "chicken", "mutton", "bus stop", "bus stand", "railway",
                "restaurant", "hotel", "cafe", "food court", "biryani", "dosa",
                "temple", "church", "mosque", "school", "college", "hospital",
                "bank", "atm", "petrol pump", "gas station", "parking"
            ]
            
            for element in data.get("elements", []):
                tags = element.get("tags", {})
                name = tags.get("name")
                
                # Skip elements without names or duplicates
                if not name or name in seen_names:
                    continue
                
                name_lower = name.lower()
                shop_type = tags.get("shop", "").lower()
                
                # Skip stores in always-exclude list
                should_skip = False
                for exclude in always_exclude:
                    if exclude in name_lower:
                        should_skip = True
                        break
                
                if should_skip:
                    continue
                    
                seen_names.add(name)
                
                # Check if store name contains any product keyword as a WHOLE WORD (high relevance)
                # Use word boundary matching to avoid false positives like "fan" matching "fancy"
                is_relevant = False
                for keyword in product_keywords:
                    # Check if keyword appears as a whole word in the name or shop type
                    pattern = r'\b' + re.escape(keyword) + r'\b'
                    if re.search(pattern, name_lower) or re.search(pattern, shop_type):
                        is_relevant = True
                        break
                
                # SKIP stores that don't have product keywords as whole words
                # This is the key fix - only include stores with actual product/brand name matches
                if not is_relevant:
                    continue
                
                # Build address from OSM tags
                address_parts = []
                if tags.get("addr:housenumber"):
                    address_parts.append(tags.get("addr:housenumber"))
                if tags.get("addr:street"):
                    address_parts.append(tags.get("addr:street"))
                if tags.get("addr:suburb") or tags.get("addr:neighbourhood"):
                    address_parts.append(tags.get("addr:suburb") or tags.get("addr:neighbourhood"))
                if tags.get("addr:city"):
                    address_parts.append(tags.get("addr:city"))
                
                full_address = ", ".join(address_parts) if address_parts else f"Near {osm_area}"
                
                # Get coordinates
                lat = element.get("lat")
                lon = element.get("lon")
                
                # Calculate approximate distance from city center
                distance_meters = None
                if lat and lon:
                    import math
                    city_lat = city_info.get("lat", 0)
                    city_lon = city_info.get("lng", 0)
                    # Haversine formula approximation
                    dlat = math.radians(lat - city_lat)
                    dlon = math.radians(lon - city_lon)
                    a = math.sin(dlat/2)**2 + math.cos(math.radians(city_lat)) * math.cos(math.radians(lat)) * math.sin(dlon/2)**2
                    c = 2 * math.asin(math.sqrt(a))
                    distance_meters = int(6371000 * c)  # Earth radius in meters
                
                # Determine business type from OSM tags
                business_type = "Retail Shop"
                business_icon = "🏪"
                
                if tags.get("industrial") or tags.get("man_made") in ["works", "factory"]:
                    if "warehouse" in str(tags.get("industrial", "")):
                        business_type = "Warehouse / Distribution"
                        business_icon = "🏭"
                    else:
                        business_type = "Factory / Manufacturing Unit"
                        business_icon = "🏭"
                elif tags.get("landuse") == "industrial":
                    business_type = "Industrial Zone"
                    business_icon = "🏭"
                elif tags.get("shop") == "wholesale":
                    business_type = "Wholesale Supplier"
                    business_icon = "📦"
                elif tags.get("trade"):
                    business_type = "Trade / B2B"
                    business_icon = "🤝"
                elif tags.get("office"):
                    business_type = "Corporate Office / Showroom"
                    business_icon = "🏢"
                elif tags.get("craft"):
                    business_type = "Manufacturing Workshop"
                    business_icon = "🔧"
                elif tags.get("shop") == "mall" or "outlet" in name.lower():
                    business_type = "Factory Outlet"
                    business_icon = "🏬"
                elif tags.get("brand"):
                    business_type = "Brand Authorized Store"
                    business_icon = "✅"
                else:
                    business_type = "Retail Shop"
                    business_icon = "🏪"
                
                # Build categories list
                categories = []
                if tags.get("shop"):
                    categories.append(tags.get("shop"))
                if tags.get("brand"):
                    categories.append(tags.get("brand"))
                if tags.get("industrial"):
                    categories.append(tags.get("industrial"))
                if tags.get("trade"):
                    categories.append(tags.get("trade"))
                if tags.get("craft"):
                    categories.append(tags.get("craft"))
                if not categories:
                    categories = [business_type]
                
                store = {
                    "place_id": str(element.get("id", "")),
                    "name": name,
                    "address": full_address,
                    "locality": osm_area,
                    "region": tags.get("addr:state", city_info.get("country", "").title()),
                    "postcode": tags.get("addr:postcode", ""),
                    "phone": tags.get("phone") or tags.get("contact:phone", ""),
                    "email": tags.get("email") or tags.get("contact:email", ""),
                    "website": tags.get("website") or tags.get("contact:website", ""),
                    "rating": None,  # OSM doesn't have ratings
                    "review_count": 0,
                    "is_open_now": None,  # Would need opening_hours parsing
                    "opening_hours": [tags.get("opening_hours")] if tags.get("opening_hours") else [],
                    "categories": categories,
                    "business_type": business_type,
                    "business_icon": business_icon,
                    "distance_meters": distance_meters,
                    "price_level": None,
                    "city": osm_area,
                    "country": city_info.get("country", ""),
                    "is_local_store": True,
                    "is_relevant": True,  # All stores here have product keyword matches
                    "data_source": "OpenStreetMap",
                    "google_maps_url": f"https://www.google.com/maps/search/?api=1&query={name.replace(' ', '+')}+{osm_area.replace(' ', '+')}" if lat and lon else f"https://www.google.com/maps/search/?api=1&query={name.replace(' ', '+')}",
                    "coordinates": {"lat": lat, "lon": lon} if lat and lon else None
                }
                
                # Add to relevant stores list (we already filtered out non-relevant ones)
                relevant_stores.append(store)
            
            # All stores are now relevant - no need to combine with other_stores
            local_stores = relevant_stores
            
            # Limit to max_results
            local_stores = local_stores[:max_results]
            
            logger.info(f"OpenStreetMap returned {len(local_stores)} local stores in {osm_area} (all matched product keywords)")
            return local_stores
            
    except Exception as e:
        logger.error(f"OpenStreetMap Overpass API error: {str(e)}")
        return []

def get_osm_categories_extended(query: str) -> Dict[str, str]:
    """
    FULLY DYNAMIC - extracts product/brand keywords from query.
    NO default shop categories - searches ONLY by product name in store names.
    Works exactly like online search - finds stores with actual product/brand names.
    """
    query_lower = query.lower()
    
    # Step 1: Detect the city from the query FIRST
    detected_city_info = get_city_from_query(query)
    detected_city_name = detected_city_info.get("name", "").lower() if detected_city_info else ""
    
    # Step 2: Build comprehensive list of location words to exclude
    location_words_to_exclude = set()
    if detected_city_name:
        location_words_to_exclude.add(detected_city_name)
        for word in detected_city_name.split():
            location_words_to_exclude.add(word)
    
    # Add all city key variations
    for city_key, city_data in CITY_COORDINATES.items():
        city_osm_name = city_data.get("name", "").lower()
        if detected_city_name and (detected_city_name in city_osm_name or city_osm_name in detected_city_name):
            location_words_to_exclude.add(city_key)
            for word in city_key.split():
                location_words_to_exclude.add(word)
    
    # Step 3: Define comprehensive stop words - DO NOT search for these
    stop_words = {
        'price', 'prices', 'in', 'at', 'the', 'a', 'an', 'for', 'of', 'to', 'and', 'or', 
        'best', 'cheap', 'buy', 'shop', 'store', 'stores', 'near', 'me', 'my', 'online', 
        'cost', 'rate', 'rates', 'where', 'find', 'get', 'local', 'nearby', 'around',
        'market', 'markets', 'dealer', 'dealers', 'seller', 'sellers', 'vendor', 'vendors',
        'showroom', 'showrooms', 'outlet', 'outlets', 'wholesale', 'retail'
    }
    
    # Step 4: Extract product/brand keywords
    words = query_lower.split()
    product_keywords = []
    for w in words:
        if w in stop_words or len(w) <= 2:
            continue
        if w in location_words_to_exclude:
            continue
        if w in CITY_COORDINATES:
            continue
        product_keywords.append(w)
    
    # Return ONLY keywords - NO default shop categories
    # The Overpass query will search for stores with these keywords in their NAME
    if product_keywords:
        return {"shop": "", "keywords": product_keywords}
    
    return {"shop": "", "keywords": []}

def get_osm_shop_category(query: str) -> str:
    """Map product query to OpenStreetMap shop tag"""
    query_lower = query.lower()
    
    if any(word in query_lower for word in ["phone", "mobile", "iphone", "samsung", "xiaomi", "oneplus", "vivo", "oppo", "realme"]):
        return "mobile_phone"
    elif any(word in query_lower for word in ["laptop", "computer", "pc", "desktop", "macbook"]):
        return "computer"
    elif any(word in query_lower for word in ["tv", "television", "led", "oled", "electronics", "camera", "headphone", "earphone", "speaker", "audio"]):
        return "electronics"
    elif any(word in query_lower for word in ["watch", "smartwatch"]):
        return "watches"
    elif any(word in query_lower for word in ["tile", "bathroom", "kitchen", "flooring", "ceramic"]):
        return "doityourself"
    elif any(word in query_lower for word in ["furniture", "sofa", "bed", "table", "chair"]):
        return "furniture"
    elif any(word in query_lower for word in ["cloth", "shirt", "pant", "dress", "fashion"]):
        return "clothes"
    elif any(word in query_lower for word in ["shoe", "footwear", "sneaker", "sandal"]):
        return "shoes"
    elif any(word in query_lower for word in ["jewel", "gold", "diamond", "ring", "necklace"]):
        return "jewelry"
    elif any(word in query_lower for word in ["grocery", "food", "vegetable", "fruit"]):
        return "supermarket"
    else:
        return "electronics"  # Default to electronics for general product searches

def is_product_relevant(product_name: str, search_query: str, query_info: Dict) -> bool:
    """
    Check if a product is relevant to the search query.
    Filters out unrelated products to ensure only matching items are shown.
    Uses lenient matching to avoid removing valid products.
    """
    import re
    
    product_lower = product_name.lower()
    query_lower = search_query.lower()
    
    # If brand is specified in search, product must contain that brand
    if query_info.get('brand'):
        brand_lower = query_info['brand'].lower()
        if brand_lower not in product_lower:
            # Allow some Apple/iPhone flexibility
            if brand_lower == 'apple' and 'iphone' not in product_lower and 'ipad' not in product_lower and 'macbook' not in product_lower:
                return False
            elif brand_lower == 'iphone' and 'iphone' not in product_lower:
                return False
            elif brand_lower not in ['apple', 'iphone'] and brand_lower not in product_lower:
                return False
    
    # Extract key search terms (filter out common words)
    stop_words = {'the', 'in', 'price', 'india', 'under', 'above', 'below', 'buy', 'online', 'best', 'new', 'latest', 'top', 'cheap', 'offer', 'for', 'with', 'and', 'or'}
    search_terms = [w for w in re.findall(r'\w+', query_lower) if len(w) > 2 and w not in stop_words]
    
    # LENIENT MATCHING: Product should contain at least one key search term OR share category
    # This ensures we don't remove valid products from category-specific sources
    if search_terms:
        # Check for direct term matches
        has_match = any(term in product_lower for term in search_terms if len(term) > 2)
        
        # If no direct match, check for partial/fuzzy matches (more lenient)
        if not has_match:
            # Check if product contains any substring of search terms or vice versa
            for term in search_terms:
                if len(term) > 3:
                    # Check for partial matches (e.g., 'shoe' in 'shoes')\n                    if any(term in word or word in term for word in product_lower.split()):
                        has_match = True
                        break
        
        # If still no match but product has valid brand/category indicators, allow it
        # This prevents removing products from category-specific sites
        if not has_match:
            category_indicators = ['men', 'women', 'kids', 'phone', 'mobile', 'laptop', 'table', 'chair', 'shirt', 'shoes']
            if any(indicator in product_lower for indicator in category_indicators):
                # If product is from a category-specific source, be more lenient
                return True
            return False
    
    return True

def parse_search_query(product_name: str) -> Dict[str, Any]:
    """
    Parse search query to extract brand, product type, price constraints, and generate search variations.
    Returns: {"primary_query": str, "variations": [str], "max_price": float, "brand": str, "product_type": str}
    """
    import re
    
    query_lower = product_name.lower()
    result = {
        "primary_query": product_name,
        "variations": [],
        "max_price": None,
        "min_price": None,
        "brand": None,
        "product_type": None
    }
    
    # Extract price constraints
    price_patterns = [
        (r'under[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'max'),
        (r'below[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'max'),
        (r'less than[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'max'),
        (r'upto[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'max'),
        (r'up to[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'max'),
        (r'above[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'min'),
        (r'over[\s]+(?:rs\.?|₹)?\s*([\d,]+)', 'min'),
    ]
    
    for pattern, price_type in price_patterns:
        match = re.search(pattern, query_lower)
        if match:
            price_str = match.group(1).replace(',', '')
            try:
                price = float(price_str)
                if price_type == 'max':
                    result['max_price'] = price
                else:
                    result['min_price'] = price
                # Remove price phrase from query
                product_name = re.sub(pattern, '', product_name, flags=re.IGNORECASE)
            except ValueError:
                pass
    
    # Extract common brands
    brands = [
        'dell', 'hp', 'lenovo', 'asus', 'acer', 'apple', 'samsung', 'lg', 'sony',
        'mi', 'xiaomi', 'realme', 'oppo', 'vivo', 'oneplus', 'nokia', 'motorola',
        'nike', 'adidas', 'puma', 'reebok', 'bosch', 'philips', 'panasonic',
        'godrej', 'whirlpool', 'haier', 'ifb', 'voltas', 'carrier', 'hitachi'
    ]
    
    for brand in brands:
        if re.search(r'\b' + brand + r'\b', query_lower):
            result['brand'] = brand.title()
            break
    
    # Extract product type
    product_types = [
        'laptop', 'laptop', 'mobile', 'phone', 'smartphone', 'tablet', 'watch',
        'shoes', 'shirt', 'jeans', 'dress', 'bag', 'refrigerator', 'fridge',
        'washing machine', 'ac', 'air conditioner', 'tv', 'television',
        'microwave', 'oven', 'mixer', 'grinder', 'camera', 'headphone', 'earphone'
    ]
    
    for ptype in product_types:
        if ptype in query_lower:
            result['product_type'] = ptype
            break
    
    # Clean the query
    cleaned = product_name.strip()
    cleaned = re.sub(r'\s+', ' ', cleaned)
    result['primary_query'] = cleaned
    
    # Generate search variations for comprehensive results
    variations = [cleaned]
    
    # Add brand + product type variation
    if result['brand'] and result['product_type']:
        variations.append(f"{result['brand']} {result['product_type']}")
    
    # Add just product type for broader results
    if result['product_type']:
        variations.append(result['product_type'])
    
    # Add brand alone if product type is generic
    if result['brand'] and result['product_type'] and len(result['product_type'].split()) == 1:
        variations.append(result['brand'])
    
    result['variations'] = list(dict.fromkeys(variations))  # Remove duplicates
    return result

def simplify_product_query(product_name: str) -> str:
    """
    Simplify technical/long product names to get better search results.
    Extracts key product terms and removes codes/abbreviations.
    """
    import re
    
    # Remove common patterns that don't help search
    # Remove codes like "BWP", "PLY", model numbers, etc.
    simplified = product_name
    
    # Remove text in parentheses or brackets
    simplified = re.sub(r'\([^)]*\)', '', simplified)
    simplified = re.sub(r'\[[^\]]*\]', '', simplified)
    
    # Remove common abbreviations and codes
    codes_to_remove = [
        r'\bBWP\b', r'\bPLY\b', r'\bMDF\b', r'\bHDF\b', r'\bSS\b',
        r'\bL\b$', r'\bM\b$', r'\bS\b$', r'\bXL\b$',  # Size codes at end
        r'\b[A-Z]{2,5}\d+\b',  # Model numbers like ABC123
        r'\b\d+[A-Z]+\b',  # Numbers followed by letters
        r'\bSIN\b', r'\bDOCLE\b',  # Specific codes
    ]
    for code in codes_to_remove:
        simplified = re.sub(code, '', simplified, flags=re.IGNORECASE)
    
    # Remove special characters but keep spaces
    simplified = re.sub(r'[^\w\s]', ' ', simplified)
    
    # Remove extra whitespace
    simplified = ' '.join(simplified.split())
    
    # If too long, take first 5-6 meaningful words
    words = simplified.split()
    if len(words) > 6:
        # Keep key product words
        important_words = []
        for word in words:
            if len(word) > 2 and word.upper() not in ['THE', 'AND', 'FOR', 'WITH']:
                important_words.append(word)
                if len(important_words) >= 5:
                    break
        simplified = ' '.join(important_words)
    
    # Add "price" to help get shopping results
    if simplified and 'price' not in simplified.lower():
        simplified = simplified + ' price india'
    
    return simplified.strip()

async def search_with_rapidapi(query: str, max_results: int = 20) -> List[Dict]:
    """
    Search for products using RapidAPI (multiple free APIs available).
    """
    if not RAPIDAPI_KEY or RAPIDAPI_KEY == '':
        logger.info("RapidAPI key not configured, skipping")
        return []
    
    products = []
    search_timestamp = datetime.now().isoformat()
    
    headers = {
        'X-RapidAPI-Key': RAPIDAPI_KEY,
        'X-RapidAPI-Host': 'real-time-product-search.p.rapidapi.com'
    }
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            url = "https://real-time-product-search.p.rapidapi.com/search"
            params = {'q': query, 'country': 'in', 'language': 'en', 'limit': str(max_results)}
            
            response = await client.get(url, headers=headers, params=params)
            
            if response.status_code == 200:
                data = response.json()
                for item in data.get('data', [])[:max_results]:
                    try:
                        title = item.get('product_title', item.get('title', ''))
                        price_str = item.get('offer_price', item.get('price', ''))
                        link = item.get('product_url', item.get('url', ''))
                        source = item.get('source', 'Unknown')
                        
                        if price_str:
                            price_clean = str(price_str).replace(',', '').replace('₹', '').replace('Rs', '').strip()
                            price = float(price_clean)
                            if 100 <= price <= 10000000 and title and link:
                                products.append({
                                    'name': title[:150],
                                    'price': price,
                                    'currency_symbol': '₹',
                                    'currency_code': 'INR',
                                    'source': source,
                                    'source_url': link,
                                    'description': title,
                                    'search_engine': 'RapidAPI',
                                    'timestamp': search_timestamp
                                })
                    except Exception:
                        pass
                logger.info(f"✅ RapidAPI found {len(products)} products")
    except Exception as e:
        logger.warning(f"RapidAPI search failed: {str(e)[:100]}")
    
    return products

async def search_with_serpapi_enhanced(query: str, original_item: str, country: str = "india", max_results: int = 30) -> List[Dict]:
    """
    Enhanced search with multiple fallback sources and query variations - NO fake data!
    Performs comprehensive search showing multiple related products like e-commerce platforms.
    1. Parse query to extract brand, product type, price constraints
    2. Search with multiple variations (broad + specific)
    3. Direct web scraping (Amazon, Flipkart, etc.)
    4. RapidAPI (if configured)
    5. SerpAPI (if configured)
    6. Filter by price constraints if specified
    """
    all_results = []
    seen_urls = set()
    
    # Parse the query to extract components and generate variations
    query_info = parse_search_query(query)
    logger.info(f"🔍 Parsed query: Brand={query_info['brand']}, Type={query_info['product_type']}, MaxPrice={query_info['max_price']}")
    logger.info(f"🔍 Search variations: {query_info['variations']}")
    
    # Search with each variation to get comprehensive results
    for search_query in query_info['variations'][:3]:  # Limit to top 3 variations
        simplified_query = simplify_product_query(search_query)
        logger.info(f"🌐 Searching for variation: '{simplified_query}'")
        
        variation_results = []
        
        # Strategy 1: FREE - Direct web scraping
        try:
            web_results = await search_real_web_prices(simplified_query, max_results)
            if web_results:
                logger.info(f"  ✅ Found {len(web_results)} results from web scraping")
                variation_results.extend(web_results)
        except Exception as e:
            logger.warning(f"  ⚠️ Web scraping failed: {e}")
        
        # Strategy 2: RapidAPI (if configured)
        try:
            if RAPIDAPI_KEY and RAPIDAPI_KEY != '':
                api_results = await search_with_rapidapi(simplified_query, max_results // 2)
                if api_results:
                    logger.info(f"  ✅ Found {len(api_results)} results from RapidAPI")
                    variation_results.extend(api_results)
        except Exception as e:
            logger.debug(f"  RapidAPI unavailable: {str(e)}")
        
        # Strategy 3: SerpAPI/Google Shopping (if configured)
        try:
            if SERPAPI_API_KEY and SERPAPI_API_KEY != 'your_serpapi_key_here':
                api_results = await search_with_serpapi(simplified_query, country, max_results // 2)
                if api_results:
                    logger.info(f"  ✅ Found {len(api_results)} results from SerpAPI")
                    variation_results.extend(api_results)
        except Exception as e:
            logger.debug(f"  SerpAPI unavailable: {str(e)}")
        
        # Deduplicate by URL
        for result in variation_results:
            url = result.get('source_url', '')
            # Create a deduplication key from URL (without query params)
            url_key = url.split('?')[0].lower() if url else ''
            if url_key and url_key not in seen_urls:
                seen_urls.add(url_key)
                all_results.append(result)
        
        # If we have good results for first variation, we can be less aggressive with others
        if len(all_results) >= 15 and search_query == query_info['variations'][0]:
            logger.info(f"✅ Got {len(all_results)} results from primary query, checking other variations")
    
    # Apply price filtering if specified
    if query_info['max_price']:
        original_count = len(all_results)
        all_results = [r for r in all_results if r.get('price', float('inf')) <= query_info['max_price']]
        if original_count > len(all_results):
            logger.info(f"💰 Filtered {original_count - len(all_results)} products above ₹{query_info['max_price']}")
    
    if query_info['min_price']:
        original_count = len(all_results)
        all_results = [r for r in all_results if r.get('price', 0) >= query_info['min_price']]
        if original_count > len(all_results):
            logger.info(f"💰 Filtered {original_count - len(all_results)} products below ₹{query_info['min_price']}")
    
    # Apply relevance filtering to remove unrelated products
    if all_results:
        original_count = len(all_results)
        all_results = [r for r in all_results if is_product_relevant(r.get('name', ''), query, query_info)]
        if original_count > len(all_results):
            logger.info(f"🎯 Filtered {original_count - len(all_results)} unrelated products (keeping only relevant matches)")
    
    if all_results:
        logger.info(f"✅ Returning {len(all_results)} comprehensive results (deduplicated & relevant)")
        return all_results[:max_results]  # Limit to max_results
    
    # No live prices found - return empty to indicate manual verification needed
    logger.warning(f"⚠️ All live price sources failed for: {query} - Manual verification required")
    return []

def extract_product_type(item_name: str) -> str:
    """
    Extract the main product type from a detailed item name.
    """
    item_lower = item_name.lower()
    
    # Common product type mappings - use word boundaries
    product_types = [
        ('sink', 'kitchen sink'),
        ('refrigerator', 'refrigerator'),
        ('fridge', 'refrigerator'),
        ('plywood', 'plywood'),
        (' ply ', 'plywood'),  # space around to avoid matching "supply"
        ('counter top', 'kitchen counter top'),
        ('countertop', 'kitchen counter top'),
        ('quartz stone', 'quartz stone slab'),
        ('quartz', 'quartz stone'),
        ('dado', 'wall dado tiles'),
        ('cabinet', 'kitchen cabinet'),
        ('chimney', 'kitchen chimney'),
        ('hob', 'kitchen hob'),
        ('mixer', 'mixer grinder'),
        ('tap', 'kitchen tap'),
        ('faucet', 'kitchen faucet'),
        ('granite', 'granite slab'),
        ('marble', 'marble slab'),
        ('tile', 'tiles'),
        ('laminate', 'laminate sheet'),
        ('hardware', 'kitchen hardware'),
        ('drawer', 'kitchen drawer'),
        ('shutter', 'kitchen shutter'),
        ('modular', 'modular kitchen'),
    ]
    
    # Brand mappings
    brand_products = {
        'bosch': 'bosch appliance',
        'carysil': 'carysil sink',
        'franke': 'franke sink',
        'lg': 'lg appliance',
        'samsung': 'samsung appliance',
        'whirlpool': 'whirlpool appliance',
        'godrej': 'godrej appliance',
        'haier': 'haier appliance',
        'ifb': 'ifb appliance',
        'siemens': 'siemens appliance',
    }
    
    # Check for brand first
    for brand, default_product in brand_products.items():
        if brand in item_lower:
            # Find what product it is
            for keyword, product_type in product_types:
                if keyword in item_lower:
                    return f"{brand} {product_type}"
            return default_product
    
    # Check product types
    for keyword, product_type in product_types:
        if keyword in item_lower:
            return product_type
    
    # Fallback: return cleaned first few words
    import re
    cleaned = re.sub(r'[^\w\s]', ' ', item_name)
    words = cleaned.split()[:4]
    return ' '.join(words)

# ================== CATEGORY DETECTION FOR INTELLIGENT SOURCE SELECTION ==================
def detect_product_category(query: str) -> str:
    """
    Intelligently detect product category from search query to trigger relevant websites.
    Uses AI-powered keyword analysis for accurate categorization.
    
    Categories:
    - fashion: Clothing, shoes, accessories, fashion items
    - electronics: Mobiles, laptops, TVs, cameras, electronics
    - furniture: Furniture, home decor, interior items
    - general: Other products
    """
    query_lower = query.lower()
    
    # Fashion keywords - clothing, footwear, accessories
    fashion_keywords = [
        # Clothing
        'shirt', 'tshirt', 't-shirt', 'pant', 'jeans', 'trouser', 'dress', 'kurti', 'saree', 'sari',
        'lehenga', 'salwar', 'kurta', 'sherwani', 'blazer', 'jacket', 'coat', 'sweater', 'hoodie',
        'shorts', 'skirt', 'top', 'blouse', 'suit', 'tracksuit', 'jogger', 'nightwear', 'innerwear',
        'lingerie', 'bra', 'underwear', 'socks', 'ethnic wear', 'western wear', 'party wear',
        # Footwear
        'shoes', 'shoe', 'sneaker', 'sandal', 'slipper', 'boot', 'heel', 'footwear', 'flip-flop',
        'loafer', 'formal shoes', 'casual shoes', 'sports shoes', 'running shoes',
        # Accessories
        'watch', 'bag', 'handbag', 'purse', 'wallet', 'belt', 'tie', 'bow tie', 'scarf', 'stole',
        'sunglasses', 'glasses', 'cap', 'hat', 'jewellery', 'jewelry', 'necklace', 'earring',
        'bracelet', 'ring', 'fashion', 'style', 'apparel', 'garment', 'clothing', 'wear'
    ]
    
    # Electronics keywords - gadgets, appliances, tech
    electronics_keywords = [
        # Mobile & Computing
        'mobile', 'phone', 'smartphone', 'iphone', 'samsung', 'oneplus', 'xiaomi', 'redmi', 'realme',
        'oppo', 'vivo', 'motorola', 'nokia', 'laptop', 'notebook', 'macbook', 'dell', 'hp', 'lenovo',
        'asus', 'acer', 'tablet', 'ipad', 'computer', 'desktop', 'pc', 'processor', 'ram', 'ssd',
        # TV & Audio
        'tv', 'television', 'smart tv', 'led tv', 'oled', 'qled', 'monitor', 'display', 'screen',
        'headphone', 'earphone', 'earbud', 'speaker', 'soundbar', 'home theatre', 'audio',
        # Cameras & Accessories
        'camera', 'dslr', 'gopro', 'webcam', 'lens', 'tripod', 'gimbal',
        # Home Appliances
        'refrigerator', 'fridge', 'washing machine', 'ac', 'air conditioner', 'microwave', 'oven',
        'dishwasher', 'vacuum cleaner', 'air purifier', 'water purifier', 'geyser', 'iron', 'fan',
        'cooler', 'mixer', 'grinder', 'juicer', 'toaster', 'kettle', 'induction', 'chimney',
        # Gadgets
        'smartwatch', 'fitness band', 'power bank', 'charger', 'bluetooth', 'wifi', 'router',
        'keyboard', 'mouse', 'gaming', 'console', 'playstation', 'xbox', 'electronic', 'gadget',
        'appliance', 'device'
    ]
    
    # Furniture & Interior keywords
    furniture_keywords = [
        # Furniture
        'sofa', 'couch', 'bed', 'mattress', 'table', 'chair', 'desk', 'cabinet', 'wardrobe',
        'cupboard', 'almirah', 'dresser', 'drawer', 'bookshelf', 'rack', 'stand', 'stool',
        'bench', 'divan', 'recliner', 'bean bag', 'furniture', 'furnishing',
        # Decor & Interior
        'curtain', 'blind', 'cushion', 'pillow', 'carpet', 'rug', 'mat', 'wallpaper', 'painting',
        'frame', 'mirror', 'lamp', 'light', 'chandelier', 'vase', 'showpiece', 'decor', 'decoration',
        'interior', 'home decor', 'wall art', 'clock', 'planter',
        # Kitchen
        'kitchen cabinet', 'modular kitchen', 'countertop', 'sink', 'tap', 'faucet',
        # Materials
        'wood', 'wooden', 'teak', 'sheesham', 'plywood', 'mdf', 'particle board'
    ]
    
    # Check fashion category
    for keyword in fashion_keywords:
        if keyword in query_lower:
            logger.info(f"Category detected: FASHION (matched: {keyword})")
            return 'fashion'
    
    # Check electronics category
    for keyword in electronics_keywords:
        if keyword in query_lower:
            logger.info(f"Category detected: ELECTRONICS (matched: {keyword})")
            return 'electronics'
    
    # Check furniture category
    for keyword in furniture_keywords:
        if keyword in query_lower:
            logger.info(f"Category detected: FURNITURE (matched: {keyword})")
            return 'furniture'
    
    # Default to general
    logger.info(f"Category detected: GENERAL (no specific match)")
    return 'general'

def get_sources_for_category(category: str) -> dict:
    """
    Get relevant website sources based on product category.
    Returns dict with source names and their scraping functions.
    """
    # Define source priorities for each category
    sources = {
        'fashion': {
            'primary': ['amazon', 'flipkart', 'myntra', 'ajio', 'tatacliq', 'meesho'],
            'secondary': ['snapdeal', 'bing', 'duckduckgo']
        },
        'electronics': {
            'primary': ['amazon', 'flipkart', 'snapdeal', 'mysmartprice', 'croma', 'reliancedigital'],
            'secondary': ['bing', 'duckduckgo', 'google']
        },
        'furniture': {
            'primary': ['amazon', 'flipkart', 'pepperfry', 'urbanladder'],
            'secondary': ['snapdeal', 'bing', 'duckduckgo']
        },
        'general': {
            'primary': ['amazon', 'flipkart', 'snapdeal', 'mysmartprice'],
            'secondary': ['bing', 'duckduckgo', 'google']
        }
    }
    
    return sources.get(category, sources['general'])

# ================== URL VALIDATION AND CLEANING HELPERS ==================
def clean_amazon_url(url: str) -> str:
    """
    Clean Amazon URL to get direct product link with ASIN.
    Removes tracking parameters but keeps product identifier.
    """
    if not url or 'amazon' not in url.lower():
        return url
    
    try:
        # Extract ASIN from URL (Amazon Standard Identification Number)
        # Pattern: /dp/ASIN or /gp/product/ASIN or /product/ASIN
        asin_match = re.search(r'/(?:dp|gp/product|product)/([A-Z0-9]{10})', url)
        if asin_match:
            asin = asin_match.group(1)
            # Return clean product URL
            return f"https://www.amazon.in/dp/{asin}"
        
        # If no ASIN found, return original URL without query params
        parsed = urlparse(url)
        return f"{parsed.scheme}://{parsed.netloc}{parsed.path}".split('?')[0]
    except:
        return url

def clean_flipkart_url(url: str) -> str:
    """
    Clean Flipkart URL to get direct product link with PID.
    Removes tracking parameters but keeps product identifier.
    """
    if not url or 'flipkart' not in url.lower():
        return url
    
    try:
        # Flipkart URLs have format: /product-name/p/PID
        parsed = urlparse(url)
        path = parsed.path
        
        # Extract PID from path
        pid_match = re.search(r'/p/([a-zA-Z0-9]+)', path)
        if pid_match:
            # Keep the path structure but remove query params
            return f"{parsed.scheme}://{parsed.netloc}{path}".split('?')[0]
        
        # If no PID, just remove query params
        return f"{parsed.scheme}://{parsed.netloc}{path}".split('?')[0]
    except:
        return url

def clean_snapdeal_url(url: str) -> str:
    """
    Clean Snapdeal URL to get direct product link.
    Removes tracking parameters.
    """
    if not url or 'snapdeal' not in url.lower():
        return url
    
    try:
        parsed = urlparse(url)
        # Remove query parameters but keep the path
        return f"{parsed.scheme}://{parsed.netloc}{parsed.path}".split('?')[0]
    except:
        return url

def is_valid_product_url(url: str, source: str = "") -> bool:
    """
    Validate if URL is a proper product page URL.
    Returns False for search pages, category pages, or broken URLs.
    """
    if not url or len(url) < 10:
        return False
    
    try:
        parsed = urlparse(url)
        
        # Must have valid scheme and netloc
        if not parsed.scheme in ['http', 'https'] or not parsed.netloc:
            return False
        
        path = parsed.path.lower()
        
        # Amazon validation
        if 'amazon' in parsed.netloc.lower():
            # Must have /dp/ or /gp/product/ for valid product pages
            if '/dp/' in path or '/gp/product/' in path or '/product/' in path:
                # Extract ASIN and validate format (10 alphanumeric)
                asin_match = re.search(r'/(?:dp|gp/product|product)/([A-Z0-9]{10})', url, re.IGNORECASE)
                return bool(asin_match)
            # Reject search pages
            if '/s?' in url or '/s/' in path or 'search' in path:
                return False
            return False
        
        # Flipkart validation
        if 'flipkart' in parsed.netloc.lower():
            # Must have /p/ for product pages
            if '/p/' in path:
                # Should have PID after /p/
                pid_match = re.search(r'/p/[a-zA-Z0-9]{10,}', path)
                return bool(pid_match)
            # Reject search pages
            if '/search?' in url or 'search' in path:
                return False
            return False
        
        # Snapdeal validation
        if 'snapdeal' in parsed.netloc.lower():
            # Product pages have /product/ in path
            if '/product/' in path:
                return True
            # Reject search pages
            if '/search?' in url or 'keyword=' in url:
                return False
            return False
        
        # For other sources, basic validation
        # Reject obvious search/category pages
        if any(term in url.lower() for term in ['/search?', '/search/', '/category/', '/categories/', '?q=', '?keyword=']):
            return False
        
        # Accept if path is not empty (likely a product page)
        return len(parsed.path) > 5
        
    except:
        return False

# ================== REAL WEB SEARCH (No Hardcoded Prices) ==================
async def search_real_web_prices(query: str, max_results: int = 30) -> List[Dict]:
    """
    Perform real web searches across multiple sources to find actual market prices.
    NO hardcoded or static prices - all prices come from live web scraping.
    
    Multi-source strategy:
    1. Direct e-commerce sites (Amazon, Flipkart, Snapdeal)
    2. Price comparison sites (MySmartPrice, PriceDekho, ComparIndia)
    3. Search engines (Bing, Google, DuckDuckGo)
    4. Shopping aggregators
    
    Returns list of price results with source URLs and vendor names.
    """
    products = []
    search_timestamp = datetime.now().isoformat()
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-IN,en-US;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Referer': 'https://www.google.com/',
    }
    
    # Clean query for better search results
    clean_query = query.strip()
    encoded_query = quote_plus(clean_query)
    encoded_query_price = quote_plus(f"{clean_query} price india")
    
    # User agent rotation to avoid detection
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    ]
    
    headers['User-Agent'] = rand_module.choice(user_agents)
    
    # ========== CATEGORY-INTELLIGENT SOURCE SELECTION ==========
    # Detect product category to trigger relevant websites
    category = detect_product_category(query)
    source_config = get_sources_for_category(category)
    
    logger.info(f"🎯 Category: {category.upper()}")
    logger.info(f"🌐 Primary sources: {', '.join(source_config['primary'])}")
    logger.info(f"🔍 Secondary sources: {', '.join(source_config['secondary'])}")
    
    # Combine primary and secondary sources
    active_sources = source_config['primary'] + source_config['secondary']
    
    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        
        # Source 1: Enhanced Amazon India search with multiple strategies
        try:
            # Try both regular search and specific category search
            amazon_urls = [
                f"https://www.amazon.in/s?k={encoded_query}",
                f"https://www.amazon.in/s?k={encoded_query}&rh=p_36%3A1000-5000000",  # Price range filter
            ]
            
            for amazon_url in amazon_urls:
                try:
                    headers['User-Agent'] = rand_module.choice(user_agents)
                    await asyncio.sleep(rand_module.uniform(0.5, 1.5))  # Random delay
                    
                    response = await client.get(amazon_url, headers=headers)
                    
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'lxml')
                        initial_count = len(products)
                        
                        # Multiple selector strategies for Amazon
                        selectors = [
                            '[data-component-type="s-search-result"]',
                            '.s-result-item[data-asin]',
                            'div[data-asin]:not([data-asin=""])'
                        ]
                        
                        items = []
                        for selector in selectors:
                            items = soup.select(selector)
                            if items:
                                break
                        
                        for item in items[:25]:
                            try:
                                # Product title - multiple strategies
                                title = None
                                title_selectors = ['h2 a span', 'h2 span', '.a-text-normal', 'h2 a']
                                for ts in title_selectors:
                                    title_elem = item.select_one(ts)
                                    if title_elem:
                                        title = title_elem.get_text(strip=True)
                                        if title:
                                            break
                                
                                if not title:
                                    continue
                                
                                # Product link - extract and validate
                                link = ""
                                link_elem = item.select_one('h2 a, a.a-link-normal')
                                if link_elem and link_elem.get('href'):
                                    href = link_elem.get('href', '')
                                    if href.startswith('http'):
                                        link = href
                                    else:
                                        link = "https://www.amazon.in" + href
                                    
                                    # Clean and validate Amazon URL
                                    link = clean_amazon_url(link)
                                    
                                    # Skip if not a valid product URL
                                    if not is_valid_product_url(link, 'Amazon'):
                                        continue
                                
                                # Price - comprehensive extraction
                                price = 0
                                price_selectors = [
                                    '.a-price-whole',
                                    '.a-price .a-offscreen',
                                    'span.a-price-whole',
                                    '.a-color-price',
                                    'span[data-a-color="price"]'
                                ]
                                
                                for ps in price_selectors:
                                    price_elem = item.select_one(ps)
                                    if price_elem:
                                        price_text = price_elem.get_text(strip=True)
                                        extracted_price = extract_price_from_text(price_text)
                                        if extracted_price > 0:
                                            price = extracted_price
                                            break
                                
                                # Rating extraction
                                rating = 0.0
                                rating_elem = item.select_one('.a-icon-alt, .a-star-small, [aria-label*="out of"]')
                                if rating_elem:
                                    rating_text = rating_elem.get_text(strip=True) or rating_elem.get('aria-label', '')
                                    import re
                                    rating_match = re.search(r'([0-9.]+)\s*out of', rating_text)
                                    if rating_match:
                                        try:
                                            rating = float(rating_match.group(1))
                                        except:
                                            pass
                                
                                # Image extraction - multiple strategies with enhanced selectors
                                image_url = None
                                image_selectors = [
                                    'img.s-image',
                                    'img[data-image-latency="s-product-image"]',
                                    '.s-image[src]',
                                    'img[src*="media-amazon"]',
                                    'img[src*="images-amazon"]',
                                    'img[data-src*="amazon"]',  # Lazy-loaded images
                                    'img.product-image',
                                    'source[srcset]',  # Picture element
                                    'img[srcset]'
                                ]
                                for img_sel in image_selectors:
                                    img_elem = item.select_one(img_sel)
                                    if img_elem:
                                        # Try multiple attributes
                                        img_src = (img_elem.get('src') or 
                                                 img_elem.get('data-src') or 
                                                 img_elem.get('srcset', ''))
                                        if img_src:
                                            # Handle srcset format (multiple URLs)
                                            if ',' in img_src:
                                                img_src = img_src.split(',')[-1].split(' ')[0]  # Get highest res
                                            # Prefer higher resolution images
                                            if 'http' in img_src and ('media-amazon' in img_src or 'images-amazon' in img_src):
                                                image_url = img_src
                                                break
                                            elif 'http' in img_src:
                                                image_url = img_src  # Fallback to any valid image
                                                break
                                
                                if price > 0 and title and link:
                                    products.append({
                                        'name': title[:200],
                                        'price': price,
                                        'currency_symbol': '₹',
                                        'currency_code': 'INR',
                                        'source': 'Amazon India',
                                        'source_url': link,  # Already cleaned and validated
                                        'description': title[:300],
                                        'rating': rating,
                                        'image_url': image_url,
                                        'search_engine': 'Direct Amazon',
                                        'timestamp': search_timestamp
                                    })
                            except Exception:
                                continue
                        
                        new_count = len(products) - initial_count
                        if new_count > 0:
                            logger.info(f"Amazon India found {new_count} products (total: {len(products)})")
                            break  # Found results, no need to try other URLs
                        
                except Exception:
                    continue
                    
        except Exception as e:
            logger.warning(f"Amazon India search failed: {e}")
        
        # Source 2: Enhanced Flipkart search
        try:
            headers['User-Agent'] = rand_module.choice(user_agents)
            await asyncio.sleep(rand_module.uniform(0.5, 1.5))
            
            flipkart_url = f"https://www.flipkart.com/search?q={encoded_query}"
            response = await client.get(flipkart_url, headers=headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Multiple selector strategies for Flipkart
                selectors = ['[data-id]', '._1AtVbE', '._13oc-S', 'div._1xHGtK', '._2kHMtA']
                items = []
                for selector in selectors:
                    items = soup.select(selector)
                    if len(items) > 5:
                        break
                
                for item in items[:25]:
                    try:
                        # Title - multiple strategies
                        title = None
                        title_selectors = ['a[title]', '.s1Q9rs', '._4rR01T', 'a.IRpwTa', 'div._4rR01T']
                        for ts in title_selectors:
                            title_elem = item.select_one(ts)
                            if title_elem:
                                title = title_elem.get('title', '') or title_elem.get_text(strip=True)
                                if title and len(title) > 5:
                                    break
                        
                        if not title:
                            continue
                        
                        # Link - extract and validate
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href
                            else:
                                link = "https://www.flipkart.com" + href
                            
                            # Clean and validate Flipkart URL
                            link = clean_flipkart_url(link)
                            
                            # Skip if not a valid product URL
                            if not is_valid_product_url(link, 'Flipkart'):
                                continue
                        
                        # Price - multiple selectors
                        price = 0
                        price_selectors = ['._30jeq3', '._1_oo_3', 'div._30jeq3', '._3I9_wc']
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('div._3LWZlK, span._2_R_DZ')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            try:
                                rating = float(rating_text)
                            except:
                                pass
                        
                        # Image extraction - enhanced with multiple strategies
                        image_url = None
                        image_selectors = [
                            'img._396cs4',
                            'img[loading="eager"]',
                            'img[src*="flipkart.com"][src*="image"]',
                            'img[src*="rukmini"]',
                            'img.CXW8mj',
                            'img[data-src*="flipkart"]',  # Lazy-loaded
                            'img.product-image',
                            'source[srcset]',
                            'img[srcset]'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                # Try multiple attributes
                                img_src = (img_elem.get('src') or 
                                         img_elem.get('data-src') or 
                                         img_elem.get('srcset', ''))
                                if img_src:
                                    # Handle srcset format
                                    if ',' in img_src:
                                        img_src = img_src.split(',')[-1].split(' ')[0]  # Highest res
                                    if 'http' in img_src or img_src.startswith('//'):
                                        # Convert protocol-relative URLs
                                        if img_src.startswith('//'):
                                            img_src = 'https:' + img_src
                                        image_url = img_src
                                        break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Flipkart',
                                'source_url': link,  # Already cleaned and validated
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Flipkart',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Flipkart found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Flipkart search failed: {e}")
        
        # Source 3: MySmartPrice (price comparison site)
        try:
            msp_url = f"https://www.mysmartprice.com/search?s={encoded_query}"
            response = await client.get(msp_url, headers=headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                for item in soup.select('.prdct-item, .product-item')[:15]:
                    try:
                        title_elem = item.select_one('a[title], .prdct-item__name')
                        price_elem = item.select_one('.prdct-item__price, .price')
                        link_elem = item.select_one('a[href]')
                        
                        if title_elem and price_elem and link_elem:
                            title = title_elem.get('title', '') or title_elem.get_text(strip=True)
                            link = link_elem.get('href', '')
                            if link and not link.startswith('http'):
                                link = "https://www.mysmartprice.com" + link
                            
                            price_text = price_elem.get_text(strip=True)
                            price = extract_price_from_text(price_text)
                            
                            if price > 0:
                                products.append({
                                    'name': title[:150],
                                    'price': price,
                                    'currency_symbol': '₹',
                                    'currency_code': 'INR',
                                    'source': 'MySmartPrice',
                                    'source_url': link,
                                    'description': title,
                                    'search_engine': 'Price Comparison',
                                    'timestamp': search_timestamp
                                })
                    except Exception:
                        continue
                
                logger.info(f"MySmartPrice found {len(products) - initial_count} additional products")
        except Exception as e:
            logger.warning(f"MySmartPrice search failed: {e}")
        
        # Source 4: Bing Search with different user agents
        try:
            bing_headers = headers.copy()
            bing_headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
            
            bing_url = f"https://www.bing.com/search?q={encoded_query_price}&count=50&setlang=en-IN"
            response = await client.get(bing_url, headers=bing_headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                
                # Try multiple selectors for Bing results
                results = soup.find_all('li', class_='b_algo')
                if not results:
                    results = soup.find_all('div', class_='b_algo')
                if not results:
                    results = soup.select('.b_algo, .b_ans, .b_top')
                
                for result in results[:max_results]:
                    try:
                        link_elem = result.find('a', href=True)
                        if not link_elem:
                            continue
                            
                        title = link_elem.get_text(strip=True)
                        link = link_elem.get('href', '')
                        
                        # Get all text from the result
                        all_text = result.get_text(separator=' ', strip=True)
                        
                        price = extract_price_from_text(all_text)
                        
                        if price > 0:
                            vendor = extract_vendor_from_url(link)
                            products.append({
                                'name': title[:150],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': vendor,
                                'source_url': link,
                                'description': all_text[:300],
                                'search_engine': 'Bing',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                        
            logger.info(f"Bing search total: {len(products)} products")
        except Exception as e:
            logger.warning(f"Bing search failed: {e}")
        
        # Source 5: DuckDuckGo HTML Search
        try:
            ddg_url = f"https://html.duckduckgo.com/html/?q={encoded_query_price}"
            response = await client.get(ddg_url, headers=headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                for result in soup.find_all('div', class_='result')[:20]:
                    try:
                        link_elem = result.find('a', class_='result__a')
                        if not link_elem:
                            continue
                            
                        title = link_elem.get_text(strip=True)
                        link = link_elem.get('href', '')
                        
                        snippet_elem = result.find('a', class_='result__snippet')
                        snippet = snippet_elem.get_text(strip=True) if snippet_elem else ''
                        
                        full_text = f"{title} {snippet}"
                        price = extract_price_from_text(full_text)
                        
                        if price > 0:
                            vendor = extract_vendor_from_url(link)
                            products.append({
                                'name': title[:150],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': vendor,
                                'source_url': link,
                                'description': snippet[:300],
                                'search_engine': 'DuckDuckGo',
                                'timestamp': search_timestamp
                            })
                    except Exception as e:
                        continue
                        
            logger.info(f"DuckDuckGo found {len(products) - initial_count} additional prices")
        except Exception as e:
            logger.warning(f"DuckDuckGo search failed: {e}")
        
        # Source 6: Enhanced Snapdeal search
        try:
            headers['User-Agent'] = rand_module.choice(user_agents)
            await asyncio.sleep(rand_module.uniform(0.3, 1.0))
            
            snapdeal_url = f"https://www.snapdeal.com/search?keyword={encoded_query}"
            response = await client.get(snapdeal_url, headers=headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                for item in soup.select('.product-tuple-listing, .col-xs-6, .product-tuple')[:25]:
                    try:
                        title_elem = item.select_one('.product-title, p[title], .prodName')
                        price_elem = item.select_one('.product-price, .lfloat.product-price, .lfloat')
                        link_elem = item.select_one('a[href]')
                        
                        if title_elem and price_elem and link_elem:
                            title = title_elem.get('title', '') or title_elem.get_text(strip=True)
                            link = link_elem.get('href', '')
                            if not link.startswith('http'):
                                link = 'https://www.snapdeal.com' + link
                            
                            # Clean and validate Snapdeal URL
                            link = clean_snapdeal_url(link)
                            
                            # Skip if not a valid product URL
                            if not is_valid_product_url(link, 'Snapdeal'):
                                continue
                            
                            price_text = price_elem.get_text(strip=True)
                            price = extract_price_from_text(price_text)
                            
                            # Rating
                            rating = 0.0
                            rating_elem = item.select_one('.filled-stars, [class*="rating"]')
                            if rating_elem:
                                rating_text = rating_elem.get_text(strip=True)
                                import re
                                rating_match = re.search(r'([0-9.]+)', rating_text)
                                if rating_match:
                                    try:
                                        rating = float(rating_match.group(1))
                                    except:
                                        pass
                            
                            # Image extraction - enhanced with multiple strategies
                            image_url = None
                            image_selectors = [
                                'img.product-image',
                                'img[src*="snapdeal.com"][src*="picture"]',
                                'img[src*="n1.sdlcdn.com"]',
                                'img[src*="n2.sdlcdn.com"]',
                                'img[data-src*="sdlcdn"]',  # Lazy-loaded
                                'source[srcset]',
                                'img[srcset]'
                            ]
                            for img_sel in image_selectors:
                                img_elem = item.select_one(img_sel)
                                if img_elem:
                                    img_src = (img_elem.get('src') or 
                                             img_elem.get('data-src') or 
                                             img_elem.get('srcset', ''))
                                    if img_src and ('http' in img_src or img_src.startswith('//')):  
                                        if img_src.startswith('//'):
                                            img_src = 'https:' + img_src
                                        # Get first URL from srcset if multiple
                                        image_url = img_src.split(',')[0].split(' ')[0]
                                        break
                            
                            if price > 0 and title:
                                products.append({
                                    'name': title[:200],
                                    'price': price,
                                    'currency_symbol': '₹',
                                    'currency_code': 'INR',
                                    'source': 'Snapdeal',
                                    'source_url': link,  # Already cleaned and validated
                                    'description': title[:300],
                                    'rating': rating,
                                    'image_url': image_url,
                                    'search_engine': 'Direct Snapdeal',
                                    'timestamp': search_timestamp
                                })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Snapdeal found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Snapdeal search failed: {e}")
        
        # Source 7: Myntra search (Fashion category)
        try:
            headers['User-Agent'] = rand_module.choice(user_agents)
            await asyncio.sleep(rand_module.uniform(0.5, 1.5))
            
            myntra_url = f"https://www.myntra.com/{encoded_query}"
            response = await client.get(myntra_url, headers=headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Myntra uses multiple container classes
                selectors = ['.product-base', '.product-productMetaInfo', 'li.product-base']
                items = []
                for selector in selectors:
                    items = soup.select(selector)
                    if len(items) > 3:
                        break
                
                for item in items[:25]:
                    try:
                        # Title extraction - multiple strategies
                        title = None
                        title_selectors = [
                            'h3.product-brand', 
                            'h4.product-product',
                            '.product-brand',
                            '.product-product',
                            'a[title]'
                        ]
                        for ts in title_selectors:
                            title_elem = item.select_one(ts)
                            if title_elem:
                                title_text = title_elem.get('title', '') or title_elem.get_text(strip=True)
                                if title_text and len(title_text) > 3:
                                    # Combine brand and product name
                                    brand = item.select_one('.product-brand')
                                    prod_name = item.select_one('.product-product')
                                    if brand and prod_name:
                                        title = f"{brand.get_text(strip=True)} {prod_name.get_text(strip=True)}"
                                    else:
                                        title = title_text
                                    break
                        
                        if not title:
                            continue
                        
                        # Link extraction
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]  # Remove query params
                            else:
                                link = f"https://www.myntra.com{href}".split('?')[0]
                        
                        if not link or '/search?' in link or not link:
                            continue
                        
                        # Price extraction - Myntra specific classes
                        price = 0
                        price_selectors = [
                            '.product-discountedPrice',
                            '.product-price',
                            'span.product-discountedPrice',
                            'div.product-price span'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating extraction
                        rating = 0.0
                        rating_elem = item.select_one('.product-rating, .product-ratingsContainer span')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image extraction - Myntra specific
                        image_url = None
                        image_selectors = [
                            'img.product-image',
                            'img[src*="myntra.com"]',
                            'img[src*="assets.myntassets"]',
                            'picture img',
                            'source[srcset]'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                img_src = img_elem.get('src') or img_elem.get('data-src') or img_elem.get('srcset', '')
                                if img_src:
                                    # Handle srcset format
                                    if ',' in img_src:
                                        img_src = img_src.split(',')[0].split(' ')[0]
                                    if 'http' in img_src or img_src.startswith('//'):
                                        if img_src.startswith('//'):
                                            img_src = 'https:' + img_src
                                        image_url = img_src
                                        break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Myntra',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Myntra',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Myntra found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Myntra search failed: {e}")
        
        # Source 8: Ajio search (Fashion category)
        try:
            # Enhanced headers for better bot protection bypass
            ajio_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.ajio.com/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            ajio_url = f"https://www.ajio.com/search/?text={encoded_query}"
            response = await client.get(ajio_url, headers=ajio_headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Ajio product containers
                items = soup.select('.item, .product, [class*="product-item"]')[:25]
                
                for item in items:
                    try:
                        # Title - combine brand and name
                        title = None
                        brand_elem = item.select_one('.brand, .brand-name, [class*="brand"]')
                        name_elem = item.select_one('.name, .product-name, .nameCls')
                        
                        if brand_elem and name_elem:
                            brand = brand_elem.get_text(strip=True)
                            name = name_elem.get_text(strip=True)
                            title = f"{brand} {name}"
                        elif name_elem:
                            title = name_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Link extraction
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            else:
                                link = f"https://www.ajio.com{href}".split('?')[0]
                        
                        if not link or '/search' in link:
                            continue
                        
                        # Price extraction
                        price = 0
                        price_selectors = [
                            '.price, .price-value',
                            '[class*="price"]',
                            'span[class*="price"]'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('.rating, [class*="rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image extraction
                        image_url = None
                        image_selectors = [
                            'img[src*="ajio"]',
                            'img.product-image',
                            'img[data-src]',
                            'picture img',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Ajio',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Ajio',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Ajio found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Ajio search failed: {e}")
        
        # Source 9: Pepperfry search (Furniture category)
        try:
            # Enhanced headers for better bot protection bypass
            pepperfry_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.pepperfry.com/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            pepperfry_url = f"https://www.pepperfry.com/search?q={encoded_query}"
            response = await client.get(pepperfry_url, headers=pepperfry_headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                items = soup.select('.product-card, .pf-product, [data-productid]')[:25]
                
                for item in items:
                    try:
                        # Title
                        title_elem = item.select_one('.product-title, .pf-product-name, h3, a[title]')
                        if not title_elem:
                            continue
                        title = title_elem.get('title', '') or title_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Link
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            else:
                                link = f"https://www.pepperfry.com{href}".split('?')[0]
                        
                        if not link or '/search' in link:
                            continue
                        
                        # Price
                        price = 0
                        price_selectors = [
                            '.product-price',
                            '.pf-selling-price',
                            '[class*="price"]'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('.rating, [class*="rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image
                        image_url = None
                        image_selectors = [
                            'img[src*="pepperfry"]',
                            'img.product-image',
                            'img[data-src]',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Pepperfry',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Pepperfry',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Pepperfry found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Pepperfry search failed: {e}")
        
        # Source 10: Tata CLiQ search (Fashion category)
        try:
            # Enhanced headers for better bot protection bypass
            tatacliq_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.tatacliq.com/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            tatacliq_url = f"https://www.tatacliq.com/search/?searchCategory=all&text={encoded_query}"
            response = await client.get(tatacliq_url, headers=tatacliq_headers, follow_redirects=True)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Tata CLiQ product containers
                items = soup.select('.ProductModule__base, .ProductModule, [class*="ProductCard"], [data-test="product"]')[:25]
                
                for item in items:
                    try:
                        # Title
                        title_elem = item.select_one('.ProductModule__productTitle, .ProductDescription__header, h2, h3, [class*="productTitle"]')
                        if not title_elem:
                            continue
                        title = title_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Brand name (if separate)
                        brand_elem = item.select_one('.ProductModule__brand, .ProductDescription__brand, [class*="brand"]')
                        if brand_elem:
                            brand = brand_elem.get_text(strip=True)
                            if brand and brand.lower() not in title.lower():
                                title = f"{brand} {title}"
                        
                        # Link
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            elif href.startswith('/'):
                                link = f"https://www.tatacliq.com{href}".split('?')[0]
                        
                        if not link or '/search' in link or link.endswith('.com'):
                            continue
                        
                        # Price
                        price = 0
                        price_selectors = [
                            '.ProductModule__price',
                            '.ProductDescription__priceHolder',
                            '[class*="actualPrice"]',
                            '[class*="price"]'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('.RatingStars__averageRating, [class*="rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image
                        image_url = None
                        image_selectors = [
                            'img[src*="tatacliq"]',
                            'img.ProductModule__image',
                            'img[data-src]',
                            'source[srcset]',
                            'img[srcset]',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                # Try srcset first (highest quality)
                                srcset = img_elem.get('srcset', '')
                                if srcset:
                                    srcset_parts = srcset.split(',')
                                    if srcset_parts:
                                        img_src = srcset_parts[-1].strip().split(' ')[0]
                                        if img_src and ('http' in img_src or img_src.startswith('//')):
                                            if img_src.startswith('//'):
                                                img_src = 'https:' + img_src
                                            image_url = img_src
                                            break
                                
                                # Try src and data-src
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Tata CLiQ',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Tata CLiQ',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Tata CLiQ found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Tata CLiQ search failed: {e}")
        
        # Source 11: Urban Ladder search (Furniture category)
        try:
            # Enhanced headers for better bot protection bypass
            urbanladder_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.urbanladder.com/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            urbanladder_url = f"https://www.urbanladder.com/search?q={encoded_query}"
            response = await client.get(urbanladder_url, headers=urbanladder_headers, follow_redirects=True)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Urban Ladder product containers
                items = soup.select('.product-item, .productcard, [data-pid], [class*="ProductCard"]')[:25]
                
                for item in items:
                    try:
                        # Title
                        title_elem = item.select_one('.product-title, .productcard__title, h2, h3, a[title]')
                        if not title_elem:
                            continue
                        title = title_elem.get('title', '') or title_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Link
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            elif href.startswith('/'):
                                link = f"https://www.urbanladder.com{href}".split('?')[0]
                        
                        if not link or '/search' in link or link.endswith('.com'):
                            continue
                        
                        # Price
                        price = 0
                        price_selectors = [
                            '.product-price',
                            '.productcard__price',
                            '[class*="price"]',
                            '.price-tag'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('.rating, [class*="rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image
                        image_url = None
                        image_selectors = [
                            'img[src*="urbanladder"]',
                            'img.product-image',
                            'img[data-src]',
                            'source[srcset]',
                            'img[srcset]',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                # Try srcset first
                                srcset = img_elem.get('srcset', '')
                                if srcset:
                                    srcset_parts = srcset.split(',')
                                    if srcset_parts:
                                        img_src = srcset_parts[-1].strip().split(' ')[0]
                                        if img_src and ('http' in img_src or img_src.startswith('//')):
                                            if img_src.startswith('//'):
                                                img_src = 'https:' + img_src
                                            image_url = img_src
                                            break
                                
                                # Try src and data-src
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Urban Ladder',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Urban Ladder',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Urban Ladder found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Urban Ladder search failed: {e}")
        
        # Source 12: Meesho search (Fashion category - affordable fashion)
        try:
            # Enhanced headers for Meesho
            meesho_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.meesho.com/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            meesho_url = f"https://www.meesho.com/search?q={encoded_query}"
            response = await client.get(meesho_url, headers=meesho_headers, follow_redirects=True)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Meesho product containers
                items = soup.select('[class*="ProductCard"], [class*="product-card"], [data-testid="product-card"]')[:25]
                
                for item in items:
                    try:
                        # Title
                        title_elem = item.select_one('[class*="ProductCard__ProductCard_Name"], [class*="product-title"], h2, h3, p[class*="title"]')
                        if not title_elem:
                            continue
                        title = title_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Link
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            elif href.startswith('/'):
                                link = f"https://www.meesho.com{href}".split('?')[0]
                        
                        if not link or '/search' in link or link.endswith('.com'):
                            continue
                        
                        # Price
                        price = 0
                        price_selectors = [
                            '[class*="ProductCard__Price"]',
                            '[class*="product-price"]',
                            '[class*="price"]',
                            'span[class*="Price"]'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('[class*="rating"], [class*="Rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image
                        image_url = None
                        image_selectors = [
                            'img[src*="meesho"]',
                            'img[data-src]',
                            'source[srcset]',
                            'img[srcset]',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                srcset = img_elem.get('srcset', '')
                                if srcset:
                                    srcset_parts = srcset.split(',')
                                    if srcset_parts:
                                        img_src = srcset_parts[-1].strip().split(' ')[0]
                                        if img_src and ('http' in img_src or img_src.startswith('//')):
                                            if img_src.startswith('//'):
                                                img_src = 'https:' + img_src
                                            image_url = img_src
                                            break
                                
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Meesho',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Meesho',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Meesho found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Meesho search failed: {e}")
        
        # Source 13: Croma search (Electronics category)
        try:
            # Enhanced headers for Croma
            croma_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.croma.com/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            croma_url = f"https://www.croma.com/searchresult?q={encoded_query}"
            response = await client.get(croma_url, headers=croma_headers, follow_redirects=True)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Croma product containers
                items = soup.select('.product, .product-item, [class*="ProductCard"], li.product')[:25]
                
                for item in items:
                    try:
                        # Title
                        title_elem = item.select_one('.product-title, .product-name, h2, h3, a[title]')
                        if not title_elem:
                            continue
                        title = title_elem.get('title', '') or title_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Link
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            elif href.startswith('/'):
                                link = f"https://www.croma.com{href}".split('?')[0]
                        
                        if not link or '/search' in link or link.endswith('.com'):
                            continue
                        
                        # Price
                        price = 0
                        price_selectors = [
                            '.amount, .price, .product-price',
                            '[class*="price"]',
                            'span.amount'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('.rating, [class*="rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image
                        image_url = None
                        image_selectors = [
                            'img[src*="croma"]',
                            'img.product-image',
                            'img[data-src]',
                            'source[srcset]',
                            'img[srcset]',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                srcset = img_elem.get('srcset', '')
                                if srcset:
                                    srcset_parts = srcset.split(',')
                                    if srcset_parts:
                                        img_src = srcset_parts[-1].strip().split(' ')[0]
                                        if img_src and ('http' in img_src or img_src.startswith('//')):
                                            if img_src.startswith('//'):
                                                img_src = 'https:' + img_src
                                            image_url = img_src
                                            break
                                
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Croma',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Croma',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Croma found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Croma search failed: {e}")
        
        # Source 14: Reliance Digital search (Electronics category)
        try:
            # Enhanced headers for Reliance Digital
            reliance_headers = {
                'User-Agent': rand_module.choice(user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Referer': 'https://www.reliancedigital.in/',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            await asyncio.sleep(rand_module.uniform(1.0, 2.0))
            
            reliance_url = f"https://www.reliancedigital.in/search?q={encoded_query}"
            response = await client.get(reliance_url, headers=reliance_headers, follow_redirects=True)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Reliance Digital product containers
                items = soup.select('.sp, .productCard, [class*="product"]')[:25]
                
                for item in items:
                    try:
                        # Title
                        title_elem = item.select_one('.sp__name, .productCard__title, h2, h3')
                        if not title_elem:
                            continue
                        title = title_elem.get_text(strip=True)
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # Link
                        link = ""
                        link_elem = item.select_one('a[href]')
                        if link_elem and link_elem.get('href'):
                            href = link_elem.get('href', '')
                            if href.startswith('http'):
                                link = href.split('?')[0]
                            elif href.startswith('/'):
                                link = f"https://www.reliancedigital.in{href}".split('?')[0]
                        
                        if not link or '/search' in link or link.endswith('.in'):
                            continue
                        
                        # Price
                        price = 0
                        price_selectors = [
                            '.sp__price',
                            '.productCard__price',
                            '[class*="price"]',
                            'span.price'
                        ]
                        for ps in price_selectors:
                            price_elem = item.select_one(ps)
                            if price_elem:
                                price_text = price_elem.get_text(strip=True)
                                price = extract_price_from_text(price_text)
                                if price > 0:
                                    break
                        
                        # Rating
                        rating = 0.0
                        rating_elem = item.select_one('.rating, [class*="rating"]')
                        if rating_elem:
                            rating_text = rating_elem.get_text(strip=True)
                            rating_match = re.search(r'([0-9.]+)', rating_text)
                            if rating_match:
                                try:
                                    rating = float(rating_match.group(1))
                                except:
                                    pass
                        
                        # Image
                        image_url = None
                        image_selectors = [
                            'img[src*="reliancedigital"]',
                            'img.sp__image',
                            'img[data-src]',
                            'source[srcset]',
                            'img[srcset]',
                            'img'
                        ]
                        for img_sel in image_selectors:
                            img_elem = item.select_one(img_sel)
                            if img_elem:
                                srcset = img_elem.get('srcset', '')
                                if srcset:
                                    srcset_parts = srcset.split(',')
                                    if srcset_parts:
                                        img_src = srcset_parts[-1].strip().split(' ')[0]
                                        if img_src and ('http' in img_src or img_src.startswith('//')):
                                            if img_src.startswith('//'):
                                                img_src = 'https:' + img_src
                                            image_url = img_src
                                            break
                                
                                img_src = img_elem.get('src') or img_elem.get('data-src') or ''
                                if img_src and ('http' in img_src or img_src.startswith('//')):
                                    if img_src.startswith('//'):
                                        img_src = 'https:' + img_src
                                    image_url = img_src
                                    break
                        
                        if price > 0 and title and link:
                            products.append({
                                'name': title[:200],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': 'Reliance Digital',
                                'source_url': link,
                                'description': title[:300],
                                'rating': rating,
                                'image_url': image_url,
                                'search_engine': 'Direct Reliance Digital',
                                'timestamp': search_timestamp
                            })
                    except Exception:
                        continue
                
                new_count = len(products) - initial_count
                if new_count > 0:
                    logger.info(f"Reliance Digital found {new_count} products (total: {len(products)})")
        except Exception as e:
            logger.warning(f"Reliance Digital search failed: {e}")
        
        # Source 15: Google Shopping search results (always active for all categories)
        ecommerce_queries = [
            f"site:amazon.in {clean_query} price",
            f"site:flipkart.com {clean_query} price",
            f"site:snapdeal.com {clean_query} price",
            f"{clean_query} buy online india price",
        ]
        
        for eq in ecommerce_queries:
            try:
                eq_encoded = quote_plus(eq)
                bing_ecom_url = f"https://www.bing.com/search?q={eq_encoded}&count=10"
                response = await client.get(bing_ecom_url, headers=headers)
                
                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, 'lxml')
                    
                    for result in soup.find_all('li', class_='b_algo')[:10]:
                        try:
                            link_elem = result.find('a')
                            if not link_elem:
                                continue
                                
                            title = link_elem.get_text(strip=True)
                            link = link_elem.get('href', '')
                            
                            snippet_elem = result.find('p')
                            snippet = snippet_elem.get_text(strip=True) if snippet_elem else ''
                            
                            full_text = f"{title} {snippet}"
                            price = extract_price_from_text(full_text)
                            
                            if price > 0:
                                vendor = extract_vendor_from_url(link)
                                products.append({
                                    'name': title[:150],
                                    'price': price,
                                    'currency_symbol': '₹',
                                    'currency_code': 'INR',
                                    'source': vendor,
                                    'source_url': link,
                                    'description': snippet[:300],
                                    'search_engine': 'Bing-Ecommerce',
                                    'timestamp': search_timestamp
                                })
                        except:
                            continue
                            
                await asyncio.sleep(0.3)  # Rate limiting
            except Exception as e:
                continue
        
        # Source 4: Google search via scraping (fallback)
        try:
            google_url = f"https://www.google.com/search?q={encoded_query}&num=20&hl=en"
            google_headers = headers.copy()
            google_headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            
            response = await client.get(google_url, headers=google_headers)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                initial_count = len(products)
                
                # Look for search result divs
                for result in soup.find_all('div', class_='g')[:15]:
                    try:
                        link_elem = result.find('a')
                        if not link_elem:
                            continue
                            
                        link = link_elem.get('href', '')
                        if not link.startswith('http'):
                            continue
                            
                        title_elem = result.find('h3')
                        title = title_elem.get_text(strip=True) if title_elem else ''
                        
                        snippet_elem = result.find('div', class_=lambda x: x and 'VwiC3b' in str(x))
                        snippet = snippet_elem.get_text(strip=True) if snippet_elem else ''
                        
                        full_text = f"{title} {snippet}"
                        price = extract_price_from_text(full_text)
                        
                        if price > 0:
                            vendor = extract_vendor_from_url(link)
                            products.append({
                                'name': title[:150],
                                'price': price,
                                'currency_symbol': '₹',
                                'currency_code': 'INR',
                                'source': vendor,
                                'source_url': link,
                                'description': snippet[:300],
                                'search_engine': 'Google',
                                'timestamp': search_timestamp
                            })
                    except:
                        continue
                        
                logger.info(f"Google found {len(products) - initial_count} additional prices")
        except Exception as e:
            logger.warning(f"Google search failed: {e}")
    
    # Remove duplicates based on price + vendor
    seen = set()
    unique_products = []
    for p in products:
        key = (round(p['price'], -2), p['source'])  # Round to nearest 100 for dedup
        if key not in seen:
            seen.add(key)
            unique_products.append(p)
    
    # Log comprehensive scraper status
    sources_found = {}
    for p in unique_products:
        source = p.get('source', 'Unknown')
        sources_found[source] = sources_found.get(source, 0) + 1
    
    logger.info(f"Real web search returned {len(unique_products)} unique prices for: {query}")
    logger.info(f"📊 Source breakdown: {', '.join([f'{src}={count}' for src, count in sorted(sources_found.items(), key=lambda x: x[1], reverse=True)])}")
    
    return unique_products


def validate_and_recheck_prices(prices: List[Dict], query: str) -> List[Dict]:
    """
    Validate prices and flag anomalies for re-checking.
    Uses statistical methods to identify outliers.
    """
    if len(prices) < 2:
        return prices
    
    price_values = [p['price'] for p in prices]
    
    # Calculate IQR for outlier detection
    sorted_prices = sorted(price_values)
    n = len(sorted_prices)
    q1_idx = n // 4
    q3_idx = (3 * n) // 4
    q1 = sorted_prices[q1_idx]
    q3 = sorted_prices[q3_idx]
    iqr = q3 - q1
    
    # Define bounds (1.5 * IQR rule)
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr
    
    # Mark anomalies
    validated = []
    for p in prices:
        price = p['price']
        is_anomaly = price < lower_bound or price > upper_bound
        
        p_copy = p.copy()
        p_copy['is_validated'] = not is_anomaly
        p_copy['validation_note'] = 'Outlier detected - verify source' if is_anomaly else 'Price verified'
        
        # Only include non-outliers or if we have very few results
        if not is_anomaly or len(prices) <= 5:
            validated.append(p_copy)
    
    return validated


def calculate_min_med_max_from_real_prices(validated_prices: List[Dict]) -> Dict:
    """
    Calculate Min, Median, Max from real web-scraped prices.
    Returns the prices along with their source information.
    """
    if not validated_prices:
        return {
            'min': {'price': None, 'source': None, 'url': None},
            'med': {'price': None, 'source': None, 'url': None},
            'max': {'price': None, 'source': None, 'url': None},
            'all_sources': []
        }
    
    # Sort by price
    sorted_by_price = sorted(validated_prices, key=lambda x: x['price'])
    
    # Get min (lowest price)
    min_item = sorted_by_price[0]
    min_price = min_item['price']
    min_source = min_item.get('source', 'Unknown')
    min_url = min_item.get('source_url', '')
    
    # Get max (highest price)
    max_item = sorted_by_price[-1]
    max_price = max_item['price']
    max_source = max_item.get('source', 'Unknown')
    max_url = max_item.get('source_url', '')
    
    # Get median
    n = len(sorted_by_price)
    mid_idx = n // 2
    if n % 2 == 0:
        med_price = (sorted_by_price[mid_idx - 1]['price'] + sorted_by_price[mid_idx]['price']) / 2
        med_item = sorted_by_price[mid_idx]  # Use the higher median item for source
    else:
        med_price = sorted_by_price[mid_idx]['price']
        med_item = sorted_by_price[mid_idx]
    
    med_source = med_item.get('source', 'Unknown')
    med_url = med_item.get('source_url', '')
    
    # Collect all sources for traceability
    all_sources = []
    for p in validated_prices:
        all_sources.append({
            'price': p['price'],
            'source': p.get('source', 'Unknown'),
            'url': p.get('source_url', ''),
            'search_engine': p.get('search_engine', 'Unknown'),
            'timestamp': p.get('timestamp', '')
        })
    
    return {
        'min': {'price': min_price, 'source': min_source, 'url': min_url},
        'med': {'price': med_price, 'source': med_source, 'url': med_url},
        'max': {'price': max_price, 'source': max_source, 'url': max_url},
        'all_sources': all_sources,
        'price_count': len(validated_prices)
    }

def extract_price_from_text(text: str) -> float:
    """Extract price from text containing INR/Rs prices with improved patterns"""
    import re
    
    # First, try to find prices with explicit currency symbols/prefixes
    explicit_patterns = [
        r'₹\s*([\d,]+(?:\.\d{1,2})?)',  # ₹1,234 or ₹1,234.00 or ₹1,234.5
        r'Rs\.?\s*([\d,]+(?:\.\d{1,2})?)',  # Rs.1,234 or Rs 1234
        r'INR\s*([\d,]+(?:\.\d{1,2})?)',  # INR 1,234
        r'(?:Price|MRP|Cost|Starts?\s*(?:at|from)?|Only)[\s:]*[₹Rs\.]*\s*([\d,]+)',  # Price: 1234
        r'([\d,]+)\s*(?:rupees|rs\.?|inr)',  # 1234 rupees
        r'\$\s*([\d,]+(?:\.\d{2})?)',  # $1,234.00 (for USD, convert to INR)
    ]
    
    all_prices = []
    
    for pattern in explicit_patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                price_str = match.group(1).replace(',', '')
                price = float(price_str)
                # Validate price is reasonable (between 500 and 10 million INR)
                # Exclude year-like numbers (2020-2030)
                if 500 <= price <= 10000000 and not (2019 <= price <= 2030):
                    all_prices.append(price)
            except (ValueError, IndexError):
                continue
    
    # Return the most reasonable price
    if all_prices:
        # Sort and return a middle-range price (avoid outliers)
        all_prices.sort()
        if len(all_prices) >= 3:
            return all_prices[len(all_prices) // 2]  # Return median
        return all_prices[0]
    
    return 0

def extract_vendor_from_url(url: str) -> str:
    """Extract vendor name from URL, including handling redirect URLs"""
    from urllib.parse import urlparse, unquote, parse_qs
    
    vendor_mappings = {
        'amazon': 'Amazon',
        'flipkart': 'Flipkart',
        'myntra': 'Myntra',
        'snapdeal': 'Snapdeal',
        'paytmmall': 'Paytm Mall',
        'tatacliq': 'Tata CLiQ',
        'reliancedigital': 'Reliance Digital',
        'croma': 'Croma',
        'vijaysales': 'Vijay Sales',
        'industrybuying': 'IndustryBuying',
        'indiamart': 'IndiaMart',
        'tradeindia': 'TradeIndia',
        'justdial': 'JustDial',
        'moglix': 'Moglix',
        'buildingmaterials': 'Building Materials',
        'buildersmart': 'BuildersMart',
        'materialtree': 'MaterialTree',
        'pepperfry': 'Pepperfry',
        'urban': 'Urban Ladder',
        'godrej': 'Godrej',
        'ikea': 'IKEA',
        'smartprix': 'Smartprix',
        'digit': 'Digit',
        '91mobiles': '91mobiles',
        'pricehistory': 'PriceHistory',
        'gadgets360': 'Gadgets360',
        'gsmarena': 'GSMArena',
        'mysmartprice': 'MySmartPrice',
        'pricebaba': 'PriceBaba',
        'apple': 'Apple Store',
        'samsung': 'Samsung Store',
        'oneplus': 'OnePlus Store',
    }
    
    try:
        # Handle DuckDuckGo redirect URLs
        if 'duckduckgo.com' in url:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'uddg' in params:
                url = unquote(params['uddg'][0])
        
        # Handle Bing redirect URLs
        if 'bing.com' in url and '/ck/a' in url:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'u' in params:
                url = unquote(params['u'][0])
        
        # Handle Google redirect URLs  
        if 'google.com/url' in url:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'url' in params:
                url = unquote(params['url'][0])
            elif 'q' in params:
                url = unquote(params['q'][0])
        
        domain = urlparse(url).netloc.lower()
        domain = domain.replace('www.', '')
        for key, vendor in vendor_mappings.items():
            if key in domain:
                return vendor
        # Return domain name if no mapping found
        return domain.replace('www.', '').split('.')[0].title()
    except:
        return 'Unknown'

# ================== REAL SERPAPI SEARCH ==================
async def search_with_serpapi(query: str, country: str = "in", max_results: int = 30, city: str = "") -> List[Dict]:
    """
    Search Google Shopping using SerpAPI for real product data.
    Returns actual prices and working product links from real marketplaces.
    Only returns REAL data from Google Shopping - no generated/fake data.
    """
    if not SERPAPI_API_KEY:
        logger.warning("SerpAPI key not configured, falling back to mock data")
        return []
    
    try:
        # Map country to Google Shopping parameters
        country_params = {
            "india": {"gl": "in", "hl": "en", "location": "India", "currency": "INR", "country": "india"},
            "usa": {"gl": "us", "hl": "en", "location": "United States", "currency": "USD", "country": "usa"},
            "uk": {"gl": "uk", "hl": "en", "location": "United Kingdom", "currency": "GBP", "country": "uk"},
            "uae": {"gl": "ae", "hl": "en", "location": "United Arab Emirates", "currency": "AED", "country": "uae"},
            "europe": {"gl": "de", "hl": "en", "location": "Germany", "currency": "EUR", "country": "europe"},
            "japan": {"gl": "jp", "hl": "en", "location": "Japan", "currency": "JPY", "country": "japan"},
            "australia": {"gl": "au", "hl": "en", "location": "Australia", "currency": "AUD", "country": "australia"},
            "canada": {"gl": "ca", "hl": "en", "location": "Canada", "currency": "CAD", "country": "canada"},
            "global": {"gl": "us", "hl": "en", "location": "United States", "currency": "USD", "country": "usa"}
        }
        
        params = country_params.get(country.lower(), country_params["india"])
        
        search_params = {
            "q": query,
            "api_key": SERPAPI_API_KEY,
            "engine": "google_shopping",
            "gl": params["gl"],
            "hl": params["hl"],
            "num": min(max_results, 100)
        }
        
        logger.info(f"SerpAPI search: query='{query}', country={params['gl']}, city={city}")
        
        # Run SerpAPI search in thread pool to avoid blocking
        loop = asyncio.get_event_loop()
        search = GoogleSearch(search_params)
        api_response = await loop.run_in_executor(None, search.get_dict)
        
        # Check for API errors
        if "error" in api_response:
            error_msg = api_response.get("error", "Unknown error")
            logger.error(f"SerpAPI Error: {error_msg}")
            if "run out of searches" in error_msg.lower() or "quota" in error_msg.lower():
                raise Exception("SerpAPI quota exhausted. Please add more credits to your SerpAPI account.")
            return []
        
        products = []
        currency_symbol = "₹" if params["currency"] == "INR" else "$" if params["currency"] == "USD" else params["currency"]
        
        # Parse inline shopping results (featured products at top)
        if "inline_shopping_results" in api_response:
            for idx, item in enumerate(api_response["inline_shopping_results"]):
                price = item.get("extracted_price", 0)
                if price and price > 0:
                    source_name = item.get("source", "Google Shopping")
                    product_title = item.get("title", "Unknown Product")
                    
                    # Try to generate direct vendor link
                    direct_link = get_direct_vendor_link(source_name, product_title)
                    
                    # Get the actual Google Shopping link from SerpAPI
                    google_shopping_link = item.get("link") or ""
                    
                    # Use our vendor link if we have a mapping, otherwise use Google Shopping link
                    final_url = direct_link if direct_link else google_shopping_link
                    
                    # Only include REAL data from SerpAPI - no fake/generated data
                    product_data = {
                        "name": product_title,
                        "price": float(price),
                        "currency_symbol": currency_symbol,
                        "currency_code": params["currency"],
                        "source": source_name,
                        "source_url": final_url,  # Best available link
                        "google_shopping_url": google_shopping_link,  # Keep Google link as backup
                        "description": item.get("snippet", ""),
                        "rating": item.get("rating") if item.get("rating") else None,
                        "availability": "In Stock",
                        "unit": "per piece",
                        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                        "image": item.get("thumbnail", ""),
                        "location": params["location"],
                        "review_count": item.get("reviews") if item.get("reviews") else None,
                        "position": idx + 1,
                        "is_real_data": True,
                        # Real vendor info from SerpAPI (only what's actually available)
                        "vendor": {
                            "vendor_name": source_name,
                            "vendor_website": final_url,
                            "is_real_data": True,
                            "data_source": "Google Shopping"
                        }
                    }
                    products.append(product_data)
        
        # Parse main shopping results
        if "shopping_results" in api_response:
            for idx, item in enumerate(api_response["shopping_results"]):
                price = item.get("extracted_price", 0)
                if price and price > 0:
                    source_name = item.get("source", "Google Shopping")
                    product_title = item.get("title", "Unknown Product")
                    
                    # Try to generate direct vendor link
                    direct_link = get_direct_vendor_link(source_name, product_title)
                    
                    # Get the actual Google Shopping link from SerpAPI
                    google_shopping_link = item.get("product_link") or item.get("link") or ""
                    
                    # Use our vendor link if we have a mapping, otherwise use Google Shopping link
                    final_url = direct_link if direct_link else google_shopping_link
                    
                    # Only include REAL data from SerpAPI - no fake/generated data
                    product_data = {
                        "name": product_title,
                        "price": float(price),
                        "currency_symbol": currency_symbol,
                        "currency_code": params["currency"],
                        "source": source_name,
                        "source_url": final_url,  # Best available link
                        "google_shopping_url": google_shopping_link,  # Keep Google link as backup
                        "description": item.get("snippet", ""),
                        "rating": item.get("rating") if item.get("rating") else None,
                        "availability": "In Stock" if not item.get("second_hand_condition") else "Used",
                        "unit": "per piece",
                        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                        "image": item.get("thumbnail", ""),
                        "location": params["location"],
                        "review_count": item.get("reviews") if item.get("reviews") else None,
                        "position": len(products) + 1,
                        "is_real_data": True,
                        # Real vendor info from SerpAPI (only what's actually available)
                        "vendor": {
                            "vendor_name": source_name,
                            "vendor_website": final_url,
                            "is_real_data": True,
                            "data_source": "Google Shopping"
                        }
                    }
                    products.append(product_data)
        
        logger.info(f"SerpAPI returned {len(products)} real products")
        return products[:max_results]
        
    except Exception as e:
        logger.error(f"SerpAPI search failed: {str(e)}")
        return []

def extract_filters_from_real_data(results: List[Dict]) -> Dict[str, Any]:
    """Extract filter options from real product data"""
    sources = list(set([r.get("source", "") for r in results if r.get("source")]))
    
    # Extract price range
    prices = [r.get("price", 0) for r in results if r.get("price", 0) > 0]
    
    return {
        "models": [],  # Real data doesn't have structured model info
        "colors": [],
        "sizes": [],
        "specifications": {},
        "materials": [],
        "brands": [],
        "sources": sources,
        "price_range": {
            "min": min(prices) if prices else 0,
            "max": max(prices) if prices else 0
        },
        "category": "Real Products"
    }

def generate_real_data_analysis(results: List[Dict], query: str, location_data: Dict, currency_info: Dict) -> str:
    """Generate market analysis for real product data from SerpAPI"""
    if not results:
        return "## No Results Found\n\nNo products found matching your search."
    
    prices = [r.get("price", 0) for r in results if r.get("price", 0) > 0]
    if not prices:
        return "## Results Found\n\nProducts found but price data unavailable."
    
    min_price = min(prices)
    max_price = max(prices)
    avg_price = sum(prices) / len(prices)
    
    # Find products with ratings (handle None ratings)
    rated_products = [r for r in results if (r.get("rating") or 0) > 0]
    best_rated = max(rated_products, key=lambda x: x.get("rating") or 0) if rated_products else None
    
    # Find cheapest product
    cheapest = min(results, key=lambda x: x.get("price") or float('inf')) if results else None
    
    # Get unique sources
    sources = list(set([r.get("source", "Unknown") for r in results]))
    
    symbol = currency_info.get("symbol", "₹")
    
    analysis = f"""# Live Prices for: {query}

## 🔴 REAL-TIME DATA from Google Shopping

**Data Source**: Live prices from {len(sources)} marketplace(s)
**Last Updated**: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}

---

## 💰 Price Summary
- **Lowest Price**: {symbol}{min_price:,.0f}
- **Highest Price**: {symbol}{max_price:,.0f}
- **Average Price**: {symbol}{avg_price:,.0f}
- **You Save Up To**: {symbol}{max_price - min_price:,.0f} ({((max_price - min_price) / max_price * 100):.0f}%)

## 📊 Market Analysis

### Where to Buy
Found **{len(results)} listings** from: {', '.join(sources[:5])}{'...' if len(sources) > 5 else ''}

"""
    
    if cheapest:
        analysis += f"""### 💡 Best Price
**{cheapest.get('name', 'Product')[:60]}{'...' if len(cheapest.get('name', '')) > 60 else ''}**
- Price: {symbol}{cheapest.get('price', 0):,.0f}
- From: {cheapest.get('source', 'Unknown')}
- [View Deal]({cheapest.get('source_url', '#')})

"""
    
    if best_rated:
        analysis += f"""### ⭐ Top Rated
**{best_rated.get('name', 'Product')[:60]}{'...' if len(best_rated.get('name', '')) > 60 else ''}**
- Rating: {best_rated.get('rating', 0)}⭐ ({best_rated.get('review_count', 0)} reviews)
- Price: {symbol}{best_rated.get('price', 0):,.0f}
- From: {best_rated.get('source', 'Unknown')}

"""
    
    analysis += """### 🛒 Buying Tips
1. Click "View" to go directly to the seller's page
2. Prices update in real-time from Google Shopping
3. Check seller ratings before purchasing
4. Compare shipping costs at checkout
"""
    
    return analysis

def generate_analysis(results: List[Dict], product_data: Dict, location_data: Dict, currency_info: Dict) -> str:
    """Generate market analysis text"""
    if not results:
        return "## Search Unavailable\n\nNo results found for your query."
    
    prices = [r["price"] for r in results]
    min_price = min(prices)
    max_price = max(prices)
    avg_price = sum(prices) / len(prices)
    
    # Find best value (good rating near average price) - handle None ratings
    best_value = None
    for r in results:
        rating = r.get("rating") or 0
        price = r.get("price") or 0
        if rating >= 4.0 and price > 0 and abs(price - avg_price) < avg_price * 0.3:
            if not best_value or rating > (best_value.get("rating") or 0):
                best_value = r
    
    symbol = currency_info["symbol"]
    
    analysis = f"""# Search Results for: {product_data.get('product_name', 'Product')}

## Price Summary
- **Lowest Price**: {symbol}{min_price:,.2f}
- **Highest Price**: {symbol}{max_price:,.2f}
- **Average Price**: {symbol}{avg_price:,.2f}
- **Price Variation**: {((max_price - min_price) / avg_price * 100):.1f}%

## Market Insights

### Price Distribution
We found **{len(results)} products** across multiple sources in {location_data['city']}, {location_data['country'].upper()}.

### Key Findings
- 📊 **Price Range**: The market shows a {symbol}{max_price - min_price:,.2f} price spread
- 🏪 **Source Diversity**: Products available from Global Suppliers, Local Markets, and Online Marketplaces
- ⭐ **Quality Options**: Multiple highly-rated options (4+ stars) available

### Best Value Recommendation
"""
    
    if best_value:
        analysis += f"""
**{best_value['name']}** from {best_value['source']}
- Price: {symbol}{best_value['price']:,.2f}
- Rating: {best_value['rating']}⭐
- {best_value['description']}

This offers the best balance of quality and price.
"""
    else:
        analysis += "\nCompare options above to find the best value for your needs."
    
    analysis += """
### Buying Tips
1. Compare prices across multiple sources
2. Check seller ratings and reviews
3. Verify warranty and return policies
4. Consider shipping costs and delivery times
"""
    
    return analysis

# ================== API ROUTES ==================

@api_router.get("/")
async def root():
    return {"status": "online", "message": "Universal Product Search API", "version": "1.0.0"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

@api_router.post("/search", response_model=SearchResponse)
async def search_products(request: SearchRequest):
    try:
        query = request.query.strip()
        if not query or len(query) < 2:
            raise HTTPException(status_code=400, detail="Query must be at least 2 characters")
        
        logger.info(f"Processing search query: {query}")
        
        # Extract location from query
        location_data = extract_location(query)
        currency_info = get_currency_info(location_data["country"])
        
        logger.info(f"Location: {location_data}, Currency: {currency_info}")
        
        # Run both searches in parallel: SerpAPI for online prices + OpenStreetMap for local stores
        serpapi_task = search_with_serpapi_enhanced(query, query, location_data["country"], request.max_results)
        places_task = search_local_stores_with_places_api(query, location_data.get("city", ""), 30)  # Get up to 30 local stores
        
        real_results, local_stores = await asyncio.gather(serpapi_task, places_task)
        
        # Get city name for local stores
        local_stores_city = None
        city_info = get_city_from_query(query)
        if city_info:
            local_stores_city = city_info.get("name")
        elif location_data.get("city"):
            local_stores_city = location_data.get("city").title()
        
        if real_results and len(real_results) > 0:
            # We have real data from Google Shopping!
            logger.info(f"Using {len(real_results)} REAL results from SerpAPI + {len(local_stores)} local stores")
            
            all_results = real_results
            
            # Generate analysis for real data
            analysis = generate_real_data_analysis(all_results, query, location_data, currency_info)
            
            # Add local stores info to analysis if available
            if local_stores and len(local_stores) > 0:
                analysis += f"\n\n## 📍 Local Stores in {local_stores_city or 'Your Area'}\n\nFound **{len(local_stores)} local stores** near you. Check the 'Local Stores' tab for addresses, phone numbers, and directions."
            
            # Prepare data sources from real results
            data_sources = []
            seen_sources = set()
            for result in all_results:
                source_name = result.get("source", "Unknown")
                if source_name not in seen_sources:
                    seen_sources.add(source_name)
                    data_sources.append({
                        "name": source_name,
                        "url": result.get("source_url", "").split("?")[0] if result.get("source_url") else "",
                        "type": "Real Marketplace",
                        "description": f"Live prices from {source_name}"
                    })
            
            # Add Foursquare as a data source if we have local stores
            if local_stores:
                data_sources.append({
                    "name": "Foursquare",
                    "url": "https://maps.google.com",
                    "type": "Local Stores",
                    "description": f"Local stores in {local_stores_city or 'your area'}"
                })
            
            # Store search in database (if available)
            if MONGODB_AVAILABLE:
                search_doc = {
                    "id": str(uuid.uuid4()),
                    "query": query,
                    "results_count": len(all_results),
                    "local_stores_count": len(local_stores),
                    "data_source": "serpapi_real",
                    "timestamp": datetime.now(timezone.utc).isoformat()
                }
                await db.searches.insert_one(search_doc)
            
            # For real data, we extract filters from actual results
            available_filters = extract_filters_from_real_data(all_results)
            
            return SearchResponse(
                success=True,
                query=query,
                message=None,
                response=analysis,
                results=all_results,
                results_count=len(all_results),
                ai_model="SerpAPI Google Shopping",
                data_sources=data_sources,
                available_filters=available_filters,
                local_stores=local_stores if local_stores else None,
                local_stores_city=local_stores_city
            )
        
        # No live data found - return empty results with message
        logger.warning(f"No live prices found for: {query}")
        
        return SearchResponse(
            success=True,
            query=query,
            message="No Live Prices Available",
            response=f"""## No Live Prices Found

We couldn't find live pricing data for **"{query}"** from real web sources.

### What This Means
- No current prices available from online marketplaces
- Product may not be widely available online
- Search terms may need refinement

### Suggestions
1. **Try different search terms** - Use generic names or brand names
2. **Check spelling** - Ensure product name is correct
3. **Use broader terms** - Try category names like "kitchen sink" instead of specific models
4. **Manual verification required** - Contact vendors directly for pricing

### For Bulk Upload
If this item is in your bulk upload Excel file, it will show:
- "Live price not available - manual verification required"

**Note**: This system only shows real, live prices from web searches. No estimated or assumed prices are provided.
""",
            results=[],
            results_count=0,
            ai_model="Live Web Search Only",
            data_sources=[{
                "name": "Web Search",
                "url": "",
                "type": "Live Search",
                "description": "Bing, DuckDuckGo, Google"
            }],
            available_filters={},
            local_stores=local_stores if local_stores else None,
            local_stores_city=local_stores_city
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Search error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/recent-searches")
async def get_recent_searches():
    """Get recent searches"""
    if not MONGODB_AVAILABLE:
        return {"searches": []}
    searches = await db.searches.find({}, {"_id": 0}).sort("timestamp", -1).to_list(10)
    return {"searches": searches}

@api_router.post("/similar-products")
async def get_similar_products(request: dict):
    """Get AI-powered similar product suggestions"""
    try:
        product_name = request.get("product_name", "")
        category = request.get("category", "General")
        
        if not product_name:
            return {"similar": [], "recommendations": []}
        
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key or not EMERGENT_AVAILABLE:
            return {"similar": [], "complementary": [], "reasons": {}}
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"similar-products-{uuid.uuid4()}",
            system_message="You are a product recommendation expert. Return ONLY valid JSON."
        )
        chat.with_model("openai", "gpt-4o")
        
        prompt = f"""Based on the product "{product_name}" in category "{category}", suggest:
1. Similar products (alternatives/competitors)
2. Complementary products (often bought together)

Return JSON:
{{
    "similar": ["product1", "product2", "product3", "product4", "product5"],
    "complementary": ["accessory1", "accessory2", "accessory3"],
    "reasons": {{
        "similar": "Why these are good alternatives",
        "complementary": "Why these go well together"
    }}
}}"""

        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        response_text = response.strip()
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(lines[1:-1] if lines[-1] == "```" else lines[1:])
        
        data = json.loads(response_text)
        return data
        
    except Exception as e:
        logger.error(f"Similar products error: {e}")
        return {"similar": [], "complementary": [], "reasons": {}}

@api_router.post("/smart-recommendations")
async def get_smart_recommendations(request: dict):
    """Get personalized recommendations based on search history"""
    try:
        recent_searches = request.get("recent_searches", [])
        current_product = request.get("current_product", "")
        
        if not recent_searches and not current_product:
            return {"recommendations": [], "trending": []}
        
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key or not EMERGENT_AVAILABLE:
            # No API key - return empty results (no static fallback data)
            return {
                "recommendations": [],
                "trending": []
            }
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"recommendations-{uuid.uuid4()}",
            system_message="You are a shopping recommendation AI. Return ONLY valid JSON."
        )
        chat.with_model("openai", "gpt-4o")
        
        search_history = ", ".join(recent_searches[-5:]) if recent_searches else "none"
        
        prompt = f"""Based on:
- Recent searches: {search_history}
- Current product: {current_product or 'none'}

Suggest personalized product recommendations and trending items.

Return JSON:
{{
    "recommendations": [
        {{"name": "Product Name", "reason": "Why recommended", "category": "Category"}},
        {{"name": "Product Name", "reason": "Why recommended", "category": "Category"}}
    ],
    "trending": ["Trending Product 1", "Trending Product 2", "Trending Product 3"]
}}

Provide 3-5 recommendations based on their interests."""

        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        response_text = response.strip()
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(lines[1:-1] if lines[-1] == "```" else lines[1:])
        
        data = json.loads(response_text)
        return data
        
    except Exception as e:
        logger.error(f"Recommendations error: {e}")
        return {"recommendations": [], "trending": []}

# ================== EXCEL BULK PROCESSING ==================

def validate_and_filter_prices(prices_with_sources: list) -> list:
    """
    Filter out unreliable/unrealistic prices using statistical methods.
    Removes outliers that are too low (likely errors, used items, or scams).
    """
    if not prices_with_sources or len(prices_with_sources) < 3:
        return prices_with_sources
    
    prices = [p['price'] for p in prices_with_sources]
    
    # Calculate IQR (Interquartile Range) for outlier detection
    sorted_prices = sorted(prices)
    n = len(sorted_prices)
    q1_idx = n // 4
    q3_idx = (3 * n) // 4
    q1 = sorted_prices[q1_idx]
    q3 = sorted_prices[q3_idx]
    iqr = q3 - q1
    
    # Define bounds - prices below lower bound are likely errors/scams
    # Using 1.5 * IQR below Q1 as lower bound
    lower_bound = max(q1 - 1.5 * iqr, sorted_prices[0] * 0.3)  # At least 30% of min price
    upper_bound = q3 + 1.5 * iqr
    
    # Also filter out prices that are suspiciously low (less than 10% of median)
    median_price = sorted_prices[n // 2]
    min_acceptable = median_price * 0.1  # At least 10% of median
    
    # Filter prices
    filtered = []
    for item in prices_with_sources:
        price = item['price']
        # Keep prices within acceptable range
        if price >= max(lower_bound, min_acceptable) and price <= upper_bound:
            filtered.append(item)
    
    # If filtering removed too many, keep at least the middle 50%
    if len(filtered) < len(prices_with_sources) * 0.3:
        # Fall back to middle 50% of prices
        start_idx = n // 4
        end_idx = (3 * n) // 4
        filtered = [prices_with_sources[i] for i in range(len(prices_with_sources)) 
                   if sorted_prices[start_idx] <= prices_with_sources[i]['price'] <= sorted_prices[end_idx]]
    
    return filtered if filtered else prices_with_sources

def generate_pdf_from_results(results, all_item_sources, total_your_amount, total_market_min_amount,
                               total_market_med_amount, total_market_max_amount,
                               your_grand_total, min_grand_total, med_grand_total, max_grand_total,
                               timestamp):
    """
    Generate a PDF report that matches the Excel format exactly.
    Returns a BytesIO buffer with the PDF content.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2F5496'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph(f"Price Comparison Report - {timestamp}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Prepare data table (first 20 items for main page)
    table_data = []
    
    # Headers
    headers = ['SL#', 'Item', 'Your Rate', 'Qty', 'Your Amt', 
               'Min Rate', 'Min Amt', 'Med Rate', 'Med Amt', 
               'Max Rate', 'Max Amt']
    table_data.append(headers)
    
    # Data rows (showing key comparison data)
    for idx, result in enumerate(results[:20], 1):  # Limit to 20 items for PDF readability
        row = [
            str(result.get('sl_no', idx)),
            result['item'][:30] + '...' if len(result['item']) > 30 else result['item'],
            f"₹{result['user_rate']:,.0f}" if isinstance(result['user_rate'], (int, float)) else str(result['user_rate']),
            str(result['quantity']),
            f"₹{result['user_amount']:,.0f}" if isinstance(result['user_amount'], (int, float)) else str(result['user_amount']),
            f"₹{result.get('market_min_rate', 'N/A'):,.0f}" if isinstance(result.get('market_min_rate'), (int, float)) else str(result.get('market_min_rate', 'N/A')),
            f"₹{result.get('market_min_total', 'N/A'):,.0f}" if isinstance(result.get('market_min_total'), (int, float)) else str(result.get('market_min_total', 'N/A')),
            f"₹{result.get('market_med_rate', 'N/A'):,.0f}" if isinstance(result.get('market_med_rate'), (int, float)) else str(result.get('market_med_rate', 'N/A')),
            f"₹{result.get('market_med_total', 'N/A'):,.0f}" if isinstance(result.get('market_med_total'), (int, float)) else str(result.get('market_med_total', 'N/A')),
            f"₹{result.get('market_max_rate', 'N/A'):,.0f}" if isinstance(result.get('market_max_rate'), (int, float)) else str(result.get('market_max_rate', 'N/A')),
            f"₹{result.get('market_max_total', 'N/A'):,.0f}" if isinstance(result.get('market_max_total'), (int, float)) else str(result.get('market_max_total', 'N/A'))
        ]
        table_data.append(row)
    
    # Create table with styling
    table = Table(table_data, colWidths=[0.4*inch, 2.2*inch, 0.7*inch, 0.4*inch, 0.8*inch,
                                         0.7*inch, 0.8*inch, 0.7*inch, 0.8*inch, 
                                         0.7*inch, 0.8*inch])
    
    # Table styling to match Excel colors
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        # Min columns - light green
        ('BACKGROUND', (5, 1), (6, -1), colors.HexColor('#E2EFDA')),
        # Med columns - light yellow
        ('BACKGROUND', (7, 1), (8, -1), colors.HexColor('#FFF2CC')),
        # Max columns - light orange
        ('BACKGROUND', (9, 1), (10, -1), colors.HexColor('#FCE4D6')),
    ]
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # GST Summary Section
    summary_title = ParagraphStyle(
        'SummaryTitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("CONSOLIDATED GST SUMMARY", summary_title))
    elements.append(Spacer(1, 0.1*inch))
    
    # Calculate GST
    cgst_rate = 0.09
    sgst_rate = 0.09
    
    your_cgst = round(total_your_amount * cgst_rate, 2)
    your_sgst = round(total_your_amount * sgst_rate, 2)
    
    min_cgst = round(total_market_min_amount * cgst_rate, 2)
    min_sgst = round(total_market_min_amount * sgst_rate, 2)
    
    med_cgst = round(total_market_med_amount * cgst_rate, 2)
    med_sgst = round(total_market_med_amount * sgst_rate, 2)
    
    max_cgst = round(total_market_max_amount * cgst_rate, 2)
    max_sgst = round(total_market_max_amount * sgst_rate, 2)
    
    # GST Summary Table
    gst_data = [
        ['', 'YOUR PRICING', 'MARKET MINIMUM', 'MARKET MEDIUM', 'MARKET MAXIMUM'],
        ['Taxable Amount', f"₹{total_your_amount:,.2f}", f"₹{total_market_min_amount:,.2f}", 
         f"₹{total_market_med_amount:,.2f}", f"₹{total_market_max_amount:,.2f}"],
        ['CGST @ 9.0%', f"₹{your_cgst:,.2f}", f"₹{min_cgst:,.2f}", 
         f"₹{med_cgst:,.2f}", f"₹{max_cgst:,.2f}"],
        ['SGST @ 9.0%', f"₹{your_sgst:,.2f}", f"₹{min_sgst:,.2f}", 
         f"₹{med_sgst:,.2f}", f"₹{max_sgst:,.2f}"],
        ['Grand Total', f"₹{your_grand_total:,.2f}", f"₹{min_grand_total:,.2f}", 
         f"₹{med_grand_total:,.2f}", f"₹{max_grand_total:,.2f}"],
    ]
    
    gst_table = Table(gst_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch, 2*inch])
    gst_style = [
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#2F5496')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#548235')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#BF8F00')),
        ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#C65911')),
        ('TEXTCOLOR', (1, 0), (4, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 4), (0, 4), colors.HexColor('#D9EAD3')),
        ('BACKGROUND', (1, 4), (1, 4), colors.HexColor('#D9EAD3')),
        ('BACKGROUND', (2, 4), (2, 4), colors.HexColor('#E2EFDA')),
        ('BACKGROUND', (3, 4), (3, 4), colors.HexColor('#FFF2CC')),
        ('BACKGROUND', (4, 4), (4, 4), colors.HexColor('#FCE4D6')),
    ]
    gst_table.setStyle(TableStyle(gst_style))
    elements.append(gst_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Comparison message
    savings_vs_min = your_grand_total - min_grand_total
    if savings_vs_min > 0:
        comparison_text = f"COMPARISON: Your Grand Total (₹{your_grand_total:,}) is ₹{savings_vs_min:,} MORE than Market Minimum - OVERPAYING"
        comp_color = colors.HexColor('#FFC7CE')
    elif savings_vs_min < 0:
        comparison_text = f"COMPARISON: Your Grand Total (₹{your_grand_total:,}) is ₹{abs(savings_vs_min):,} LESS than Market Minimum - GOOD DEAL"
        comp_color = colors.HexColor('#C6EFCE')
    else:
        comparison_text = f"COMPARISON: Your Grand Total matches Market Minimum"
        comp_color = colors.white
    
    comp_style = ParagraphStyle(
        'Comparison',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        backColor=comp_color,
        borderPadding=10
    )
    elements.append(Paragraph(comparison_text, comp_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Legend
    legend_text = """
    <b>Legend:</b><br/>
    <font color="#006100">Green = Good Deal (Your rate is LOWER than market rate)</font><br/>
    <font color="#9C0006">Red = Overpaying (Your rate is HIGHER than market rate)</font><br/><br/>
    <i>Note: All prices sourced from real web searches. See Excel file for complete details and clickable source links.</i>
    """
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=8,
        spaceAfter=10
    )
    elements.append(Paragraph(legend_text, legend_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

@api_router.post("/bulk-search/upload")
async def bulk_search_upload(file: UploadFile = File(...)):
    """
    Upload an Excel file with product entries and compare with market rates.
    
    Expected Excel format:
    - SL No (Serial Number)
    - Item (Product Name)
    - Rate/Item (Your rate per item)
    - Qty (Quantity)
    - Amount (Your total = Rate × Qty)
    
    Output Excel contains:
    - Your uploaded data (SL No, Item, Your Rate, Qty, Your Amount)
    - Market rates (Min Rate, Medium Rate, Max Rate)
    - Market totals (Min Total, Med Total, Max Total)
    - Comparison (Rate Difference, Amount Difference)
    - Highlighting: Green if you're paying less, Red if you're paying more
    - Website Links, Vendor Details
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
        
        # Read the uploaded file
        contents = await file.read()
        
        # Parse Excel file
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = workbook.active
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
        
        # Extract product entries from Excel
        # First, detect column positions from header row
        products = []
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Find column indices (case-insensitive matching)
        sl_col = None
        item_col = None
        qty_col = None
        rate_col = None  # User's rate per item
        amount_col = None  # User's total amount
        
        for idx, header in enumerate(header_row):
            if header:
                header_lower = str(header).lower().strip()
                if 'sl' in header_lower or 'serial' in header_lower or 's.no' in header_lower:
                    if sl_col is None:
                        sl_col = idx
                elif 'item' in header_lower or 'product' in header_lower or 'description' in header_lower:
                    if item_col is None:
                        item_col = idx
                elif 'rate' in header_lower and 'item' in header_lower:
                    rate_col = idx
                elif 'rate' in header_lower and rate_col is None:
                    rate_col = idx
                elif 'qty' in header_lower or 'quantity' in header_lower:
                    qty_col = idx
                elif 'amount' in header_lower or 'total' in header_lower:
                    amount_col = idx
        
        # Fallback to default positions if headers not found
        if sl_col is None:
            sl_col = 0
        if item_col is None:
            item_col = 1
        if rate_col is None:
            rate_col = 2
        if qty_col is None:
            qty_col = 3
        if amount_col is None:
            amount_col = 4
        
        logger.info(f"Detected columns - SL: {sl_col}, Item: {item_col}, Rate: {rate_col}, Qty: {qty_col}, Amount: {amount_col}")
        logger.info(f"Headers found: {header_row}")
        
        # Now extract data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row and len(row) > item_col and row[item_col]:
                # Get SL No
                sl_no = str(row[sl_col]).strip() if sl_col < len(row) and row[sl_col] else str(row_idx - 1)
                
                # Get Item name
                item_name = str(row[item_col]).strip()
                
                # Get User's Rate per Item
                user_rate = 0
                if rate_col is not None and rate_col < len(row) and row[rate_col] is not None:
                    try:
                        rate_value = row[rate_col]
                        if isinstance(rate_value, (int, float)):
                            user_rate = float(rate_value)
                        else:
                            user_rate = float(str(rate_value).strip().replace(',', ''))
                    except (ValueError, TypeError):
                        user_rate = 0
                
                # Get Quantity
                quantity = 1
                if qty_col is not None and qty_col < len(row) and row[qty_col] is not None:
                    try:
                        qty_value = row[qty_col]
                        if isinstance(qty_value, (int, float)):
                            quantity = int(qty_value)
                        else:
                            quantity = int(float(str(qty_value).strip().replace(',', '')))
                        if quantity < 1:
                            quantity = 1
                    except (ValueError, TypeError):
                        quantity = 1
                
                # Get User's Amount (or calculate from rate × qty)
                user_amount = 0
                if amount_col is not None and amount_col < len(row) and row[amount_col] is not None:
                    try:
                        amount_value = row[amount_col]
                        if isinstance(amount_value, (int, float)):
                            user_amount = float(amount_value)
                        else:
                            user_amount = float(str(amount_value).strip().replace(',', ''))
                    except (ValueError, TypeError):
                        user_amount = user_rate * quantity
                else:
                    user_amount = user_rate * quantity
                
                if item_name and item_name.lower() not in ['item', 'product', 'name', 'description']:
                    logger.info(f"Row {row_idx}: SL={sl_no}, Item={item_name[:30]}..., Rate={user_rate}, Qty={quantity}, Amount={user_amount}")
                    products.append({
                        "row": row_idx,
                        "sl_no": sl_no,
                        "item": item_name,
                        "user_rate": user_rate,
                        "quantity": quantity,
                        "user_amount": user_amount,
                        "query": item_name
                    })
        
        if not products:
            raise HTTPException(status_code=400, detail="No items found in Excel file. Please ensure your Excel has columns for SL No, Item, Rate/Item, Qty, and Amount.")
        
        logger.info(f"Processing {len(products)} items from Excel upload")
        
        # Process products concurrently in batches for better performance
        async def process_single_product(idx, product_info):
            """Process a single product and return its result"""
            try:
                logger.info(f"Processing {idx + 1}/{len(products)}: {product_info['item']} (Qty: {product_info['quantity']})")
                
                # Search using enhanced search (with fallback to free search)
                search_results = await search_with_serpapi_enhanced(
                    query=product_info['query'],
                    original_item=product_info['item'],
                    country="india",
                    max_results=30
                )
                
                user_rate = product_info['user_rate']
                user_amount = product_info['user_amount']
                quantity = product_info['quantity']
                
                if search_results:
                    # Collect prices with their sources for validation
                    prices_with_sources = []
                    for r in search_results:
                        price = r.get('price', 0)
                        if price > 0:
                            prices_with_sources.append({
                                'price': price,
                                'vendor': r.get('source', ''),
                                'website': r.get('source_url', ''),
                                'name': r.get('name', '')
                            })
                    
                    # Validate and filter prices to remove outliers/errors
                    validated_prices = validate_and_filter_prices(prices_with_sources)
                    
                    if validated_prices:
                        prices = [p['price'] for p in validated_prices]
                        min_price = min(prices)
                        max_price = max(prices)
                        sorted_prices = sorted(prices)
                        mid_idx = len(sorted_prices) // 2
                        med_price = sorted_prices[mid_idx] if len(sorted_prices) % 2 == 1 else (sorted_prices[mid_idx - 1] + sorted_prices[mid_idx]) / 2
                        
                        # Calculate market totals based on quantity for ALL rate types
                        min_total = min_price * quantity
                        med_total = med_price * quantity
                        max_total = max_price * quantity
                        
                        # Calculate differences (Your Amount - Market Min Amount)
                        # Positive = you're paying MORE than market (bad)
                        # Negative = you're paying LESS than market (good deal)
                        # Calculate differences for ALL rate types (Min, Med, Max)
                        rate_diff_min = user_rate - min_price if user_rate > 0 else 0
                        amount_diff_min = user_amount - min_total if user_amount > 0 else 0
                        
                        rate_diff_med = user_rate - med_price if user_rate > 0 else 0
                        amount_diff_med = user_amount - med_total if user_amount > 0 else 0
                        
                        rate_diff_max = user_rate - max_price if user_rate > 0 else 0
                        amount_diff_max = user_amount - max_total if user_amount > 0 else 0
                        
                        # Get vendor/website info from validated results only
                        vendors = []
                        websites = []
                        
                        for item in validated_prices:
                            vendor = item.get('vendor', '')
                            website = item.get('website', '')
                            if vendor and vendor not in vendors:
                                vendors.append(vendor)
                            if website and website not in websites:
                                websites.append(website)
                        
                        # Find specific source for each price point (min, med, max)
                        sorted_validated = sorted(validated_prices, key=lambda x: x['price'])
                        min_source_item = sorted_validated[0]
                        max_source_item = sorted_validated[-1]
                        mid_idx_src = len(sorted_validated) // 2
                        med_source_item = sorted_validated[mid_idx_src]
                        
                        return {
                            "sl_no": product_info['sl_no'],
                            "item": product_info['item'],
                            "user_rate": round(user_rate, 2),
                            "quantity": quantity,
                            "user_amount": round(user_amount, 2),
                            "market_min_rate": round(min_price, 2),
                            "market_min_total": round(min_total, 2),
                            "min_source": min_source_item.get('vendor', 'Unknown'),
                            "min_url": min_source_item.get('website', ''),
                            "rate_diff_min": round(rate_diff_min, 2),
                            "amount_diff_min": round(amount_diff_min, 2),
                            "market_med_rate": round(med_price, 2),
                            "market_med_total": round(med_total, 2),
                            "med_source": med_source_item.get('vendor', 'Unknown'),
                            "med_url": med_source_item.get('website', ''),
                            "rate_diff_med": round(rate_diff_med, 2),
                            "amount_diff_med": round(amount_diff_med, 2),
                            "market_max_rate": round(max_price, 2),
                            "market_max_total": round(max_total, 2),
                            "max_source": max_source_item.get('vendor', 'Unknown'),
                            "max_url": max_source_item.get('website', ''),
                            "rate_diff_max": round(rate_diff_max, 2),
                            "amount_diff_max": round(amount_diff_max, 2),
                            "website_links": "\n".join(websites[:5]),
                            "vendor_details": ", ".join(vendors[:10]),
                            "all_sources": validated_prices,  # For Sources sheet
                            "search_timestamp": datetime.now().isoformat(),
                            "price_adjusted": False,
                            "adjustment_note": ""
                        }
                    else:
                        # No validated prices
                        return {
                            "sl_no": product_info['sl_no'],
                            "item": product_info['item'],
                            "user_rate": round(user_rate, 2),
                            "quantity": quantity,
                            "user_amount": round(user_amount, 2),
                            "market_min_rate": "Live price not available",
                            "market_min_total": "Live price not available",
                            "rate_diff_min": "N/A",
                            "amount_diff_min": "N/A",
                            "market_med_rate": "Live price not available",
                            "market_med_total": "Live price not available",
                            "rate_diff_med": "N/A",
                            "amount_diff_med": "N/A",
                            "market_max_rate": "Live price not available",
                            "market_max_total": "Live price not available",
                            "rate_diff_max": "N/A",
                            "amount_diff_max": "N/A",
                            "website_links": "Live price not available - manual verification required",
                            "vendor_details": "Live price not available - manual verification required"
                        }
                else:
                    # No search results
                    return {
                        "sl_no": product_info['sl_no'],
                        "item": product_info['item'],
                        "user_rate": round(user_rate, 2),
                        "quantity": quantity,
                        "user_amount": round(user_amount, 2),
                        "market_min_rate": "Live price not available",
                        "market_min_total": "Live price not available",
                        "rate_diff_min": "N/A",
                        "amount_diff_min": "N/A",
                        "market_med_rate": "Live price not available",
                        "market_med_total": "Live price not available",
                        "rate_diff_med": "N/A",
                        "amount_diff_med": "N/A",
                        "market_max_rate": "Live price not available",
                        "market_max_total": "Live price not available",
                        "rate_diff_max": "N/A",
                        "amount_diff_max": "N/A",
                        "website_links": "Live price not available - manual verification required",
                        "vendor_details": "Live price not available - manual verification required"
                    }
                
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Error processing item {product_info['item']}: {error_msg}")
                
                return {
                    "sl_no": product_info['sl_no'],
                    "item": product_info['item'],
                    "user_rate": product_info.get('user_rate', 0),
                    "quantity": product_info.get('quantity', 1),
                    "user_amount": product_info.get('user_amount', 0),
                    "market_min_rate": "Error",
                    "market_min_total": "Error",
                    "rate_diff_min": "Error",
                    "amount_diff_min": "Error",
                    "market_med_rate": "Error",
                    "market_med_total": "Error",
                    "rate_diff_med": "Error",
                    "amount_diff_med": "Error",
                    "market_max_rate": "Error",
                    "market_max_total": "Error",
                    "rate_diff_max": "Error",
                    "amount_diff_max": "Error",
                    "website_links": f"Error: {error_msg}",
                    "vendor_details": "Error"
                }
        
        # Process products in batches concurrently (5 at a time to avoid overwhelming the system)
        results = []
        batch_size = 5
        
        for i in range(0, len(products), batch_size):
            batch = products[i:i + batch_size]
            batch_tasks = [process_single_product(idx + i, product_info) for idx, product_info in enumerate(batch)]
            batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
            
            # Handle results and exceptions
            for result in batch_results:
                if isinstance(result, Exception):
                    logger.error(f"Batch processing error: {result}")
                else:
                    results.append(result)
            
            # Small delay between batches to avoid rate limiting
            if i + batch_size < len(products):
                await asyncio.sleep(0.5)
        
        logger.info(f"Completed processing {len(results)} items")
        
        # Generate output Excel with consolidated GST summary
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Price Comparison"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # Colors for comparison highlighting
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Good deal
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Overpaying
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")
        summary_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # Summary section
        min_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Min section - light green
        med_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Med section - light yellow
        max_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Max section - light orange
        adjusted_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Lavender for adjusted prices
        
        # Headers with Min, Med, Max rate columns + combined Website Links & Vendor Details
        headers = [
            "SL No", "Item", 
            "Your Rate (₹)", "Qty", "Your Amount (₹)",
            "Min Rate (₹)", "Min Amount (₹)", "Rate Diff (Min) (₹)", "Amount Diff (Min) (₹)",
            "Med Rate (₹)", "Med Amount (₹)", "Rate Diff (Med) (₹)", "Amount Diff (Med) (₹)",
            "Max Rate (₹)", "Max Amount (₹)", "Rate Diff (Max) (₹)", "Amount Diff (Max) (₹)",
            "Website Links", "Vendor Details"
        ]
        
        # Write column headers (row 1)
        for col_idx, header in enumerate(headers, start=1):
            cell = output_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            # Color-code market rate columns
            if 'Min' in header and 'Website' not in header and 'Vendor' not in header:
                cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            elif 'Med' in header and 'Website' not in header and 'Vendor' not in header:
                cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
            elif 'Max' in header and 'Website' not in header and 'Vendor' not in header:
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        
        # Calculate totals while writing data
        total_your_amount = 0
        total_market_min_amount = 0
        total_market_med_amount = 0
        total_market_max_amount = 0
        
        # Collect all sources for Sources sheet
        all_item_sources = []
        
        # Write data (starting row 2)
        for row_idx, result in enumerate(results, start=2):
            # Accumulate totals
            if isinstance(result['user_amount'], (int, float)):
                total_your_amount += result['user_amount']
            if isinstance(result.get('market_min_total'), (int, float)):
                total_market_min_amount += result['market_min_total']
            if isinstance(result.get('market_med_total'), (int, float)):
                total_market_med_amount += result['market_med_total']
            if isinstance(result.get('market_max_total'), (int, float)):
                total_market_max_amount += result['market_max_total']
            
            # Collect sources for Sources sheet
            if result.get('all_sources'):
                for src in result['all_sources']:
                    all_item_sources.append({
                        'item': result['item'],
                        'price': src.get('price', 0),
                        'vendor': src.get('vendor', 'Unknown'),
                        'url': src.get('website', ''),
                        'timestamp': result.get('search_timestamp', '')
                    })
            
            data = [
                result['sl_no'],                              # A - SL No
                result['item'],                               # B - Item
                result['user_rate'],                          # C - Your Rate
                result['quantity'],                           # D - Qty
                result['user_amount'],                        # E - Your Amount
                result.get('market_min_rate', 'N/A'),         # F - Min Rate
                result.get('market_min_total', 'N/A'),        # G - Min Amount
                result.get('rate_diff_min', 'N/A'),           # H - Rate Diff (Min)
                result.get('amount_diff_min', 'N/A'),         # I - Amount Diff (Min)
                result.get('market_med_rate', 'N/A'),         # J - Med Rate
                result.get('market_med_total', 'N/A'),        # K - Med Amount
                result.get('rate_diff_med', 'N/A'),           # L - Rate Diff (Med)
                result.get('amount_diff_med', 'N/A'),         # M - Amount Diff (Med)
                result.get('market_max_rate', 'N/A'),         # N - Max Rate
                result.get('market_max_total', 'N/A'),        # O - Max Amount
                result.get('rate_diff_max', 'N/A'),           # P - Rate Diff (Max)
                result.get('amount_diff_max', 'N/A'),         # Q - Amount Diff (Max)
                result.get('website_links', ''),              # R - Website Links (combined)
                result.get('vendor_details', '')              # S - Vendor Details (combined)
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = output_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Number columns - right align (all numeric columns)
                if col_idx in [3, 4, 5, 6, 7, 9, 10, 11, 12, 14, 15, 16, 17, 19, 20]:
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [8, 13, 18, 218, 9, 10, 11, 12, 13, 14, 15, 16, 17]:
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [18, 19]:  # Website Links and Vendor Details - wrap text
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Color-code market rate columns with light backgrounds
                if col_idx in [6, 7]:  # Min Rate & Amount columns
                    if isinstance(value, (int, float)):
                        cell.fill = min_fill
                elif col_idx in [10, 11]:  # Med Rate & Amount columns
                    if isinstance(value, (int, float)):
                        cell.fill = med_fill
                elif col_idx in [14, 15]:  # Max Rate & Amount columns
                    if isinstance(value, (int, float)):
                        cell.fill = max_fill
                
                # Apply color highlighting for ALL difference columns
                # Min Diff (H, I), Med Diff (L, M), Max Diff (P, Q)
                if col_idx in [8, 9, 12, 13, 16, 17]:
                    if isinstance(value, (int, float)):
                        if value > 0:  # Positive = You're paying MORE than market (Overpaying)
                            cell.fill = red_fill
                            cell.font = red_font
                        elif value < 0:  # Negative = You're paying LESS than market (Good deal)
                            cell.fill = green_fill
                            cell.font = green_font
            
            # Flag row highlighting
            user_rate = result.get('user_rate', 0)
            min_rate = result.get('market_min_rate', 0)
            max_rate = result.get('market_max_rate', 0)
            
            if isinstance(user_rate, (int, float)) and isinstance(min_rate, (int, float)):
                if user_rate < min_rate:
                    # Good deal - highlight row lightly
                    for col in range(1, 20):  # Updated column count (19 columns total)
                        cell = output_sheet.cell(row=row_idx, column=col)
                        if not cell.fill.start_color.index or cell.fill.start_color.index == '00000000':
                            cell.fill = green_fill
            
            if isinstance(user_rate, (int, float)) and isinstance(max_rate, (int, float)):
                if user_rate > max_rate:
                    # Bad - overpaying beyond max
                    for col in [2]:  # Highlight item name
                        cell = output_sheet.cell(row=row_idx, column=col)
                        cell.fill = red_fill
                        cell.font = Font(bold=True, color="9C0006")
                
        # Adjust column widths (19 columns now - removed 3 separate source columns, kept 2 combined)
        column_widths = [8, 30, 12, 6, 14, 12, 13, 12, 14, 12, 13, 12, 14, 12, 13, 12, 14, 40, 30]
        
        # Set row heights
        output_sheet.row_dimensions[1].height = 40  # Header row (taller for wrapped text)
        for row_idx in range(2, len(results) + 2):
            output_sheet.row_dimensions[row_idx].height = 45
        
        # ========== CONSOLIDATED GST SUMMARY SECTION ==========
        summary_start_row = len(results) + 4
        
        # Calculate GST on totals
        cgst_rate = 0.09
        sgst_rate = 0.09
        
        # Your pricing GST
        your_cgst = round(total_your_amount * cgst_rate, 2)
        your_sgst = round(total_your_amount * sgst_rate, 2)
        your_subtotal = total_your_amount + your_cgst + your_sgst
        your_round_off = round(your_subtotal) - your_subtotal
        your_grand_total = round(your_subtotal)
        
        # Market MINIMUM pricing GST
        min_cgst = round(total_market_min_amount * cgst_rate, 2)
        min_sgst = round(total_market_min_amount * sgst_rate, 2)
        min_subtotal = total_market_min_amount + min_cgst + min_sgst
        min_round_off = round(min_subtotal) - min_subtotal
        min_grand_total = round(min_subtotal)
        
        # Market MEDIUM pricing GST
        med_cgst = round(total_market_med_amount * cgst_rate, 2)
        med_sgst = round(total_market_med_amount * sgst_rate, 2)
        med_subtotal = total_market_med_amount + med_cgst + med_sgst
        med_round_off = round(med_subtotal) - med_subtotal
        med_grand_total = round(med_subtotal)
        
        # Market MAXIMUM pricing GST
        max_cgst = round(total_market_max_amount * cgst_rate, 2)
        max_sgst = round(total_market_max_amount * sgst_rate, 2)
        max_subtotal = total_market_max_amount + max_cgst + max_sgst
        max_round_off = round(max_subtotal) - max_subtotal
        max_grand_total = round(max_subtotal)
        
        # Difference (vs minimum)
        savings_vs_min = your_grand_total - min_grand_total
        
        # Summary Header
        output_sheet.merge_cells(f'A{summary_start_row}:M{summary_start_row}')
        summary_header = output_sheet.cell(row=summary_start_row, column=1, value="CONSOLIDATED GST SUMMARY")
        summary_header.font = Font(bold=True, size=14, color="FFFFFF")
        summary_header.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        summary_header.alignment = Alignment(horizontal="center", vertical="center")
        
        # Four-column summary layout (using wider spacing)
        summary_row = summary_start_row + 2
        
        # Column headers for summary - using columns A-B, D-E, G-H, J-K
        summary_headers = [
            ("A", "B", "YOUR PRICING", PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")),
            ("D", "E", "MARKET MINIMUM", PatternFill(start_color="548235", end_color="548235", fill_type="solid")),
            ("G", "H", "MARKET MEDIUM", PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")),
            ("J", "K", "MARKET MAXIMUM", PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")),
        ]
        
        for start_col, end_col, title, fill_color in summary_headers:
            output_sheet.merge_cells(f'{start_col}{summary_row}:{end_col}{summary_row}')
            header_cell = output_sheet[f'{start_col}{summary_row}']
            header_cell.value = title
            header_cell.font = Font(bold=True, size=11, color="FFFFFF")
            header_cell.fill = fill_color
            header_cell.alignment = Alignment(horizontal="center")
        
        # Summary data
        summary_labels = ["Taxable Amount", "CGST @ 9.0%", "SGST @ 9.0%", "Round Off", "Grand Total"]
        your_values = [total_your_amount, your_cgst, your_sgst, your_round_off, your_grand_total]
        min_values = [total_market_min_amount, min_cgst, min_sgst, min_round_off, min_grand_total]
        med_values = [total_market_med_amount, med_cgst, med_sgst, med_round_off, med_grand_total]
        max_values = [total_market_max_amount, max_cgst, max_sgst, max_round_off, max_grand_total]
        
        for i, label in enumerate(summary_labels):
            row = summary_row + 1 + i
            is_grand_total = (label == "Grand Total")
            
            # Your Pricing (Col A-B)
            output_sheet.cell(row=row, column=1, value=label).border = thin_border
            val_cell = output_sheet.cell(row=row, column=2, value=f"₹{your_values[i]:,.2f}")
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal="right")
            if is_grand_total:
                output_sheet.cell(row=row, column=1).font = Font(bold=True)
                val_cell.font = Font(bold=True)
                val_cell.fill = summary_fill
            
            # Market Min (Col D-E)
            output_sheet.cell(row=row, column=4, value=label).border = thin_border
            val_cell = output_sheet.cell(row=row, column=5, value=f"₹{min_values[i]:,.2f}")
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal="right")
            if is_grand_total:
                output_sheet.cell(row=row, column=4).font = Font(bold=True)
                val_cell.font = Font(bold=True)
                val_cell.fill = min_fill
            
            # Market Med (Col G-H)
            output_sheet.cell(row=row, column=7, value=label).border = thin_border
            val_cell = output_sheet.cell(row=row, column=8, value=f"₹{med_values[i]:,.2f}")
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal="right")
            if is_grand_total:
                output_sheet.cell(row=row, column=7).font = Font(bold=True)
                val_cell.font = Font(bold=True)
                val_cell.fill = med_fill
            
            # Market Max (Col J-K)
            output_sheet.cell(row=row, column=10, value=label).border = thin_border
            val_cell = output_sheet.cell(row=row, column=11, value=f"₹{max_values[i]:,.2f}")
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal="right")
            if is_grand_total:
                output_sheet.cell(row=row, column=10).font = Font(bold=True)
                val_cell.font = Font(bold=True)
                val_cell.fill = max_fill
        
        # Comparison / Savings section
        comparison_row = summary_row + 8
        output_sheet.merge_cells(f'A{comparison_row}:M{comparison_row}')
        if savings_vs_min > 0:
            comparison_text = f"COMPARISON: Your Grand Total (₹{your_grand_total:,}) is ₹{savings_vs_min:,} MORE than Market Minimum (₹{min_grand_total:,}) - OVERPAYING"
            comparison_cell = output_sheet.cell(row=comparison_row, column=1, value=comparison_text)
            comparison_cell.fill = red_fill
            comparison_cell.font = Font(bold=True, color="9C0006")
        elif savings_vs_min < 0:
            comparison_text = f"COMPARISON: Your Grand Total (₹{your_grand_total:,}) is ₹{abs(savings_vs_min):,} LESS than Market Minimum (₹{min_grand_total:,}) - GOOD DEAL"
            comparison_cell = output_sheet.cell(row=comparison_row, column=1, value=comparison_text)
            comparison_cell.fill = green_fill
            comparison_cell.font = Font(bold=True, color="006100")
        else:
            comparison_text = f"COMPARISON: Your Grand Total (₹{your_grand_total:,}) matches Market Minimum (₹{min_grand_total:,})"
            comparison_cell = output_sheet.cell(row=comparison_row, column=1, value=comparison_text)
            comparison_cell.font = Font(bold=True)
        comparison_cell.alignment = Alignment(horizontal="center")
        comparison_cell.border = thick_border
        
        # Legend
        legend_row = comparison_row + 3
        output_sheet.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
        
        legend_green = output_sheet.cell(row=legend_row + 1, column=1, value="Green = Good Deal (Your rate is LOWER than market rate)")
        legend_green.fill = green_fill
        legend_green.font = green_font
        
        legend_red = output_sheet.cell(row=legend_row + 2, column=1, value="Red = Overpaying (Your rate is HIGHER than market rate)")
        legend_red.fill = red_fill
        legend_red.font = red_font
        
        output_sheet.cell(row=legend_row + 4, column=1, value="Note: All prices sourced from real web searches. Rate Diff and Amount Diff columns show comparison for each rate type (Min, Med, Max).").font = Font(italic=True)
        output_sheet.cell(row=legend_row + 5, column=1, value="See 'Sources' sheet for detailed source URLs and timestamps.").font = Font(italic=True)
        
        # ========== CREATE SOURCES SHEET FOR TRACEABILITY ==========
        sources_sheet = output_workbook.create_sheet(title="Sources")
        
        # Sources sheet headers
        sources_headers = ["Item", "Detected Price (₹)", "Vendor/Source", "URL", "Timestamp"]
        for col_idx, header in enumerate(sources_headers, start=1):
            cell = sources_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write all sources data
        source_row = 2
        for source in all_item_sources:
            sources_sheet.cell(row=source_row, column=1, value=source.get('item', '')).border = thin_border
            sources_sheet.cell(row=source_row, column=2, value=source.get('price', 0)).border = thin_border
            sources_sheet.cell(row=source_row, column=3, value=source.get('vendor', 'Unknown')).border = thin_border
            url_cell = sources_sheet.cell(row=source_row, column=4, value=source.get('url', ''))
            url_cell.border = thin_border
            url_cell.alignment = Alignment(wrap_text=True)
            sources_sheet.cell(row=source_row, column=5, value=source.get('timestamp', '')).border = thin_border
            source_row += 1
        
        # Adjust sources sheet column widths
        sources_sheet.column_dimensions['A'].width = 35
        sources_sheet.column_dimensions['B'].width = 15
        sources_sheet.column_dimensions['C'].width = 20
        sources_sheet.column_dimensions['D'].width = 60
        sources_sheet.column_dimensions['E'].width = 25
        
        # Save Excel to BytesIO
        excel_buffer = io.BytesIO()
        output_workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"PriceComparison_Results_{timestamp}.xlsx"
        pdf_filename = f"PriceComparison_Results_{timestamp}.pdf"
        zip_filename = f"PriceComparison_Results_{timestamp}.zip"
        
        # Generate PDF version
        pdf_buffer = generate_pdf_from_results(results, all_item_sources, 
                                                total_your_amount, total_market_min_amount,
                                                total_market_med_amount, total_market_max_amount,
                                                your_grand_total, min_grand_total, med_grand_total, max_grand_total,
                                                timestamp)
        
        # Create ZIP file containing both Excel and PDF
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel file
            zip_file.writestr(excel_filename, excel_buffer.getvalue())
            # Add PDF file
            zip_file.writestr(pdf_filename, pdf_buffer.getvalue())
        
        zip_buffer.seek(0)
        
        logger.info(f"Processing complete. Generated {excel_filename} and {pdf_filename} with {len(results)} results and {len(all_item_sources)} source entries.")
        
        # Return ZIP file containing both Excel and PDF
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={zip_filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk search error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@api_router.get("/bulk-search/template")
async def download_template():
    """Download a sample Excel template for bulk search with SL No, Item and Quantity columns"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items"
    
    # Headers - matching user's format with Quantity
    headers = ["SL No", "Item", "Quantity"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # No sample data - user must provide their own products
    # Template contains only headers for user to fill in
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 12
    
    # Add instructions
    sheet.cell(row=9, column=1, value="Instructions:").font = Font(bold=True)
    sheet.cell(row=10, column=1, value="1. Add serial numbers in Column A (SL No)")
    sheet.cell(row=11, column=1, value="2. Add item/product names in Column B (Item)")
    sheet.cell(row=12, column=1, value="3. Add quantity in Column C (Quantity) - defaults to 1 if not provided")
    sheet.cell(row=13, column=1, value="4. Save and upload this file to get price results")
    
    # Save to BytesIO
    output_buffer = io.BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)
    
    return StreamingResponse(
        output_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=PriceSearch_Template.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
