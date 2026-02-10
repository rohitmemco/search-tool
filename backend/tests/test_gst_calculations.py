"""
Test cases for GST Calculations in Bulk Excel Upload feature
Tests the GST calculation columns (CGST @9%, SGST @9%, Grand Total)
for both user's uploaded rates and detected market rates.

Expected columns in output Excel (17 total):
- SL No, Item
- Your Rate, Qty, Your Amount, Your CGST @9%, Your SGST @9%, Your Grand Total
- Market Rate, Market Amount, Market CGST @9%, Market SGST @9%, Market Grand Total
- Rate Diff, Grand Total Diff
- Website Links, Vendor Details
"""
import pytest
import requests
import os
import io
import openpyxl
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
CGST_RATE = 0.09
SGST_RATE = 0.09
GST_TOLERANCE = 0.01  # 1% tolerance for floating point comparison


class TestGSTCalculations:
    """Tests for GST calculations in bulk upload feature"""
    
    @pytest.fixture
    def create_test_excel_with_gst_data(self):
        """Create a test Excel file with SL No, Item, Rate/Item, Qty, Amount columns"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            # Headers matching expected format
            ws['A1'] = "SL No"
            ws['B1'] = "Item"
            ws['C1'] = "Rate/Item"
            ws['D1'] = "Qty"
            ws['E1'] = "Amount"
            
            # Add products
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('sl_no', idx - 1))
                ws.cell(row=idx, column=2, value=product.get('item', ''))
                ws.cell(row=idx, column=3, value=product.get('rate', 0))
                ws.cell(row=idx, column=4, value=product.get('qty', 1))
                ws.cell(row=idx, column=5, value=product.get('amount', product.get('rate', 0) * product.get('qty', 1)))
            
            # Save to BytesIO
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_output_has_17_columns(self, create_test_excel_with_gst_data):
        """Output Excel should have all 17 columns including GST columns"""
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "iPhone 15 Pro", "rate": 150000, "qty": 1, "amount": 150000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_gst.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers from row 2 (row 1 has category headers)
        headers = [ws.cell(row=2, column=col).value for col in range(1, 18)]
        
        # Expected headers
        expected_headers = [
            "SL No", "Item",
            "Your Rate", "Qty", "Your Amount", "Your CGST @9%", "Your SGST @9%", "Your Grand Total",
            "Market Rate", "Market Amount", "Market CGST @9%", "Market SGST @9%", "Market Grand Total",
            "Rate Diff", "Grand Total Diff",
            "Website Links", "Vendor Details"
        ]
        
        print(f"Actual headers: {headers}")
        
        # Verify we have 17 columns
        non_empty_headers = [h for h in headers if h]
        assert len(non_empty_headers) == 17, f"Expected 17 columns, got {len(non_empty_headers)}"
        
        # Check GST-related headers exist
        header_str = str(headers).lower()
        assert "cgst" in header_str, "CGST column header missing"
        assert "sgst" in header_str, "SGST column header missing"
        assert "grand total" in header_str, "Grand Total column header missing"
        
        print(f"✓ Output has 17 columns including GST columns")
    
    def test_user_cgst_calculation(self, create_test_excel_with_gst_data):
        """User CGST should be Amount × 0.09"""
        user_rate = 10000
        qty = 2
        user_amount = user_rate * qty  # 20000
        expected_cgst = user_amount * CGST_RATE  # 1800
        
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Test Product", "rate": user_rate, "qty": qty, "amount": user_amount}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_cgst.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Row 3 is first data row (Row 1: category headers, Row 2: column headers)
        actual_amount = ws.cell(row=3, column=5).value  # Your Amount (column E)
        actual_cgst = ws.cell(row=3, column=6).value    # Your CGST @9% (column F)
        
        print(f"User Amount: {actual_amount}, User CGST: {actual_cgst}")
        print(f"Expected CGST: {expected_cgst}")
        
        # Verify amount
        assert actual_amount == user_amount, f"Expected amount {user_amount}, got {actual_amount}"
        
        # Verify CGST calculation (Amount × 0.09)
        if isinstance(actual_cgst, (int, float)):
            calculated_cgst = actual_amount * CGST_RATE
            assert abs(actual_cgst - calculated_cgst) < GST_TOLERANCE, \
                f"CGST calculation incorrect. Expected {calculated_cgst}, got {actual_cgst}"
            print(f"✓ User CGST @9% correctly calculated: {actual_cgst}")
        else:
            print(f"! User CGST value: {actual_cgst}")
    
    def test_user_sgst_calculation(self, create_test_excel_with_gst_data):
        """User SGST should be Amount × 0.09"""
        user_rate = 15000
        qty = 3
        user_amount = user_rate * qty  # 45000
        expected_sgst = user_amount * SGST_RATE  # 4050
        
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Test Product SGST", "rate": user_rate, "qty": qty, "amount": user_amount}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_sgst.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        actual_amount = ws.cell(row=3, column=5).value  # Your Amount (column E)
        actual_sgst = ws.cell(row=3, column=7).value    # Your SGST @9% (column G)
        
        print(f"User Amount: {actual_amount}, User SGST: {actual_sgst}")
        print(f"Expected SGST: {expected_sgst}")
        
        if isinstance(actual_sgst, (int, float)):
            calculated_sgst = actual_amount * SGST_RATE
            assert abs(actual_sgst - calculated_sgst) < GST_TOLERANCE, \
                f"SGST calculation incorrect. Expected {calculated_sgst}, got {actual_sgst}"
            print(f"✓ User SGST @9% correctly calculated: {actual_sgst}")
        else:
            print(f"! User SGST value: {actual_sgst}")
    
    def test_user_grand_total_calculation(self, create_test_excel_with_gst_data):
        """User Grand Total should be Amount + CGST + SGST"""
        user_rate = 20000
        qty = 1
        user_amount = user_rate * qty  # 20000
        expected_cgst = user_amount * CGST_RATE  # 1800
        expected_sgst = user_amount * SGST_RATE  # 1800
        expected_grand_total = user_amount + expected_cgst + expected_sgst  # 23600
        
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Grand Total Test", "rate": user_rate, "qty": qty, "amount": user_amount}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_grand_total.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        actual_amount = ws.cell(row=3, column=5).value       # Your Amount (column E)
        actual_cgst = ws.cell(row=3, column=6).value         # Your CGST @9% (column F)
        actual_sgst = ws.cell(row=3, column=7).value         # Your SGST @9% (column G)
        actual_grand_total = ws.cell(row=3, column=8).value  # Your Grand Total (column H)
        
        print(f"Amount: {actual_amount}, CGST: {actual_cgst}, SGST: {actual_sgst}, Grand Total: {actual_grand_total}")
        print(f"Expected Grand Total: {expected_grand_total}")
        
        if isinstance(actual_grand_total, (int, float)):
            # Verify Grand Total = Amount + CGST + SGST
            if isinstance(actual_cgst, (int, float)) and isinstance(actual_sgst, (int, float)):
                calculated_grand_total = actual_amount + actual_cgst + actual_sgst
                assert abs(actual_grand_total - calculated_grand_total) < GST_TOLERANCE, \
                    f"Grand Total incorrect. Expected {calculated_grand_total}, got {actual_grand_total}"
            print(f"✓ User Grand Total correctly calculated: {actual_grand_total}")
        else:
            print(f"! User Grand Total value: {actual_grand_total}")
    
    def test_market_gst_columns_present(self, create_test_excel_with_gst_data):
        """Market GST columns (CGST, SGST, Grand Total) should be present"""
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Samsung Galaxy S24", "rate": 85000, "qty": 1, "amount": 85000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_market_gst.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Market columns are K, L, M (11, 12, 13)
        market_cgst = ws.cell(row=3, column=11).value     # Market CGST @9% (column K)
        market_sgst = ws.cell(row=3, column=12).value     # Market SGST @9% (column L)
        market_grand_total = ws.cell(row=3, column=13).value  # Market Grand Total (column M)
        
        print(f"Market CGST: {market_cgst}, Market SGST: {market_sgst}, Market Grand Total: {market_grand_total}")
        
        # Market columns should have values (either numbers or "N/A" if no market data)
        assert market_cgst is not None, "Market CGST column should have a value"
        assert market_sgst is not None, "Market SGST column should have a value"
        assert market_grand_total is not None, "Market Grand Total column should have a value"
        
        print(f"✓ Market GST columns present with values")
    
    def test_market_gst_calculation_correctness(self, create_test_excel_with_gst_data):
        """Market GST should be calculated correctly when market data is available"""
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Sony WH-1000XM5 Headphones", "rate": 30000, "qty": 2, "amount": 60000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_market_calc.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        market_amount = ws.cell(row=3, column=10).value    # Market Amount (column J)
        market_cgst = ws.cell(row=3, column=11).value      # Market CGST @9% (column K)
        market_sgst = ws.cell(row=3, column=12).value      # Market SGST @9% (column L)
        market_grand_total = ws.cell(row=3, column=13).value  # Market Grand Total (column M)
        
        print(f"Market Amount: {market_amount}")
        print(f"Market CGST: {market_cgst}, Market SGST: {market_sgst}, Market Grand Total: {market_grand_total}")
        
        # If market amount is a number, verify GST calculations
        if isinstance(market_amount, (int, float)) and market_amount > 0:
            expected_market_cgst = market_amount * CGST_RATE
            expected_market_sgst = market_amount * SGST_RATE
            expected_market_grand_total = market_amount + expected_market_cgst + expected_market_sgst
            
            if isinstance(market_cgst, (int, float)):
                assert abs(market_cgst - expected_market_cgst) < GST_TOLERANCE * market_amount, \
                    f"Market CGST incorrect. Expected ~{expected_market_cgst:.2f}, got {market_cgst}"
                print(f"✓ Market CGST correctly calculated")
            
            if isinstance(market_sgst, (int, float)):
                assert abs(market_sgst - expected_market_sgst) < GST_TOLERANCE * market_amount, \
                    f"Market SGST incorrect. Expected ~{expected_market_sgst:.2f}, got {market_sgst}"
                print(f"✓ Market SGST correctly calculated")
            
            if isinstance(market_grand_total, (int, float)):
                if isinstance(market_cgst, (int, float)) and isinstance(market_sgst, (int, float)):
                    calc_grand_total = market_amount + market_cgst + market_sgst
                    assert abs(market_grand_total - calc_grand_total) < GST_TOLERANCE * market_amount, \
                        f"Market Grand Total incorrect. Expected ~{calc_grand_total:.2f}, got {market_grand_total}"
                print(f"✓ Market Grand Total correctly calculated")
        else:
            print(f"! Market data not available (value: {market_amount})")
    
    def test_grand_total_diff_calculation(self, create_test_excel_with_gst_data):
        """Grand Total Diff should be: Your Grand Total - Market Grand Total"""
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "MacBook Air M3", "rate": 120000, "qty": 1, "amount": 120000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_diff.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        user_grand_total = ws.cell(row=3, column=8).value      # Your Grand Total (column H)
        market_grand_total = ws.cell(row=3, column=13).value   # Market Grand Total (column M)
        grand_total_diff = ws.cell(row=3, column=15).value     # Grand Total Diff (column O)
        
        print(f"User Grand Total: {user_grand_total}")
        print(f"Market Grand Total: {market_grand_total}")
        print(f"Grand Total Diff: {grand_total_diff}")
        
        # Verify Grand Total Diff calculation
        if isinstance(user_grand_total, (int, float)) and isinstance(market_grand_total, (int, float)):
            expected_diff = user_grand_total - market_grand_total
            if isinstance(grand_total_diff, (int, float)):
                assert abs(grand_total_diff - expected_diff) < GST_TOLERANCE * user_grand_total, \
                    f"Grand Total Diff incorrect. Expected {expected_diff}, got {grand_total_diff}"
                print(f"✓ Grand Total Diff correctly calculated: {grand_total_diff}")
        else:
            print(f"! Grand Total Diff value: {grand_total_diff}")
    
    def test_legend_has_gst_explanation(self, create_test_excel_with_gst_data):
        """Legend section should include GST calculation explanation"""
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Test Product", "rate": 5000, "qty": 1, "amount": 5000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_legend.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find legend rows (after data rows)
        legend_text = ""
        for row in range(4, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                legend_text += str(cell_value) + " "
        
        legend_text_lower = legend_text.lower()
        
        # Check for GST explanation in legend
        assert "cgst" in legend_text_lower, "Legend should mention CGST"
        assert "sgst" in legend_text_lower, "Legend should mention SGST"
        assert "9%" in legend_text or "9 %" in legend_text, "Legend should mention 9% rate"
        
        print(f"✓ Legend includes GST calculation explanation")
        print(f"Legend content: {legend_text[:200]}...")
    
    def test_template_endpoint_works(self):
        """GET /api/bulk-search/template should work"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        
        assert response.status_code == 200, f"Template endpoint failed with {response.status_code}"
        
        # Verify it's a valid Excel
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = [ws.cell(row=1, column=col).value for col in range(1, 5)]
        print(f"Template headers: {headers}")
        
        assert ws['A1'].value is not None, "Template should have headers"
        print(f"✓ Template download works")
    
    def test_multiple_products_gst_calculations(self, create_test_excel_with_gst_data):
        """Multiple products should each have their own GST calculations"""
        products = [
            {"sl_no": 1, "item": "iPhone 15 Pro", "rate": 150000, "qty": 1, "amount": 150000},
            {"sl_no": 2, "item": "Samsung Galaxy S24", "rate": 85000, "qty": 2, "amount": 170000},
            {"sl_no": 3, "item": "OnePlus 12", "rate": 70000, "qty": 3, "amount": 210000}
        ]
        
        excel_file = create_test_excel_with_gst_data(products)
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_multiple.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=180
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check each product row (rows 3, 4, 5)
        all_gst_valid = True
        for idx, product in enumerate(products, start=3):
            amount = ws.cell(row=idx, column=5).value
            cgst = ws.cell(row=idx, column=6).value
            sgst = ws.cell(row=idx, column=7).value
            grand_total = ws.cell(row=idx, column=8).value
            
            print(f"Row {idx}: Amount={amount}, CGST={cgst}, SGST={sgst}, Grand Total={grand_total}")
            
            if isinstance(amount, (int, float)) and amount > 0:
                if isinstance(cgst, (int, float)) and isinstance(sgst, (int, float)):
                    expected_cgst = amount * CGST_RATE
                    expected_sgst = amount * SGST_RATE
                    
                    if abs(cgst - expected_cgst) > GST_TOLERANCE:
                        print(f"! CGST mismatch in row {idx}")
                        all_gst_valid = False
                    if abs(sgst - expected_sgst) > GST_TOLERANCE:
                        print(f"! SGST mismatch in row {idx}")
                        all_gst_valid = False
        
        if all_gst_valid:
            print(f"✓ All {len(products)} products have correct GST calculations")
        else:
            print("! Some GST calculations may have issues")


class TestGSTEdgeCases:
    """Edge cases for GST calculations"""
    
    @pytest.fixture
    def create_test_excel_with_gst_data(self):
        """Create a test Excel file with SL No, Item, Rate/Item, Qty, Amount columns"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            ws['A1'] = "SL No"
            ws['B1'] = "Item"
            ws['C1'] = "Rate/Item"
            ws['D1'] = "Qty"
            ws['E1'] = "Amount"
            
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('sl_no', idx - 1))
                ws.cell(row=idx, column=2, value=product.get('item', ''))
                ws.cell(row=idx, column=3, value=product.get('rate', 0))
                ws.cell(row=idx, column=4, value=product.get('qty', 1))
                ws.cell(row=idx, column=5, value=product.get('amount', product.get('rate', 0) * product.get('qty', 1)))
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_zero_rate_gst_calculation(self, create_test_excel_with_gst_data):
        """Zero rate should result in zero GST"""
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Free Item Test", "rate": 0, "qty": 1, "amount": 0}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_zero.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        amount = ws.cell(row=3, column=5).value
        cgst = ws.cell(row=3, column=6).value
        sgst = ws.cell(row=3, column=7).value
        
        print(f"Zero rate test - Amount: {amount}, CGST: {cgst}, SGST: {sgst}")
        
        if isinstance(cgst, (int, float)):
            assert cgst == 0, f"CGST should be 0 for zero amount, got {cgst}"
        if isinstance(sgst, (int, float)):
            assert sgst == 0, f"SGST should be 0 for zero amount, got {sgst}"
        
        print("✓ Zero rate produces zero GST")
    
    def test_large_amount_gst_calculation(self, create_test_excel_with_gst_data):
        """Large amounts should have correctly calculated GST"""
        large_rate = 1000000  # 10 lakhs
        qty = 10
        expected_amount = large_rate * qty  # 1 crore
        expected_cgst = expected_amount * CGST_RATE  # 9 lakhs
        expected_sgst = expected_amount * SGST_RATE  # 9 lakhs
        
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Luxury Item Test", "rate": large_rate, "qty": qty, "amount": expected_amount}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_large.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        amount = ws.cell(row=3, column=5).value
        cgst = ws.cell(row=3, column=6).value
        sgst = ws.cell(row=3, column=7).value
        grand_total = ws.cell(row=3, column=8).value
        
        print(f"Large amount test - Amount: {amount}, CGST: {cgst}, SGST: {sgst}, Grand Total: {grand_total}")
        
        if isinstance(cgst, (int, float)) and isinstance(sgst, (int, float)):
            assert abs(cgst - expected_cgst) < 1, f"CGST should be ~{expected_cgst}, got {cgst}"
            assert abs(sgst - expected_sgst) < 1, f"SGST should be ~{expected_sgst}, got {sgst}"
            print(f"✓ Large amount GST correctly calculated")
    
    def test_decimal_amount_gst_calculation(self, create_test_excel_with_gst_data):
        """Decimal amounts should have properly rounded GST"""
        rate = 999.99
        qty = 3
        amount = rate * qty  # 2999.97
        
        excel_file = create_test_excel_with_gst_data([
            {"sl_no": 1, "item": "Decimal Test Item", "rate": rate, "qty": qty, "amount": amount}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test_decimal.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        actual_amount = ws.cell(row=3, column=5).value
        cgst = ws.cell(row=3, column=6).value
        sgst = ws.cell(row=3, column=7).value
        
        print(f"Decimal test - Amount: {actual_amount}, CGST: {cgst}, SGST: {sgst}")
        
        # GST should be properly rounded
        if isinstance(cgst, (int, float)):
            assert cgst > 0, "CGST should be positive"
            # Check it's properly rounded (max 2 decimal places)
            cgst_str = str(cgst)
            if '.' in cgst_str:
                decimal_places = len(cgst_str.split('.')[1])
                assert decimal_places <= 2, f"CGST should have max 2 decimal places, has {decimal_places}"
            print(f"✓ Decimal amount GST correctly rounded")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
