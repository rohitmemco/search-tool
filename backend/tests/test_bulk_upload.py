"""
Test cases for Bulk Excel Upload feature
Tests the /api/bulk-search/template and /api/bulk-search/upload endpoints
"""
import pytest
import requests
import os
import io
import openpyxl
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestBulkSearchTemplate:
    """Tests for GET /api/bulk-search/template endpoint"""
    
    def test_template_download_returns_200(self):
        """Template download endpoint should return 200"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("✓ Template download returns 200")
    
    def test_template_is_valid_excel(self):
        """Downloaded template should be a valid Excel file"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200
        
        # Try to load as Excel
        try:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            assert ws is not None
            print(f"✓ Template is valid Excel with sheet: {ws.title}")
        except Exception as e:
            pytest.fail(f"Template is not a valid Excel file: {e}")
    
    def test_template_has_correct_headers(self):
        """Template should have SL No, Item, Quantity headers"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        header_a = ws['A1'].value
        header_b = ws['B1'].value
        header_c = ws['C1'].value
        
        assert header_a == "SL No", f"Expected 'SL No', got '{header_a}'"
        assert header_b == "Item", f"Expected 'Item' in header B, got '{header_b}'"
        assert header_c == "Quantity", f"Expected 'Quantity' in header C, got '{header_c}'"
        print(f"✓ Template has correct headers: '{header_a}', '{header_b}', '{header_c}'")
    
    def test_template_has_sample_data(self):
        """Template should have sample data rows"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check row 2 has sample data
        sample_product = ws['A2'].value
        assert sample_product is not None, "Template should have sample data in row 2"
        print(f"✓ Template has sample data: '{sample_product}'")
    
    def test_template_content_disposition_header(self):
        """Template response should have Content-Disposition header for download"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200
        
        content_disp = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disp, f"Expected 'attachment' in Content-Disposition, got '{content_disp}'"
        assert '.xlsx' in content_disp, f"Expected '.xlsx' in filename, got '{content_disp}'"
        print(f"✓ Content-Disposition header correct: {content_disp}")


class TestBulkSearchUpload:
    """Tests for POST /api/bulk-search/upload endpoint"""
    
    @pytest.fixture
    def create_test_excel(self):
        """Create a test Excel file for upload"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Headers
            ws['A1'] = "Product Name"
            ws['B1'] = "Location (Optional)"
            
            # Add products
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('name', ''))
                ws.cell(row=idx, column=2, value=product.get('location', ''))
            
            # Save to BytesIO
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_upload_returns_200_with_valid_excel(self, create_test_excel):
        """Upload with valid Excel should return 200"""
        excel_file = create_test_excel([
            {"name": "iPhone 15", "location": "Mumbai"}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("✓ Upload with valid Excel returns 200")
    
    def test_upload_returns_excel_file(self, create_test_excel):
        """Upload should return an Excel file as response"""
        excel_file = create_test_excel([
            {"name": "Samsung TV", "location": ""}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Verify response is valid Excel
        try:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            assert ws is not None
            print(f"✓ Response is valid Excel with sheet: {ws.title}")
        except Exception as e:
            pytest.fail(f"Response is not a valid Excel file: {e}")
    
    def test_upload_result_has_correct_columns(self, create_test_excel):
        """Result Excel should have all required columns including GST columns"""
        excel_file = create_test_excel([
            {"name": "Sony Headphones", "location": "Delhi"}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers from row 2 (row 1 has category headers)
        headers = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
        
        # Required columns for GST format
        required_columns = [
            "SL No", "Item", "Your Rate", "Qty", "Your Amount", "CGST", "SGST", "Grand Total",
            "Market Rate", "Market Amount", "Rate Diff", "Grand Total Diff"
        ]
        
        for req_col in required_columns:
            found = any(req_col.lower() in str(h).lower() for h in headers if h)
            assert found, f"Required column '{req_col}' not found in headers: {headers}"
        
        print(f"✓ Result has all required columns: {headers}")
    
    def test_upload_result_has_price_data(self, create_test_excel):
        """Result Excel should have price data for products"""
        excel_file = create_test_excel([
            {"name": "iPhone 15", "location": "Mumbai"}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check row 2 (first data row) has price data
        # Columns: S.No, Product Name, Location, Min Price, Median Price, Max Price, ...
        min_price = ws.cell(row=2, column=4).value  # Min Price column
        med_price = ws.cell(row=2, column=5).value  # Median Price column
        max_price = ws.cell(row=2, column=6).value  # Max Price column
        status = ws.cell(row=2, column=12).value    # Status column
        
        print(f"  Min: {min_price}, Median: {med_price}, Max: {max_price}, Status: {status}")
        
        # Either we have prices or status indicates why not
        if status == "Success":
            assert min_price is not None and min_price > 0, f"Min price should be > 0, got {min_price}"
            assert med_price is not None and med_price > 0, f"Median price should be > 0, got {med_price}"
            assert max_price is not None and max_price > 0, f"Max price should be > 0, got {max_price}"
            assert min_price <= med_price <= max_price, "Prices should be in order: min <= median <= max"
            print(f"✓ Result has valid price data: Min={min_price}, Med={med_price}, Max={max_price}")
        else:
            print(f"✓ Product search status: {status}")
    
    def test_upload_result_has_vendor_info(self, create_test_excel):
        """Result Excel should have vendor information"""
        excel_file = create_test_excel([
            {"name": "MacBook Air", "location": ""}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check vendor columns
        cheapest_vendor = ws.cell(row=2, column=7).value  # Cheapest Vendor
        all_vendors = ws.cell(row=2, column=9).value      # All Vendors
        status = ws.cell(row=2, column=12).value          # Status
        
        if status == "Success":
            assert all_vendors is not None and len(str(all_vendors)) > 0, "All Vendors should have data"
            print(f"✓ Result has vendor info: Cheapest={cheapest_vendor}, All={all_vendors[:50]}...")
        else:
            print(f"✓ Product search status: {status}")
    
    def test_upload_rejects_non_excel_file(self):
        """Upload should reject non-Excel files"""
        # Create a text file
        text_content = b"This is not an Excel file"
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.txt", io.BytesIO(text_content), "text/plain")},
            timeout=30
        )
        
        assert response.status_code == 400, f"Expected 400 for non-Excel file, got {response.status_code}"
        print("✓ Non-Excel file rejected with 400")
    
    def test_upload_rejects_empty_excel(self, create_test_excel):
        """Upload should reject Excel with no products"""
        excel_file = create_test_excel([])  # Empty products list
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30
        )
        
        assert response.status_code == 400, f"Expected 400 for empty Excel, got {response.status_code}"
        print("✓ Empty Excel rejected with 400")
    
    def test_upload_multiple_products(self, create_test_excel):
        """Upload with multiple products should process all"""
        excel_file = create_test_excel([
            {"name": "iPhone 15", "location": "Mumbai"},
            {"name": "Samsung TV", "location": "Delhi"}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=180  # Longer timeout for multiple products
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Count data rows (excluding header and summary)
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            product_name = ws.cell(row=row, column=2).value
            if product_name and product_name not in ["Summary", "Total Products Processed:", "Successful Searches:", "Generated On:"]:
                data_rows += 1
        
        assert data_rows >= 2, f"Expected at least 2 data rows, got {data_rows}"
        print(f"✓ Multiple products processed: {data_rows} rows")
    
    def test_upload_content_disposition_header(self, create_test_excel):
        """Result response should have Content-Disposition header"""
        excel_file = create_test_excel([
            {"name": "Test Product", "location": ""}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        content_disp = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disp, f"Expected 'attachment' in Content-Disposition"
        assert 'PriceComparison_Results' in content_disp, f"Expected 'PriceComparison_Results' in filename, got {content_disp}"
        print(f"✓ Content-Disposition header correct: {content_disp}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
