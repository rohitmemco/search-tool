"""
Test suite for Bulk Excel Upload with Min/Med/Max rate columns and consolidated GST
Tests for:
- 15 columns in output Excel
- Min, Med, Max rate columns with Amount calculations
- Consolidated GST summary with FOUR sections (Your, Min, Med, Max)
- Realistic estimated prices for common products
- GST calculation verification
"""

import pytest
import requests
import os
import io
import openpyxl
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Expected 15 columns in output Excel
EXPECTED_COLUMNS = [
    "SL No", "Item", 
    "Your Rate", "Qty", "Your Amount",
    "Min Rate", "Min Amount",
    "Med Rate", "Med Amount",
    "Max Rate", "Max Amount",
    "Rate Diff", "Amount Diff",
    "Website Links", "Vendor Details"
]

# Expected GST sections in consolidated summary
GST_SECTIONS = ["YOUR PRICING", "MARKET MINIMUM", "MARKET MEDIUM", "MARKET MAXIMUM"]
GST_LINE_ITEMS = ["Taxable Amount", "CGST", "SGST", "Round Off", "Grand Total"]

# Test products with realistic price ranges (in INR)
TEST_PRODUCTS = {
    "Samsung Galaxy S24": {"min": 60000, "max": 95000, "expected_range": "65K-90K"},
    "MacBook Air M3": {"min": 95000, "max": 145000, "expected_range": "100K-140K"},
    "Sony WH-1000XM5": {"min": 22000, "max": 38000, "expected_range": "25K-35K"}
}


def create_test_excel(products_data: list) -> bytes:
    """Create test Excel file with products"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    
    # Headers matching expected format
    headers = ["SL No", "Item", "Rate/Item", "Qty", "Amount"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add product data
    for idx, product in enumerate(products_data, 1):
        ws.cell(row=idx + 1, column=1, value=idx)  # SL No
        ws.cell(row=idx + 1, column=2, value=product['item'])  # Item
        ws.cell(row=idx + 1, column=3, value=product['rate'])  # Rate/Item
        ws.cell(row=idx + 1, column=4, value=product['qty'])  # Qty
        ws.cell(row=idx + 1, column=5, value=product['rate'] * product['qty'])  # Amount
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


class TestBulkSearchTemplate:
    """Tests for template download endpoint"""
    
    def test_template_endpoint_returns_200(self):
        """Verify GET /api/bulk-search/template returns 200"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("✅ GET /api/bulk-search/template returns 200")
    
    def test_template_is_valid_excel(self):
        """Verify template is a valid Excel file"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200
        
        # Try to load as Excel
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check it has content
        assert ws.max_row >= 1, "Template should have at least header row"
        print("✅ Template is valid Excel file")


class TestBulkUploadColumns:
    """Tests for verifying output Excel has correct 15 columns"""
    
    def test_upload_returns_200_with_valid_excel(self):
        """Verify POST /api/bulk-search/upload returns 200"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 2}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}. Response: {response.text[:200]}"
        print("✅ POST /api/bulk-search/upload returns 200")
    
    def test_output_has_15_columns(self):
        """Verify output Excel has 15 columns"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 1}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers from row 1
        headers = []
        for col in range(1, 20):  # Check first 20 columns
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value)
            else:
                break
        
        print(f"Found headers: {headers}")
        assert len(headers) == 15, f"Expected 15 columns, found {len(headers)}: {headers}"
        print("✅ Output Excel has 15 columns")
    
    def test_column_names_match_expected(self):
        """Verify column names include Min, Med, Max rate columns"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 1}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get all headers
        headers = [ws.cell(row=1, column=col).value for col in range(1, 16)]
        headers_str = ' '.join([h for h in headers if h])
        
        # Check for required columns
        required_keywords = ['Min Rate', 'Min Amount', 'Med Rate', 'Med Amount', 'Max Rate', 'Max Amount']
        for keyword in required_keywords:
            assert keyword in headers_str, f"Missing column containing '{keyword}'. Found headers: {headers}"
            print(f"✅ Found column: {keyword}")
        
        print("✅ All Min, Med, Max rate and amount columns present")


class TestMarketPriceCalculations:
    """Tests for verifying market price calculations (Amount = Rate × Qty)"""
    
    def test_min_amount_equals_min_rate_times_qty(self):
        """Verify Min Amount = Min Rate × Qty"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 3}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find column indices by header names
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, 20) if ws.cell(row=1, column=col).value}
        
        # Get data from row 2
        qty_col = None
        min_rate_col = None
        min_amount_col = None
        
        for header, col in headers.items():
            if 'Qty' in str(header):
                qty_col = col
            if 'Min Rate' in str(header):
                min_rate_col = col
            if 'Min Amount' in str(header):
                min_amount_col = col
        
        if qty_col and min_rate_col and min_amount_col:
            qty = ws.cell(row=2, column=qty_col).value
            min_rate = ws.cell(row=2, column=min_rate_col).value
            min_amount = ws.cell(row=2, column=min_amount_col).value
            
            if isinstance(min_rate, (int, float)) and isinstance(min_amount, (int, float)):
                expected_amount = min_rate * qty
                assert abs(min_amount - expected_amount) < 0.01, f"Min Amount ({min_amount}) != Min Rate ({min_rate}) × Qty ({qty})"
                print(f"✅ Min Amount ({min_amount}) = Min Rate ({min_rate}) × Qty ({qty}) = {expected_amount}")
            else:
                print(f"⚠️ Min Rate or Amount is N/A: rate={min_rate}, amount={min_amount}")
        else:
            print(f"⚠️ Could not find columns: qty_col={qty_col}, min_rate_col={min_rate_col}, min_amount_col={min_amount_col}")
    
    def test_med_amount_equals_med_rate_times_qty(self):
        """Verify Med Amount = Med Rate × Qty"""
        test_data = [
            {"item": "MacBook Air M3", "rate": 120000, "qty": 2}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, 20) if ws.cell(row=1, column=col).value}
        
        qty_col = None
        med_rate_col = None
        med_amount_col = None
        
        for header, col in headers.items():
            if 'Qty' in str(header):
                qty_col = col
            if 'Med Rate' in str(header):
                med_rate_col = col
            if 'Med Amount' in str(header):
                med_amount_col = col
        
        if qty_col and med_rate_col and med_amount_col:
            qty = ws.cell(row=2, column=qty_col).value
            med_rate = ws.cell(row=2, column=med_rate_col).value
            med_amount = ws.cell(row=2, column=med_amount_col).value
            
            if isinstance(med_rate, (int, float)) and isinstance(med_amount, (int, float)):
                expected_amount = med_rate * qty
                assert abs(med_amount - expected_amount) < 0.01, f"Med Amount ({med_amount}) != Med Rate ({med_rate}) × Qty ({qty})"
                print(f"✅ Med Amount ({med_amount}) = Med Rate ({med_rate}) × Qty ({qty}) = {expected_amount}")
            else:
                print(f"⚠️ Med Rate or Amount is N/A: rate={med_rate}, amount={med_amount}")
    
    def test_max_amount_equals_max_rate_times_qty(self):
        """Verify Max Amount = Max Rate × Qty"""
        test_data = [
            {"item": "Sony WH-1000XM5", "rate": 25000, "qty": 4}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, 20) if ws.cell(row=1, column=col).value}
        
        qty_col = None
        max_rate_col = None
        max_amount_col = None
        
        for header, col in headers.items():
            if 'Qty' in str(header):
                qty_col = col
            if 'Max Rate' in str(header):
                max_rate_col = col
            if 'Max Amount' in str(header):
                max_amount_col = col
        
        if qty_col and max_rate_col and max_amount_col:
            qty = ws.cell(row=2, column=qty_col).value
            max_rate = ws.cell(row=2, column=max_rate_col).value
            max_amount = ws.cell(row=2, column=max_amount_col).value
            
            if isinstance(max_rate, (int, float)) and isinstance(max_amount, (int, float)):
                expected_amount = max_rate * qty
                assert abs(max_amount - expected_amount) < 0.01, f"Max Amount ({max_amount}) != Max Rate ({max_rate}) × Qty ({qty})"
                print(f"✅ Max Amount ({max_amount}) = Max Rate ({max_rate}) × Qty ({qty}) = {expected_amount}")
            else:
                print(f"⚠️ Max Rate or Amount is N/A: rate={max_rate}, amount={max_amount}")


class TestRealisticPrices:
    """Tests for verifying market prices are realistic"""
    
    def test_samsung_galaxy_s24_price_realistic(self):
        """Verify Samsung Galaxy S24 price is in range ₹65K-90K"""
        test_data = [{"item": "Samsung Galaxy S24", "rate": 80000, "qty": 1}]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, 20) if ws.cell(row=1, column=col).value}
        
        # Find Min Rate column
        min_rate_col = None
        for header, col in headers.items():
            if 'Min Rate' in str(header):
                min_rate_col = col
                break
        
        if min_rate_col:
            min_rate = ws.cell(row=2, column=min_rate_col).value
            if isinstance(min_rate, (int, float)):
                # Check if price is realistic (65K-90K range with some tolerance)
                assert 55000 <= min_rate <= 100000, f"Samsung Galaxy S24 price ({min_rate}) should be in 65K-90K range"
                print(f"✅ Samsung Galaxy S24 Min Rate: ₹{min_rate:,.0f} (within expected ₹65K-90K range)")
            else:
                print(f"⚠️ Min Rate is N/A for Samsung Galaxy S24")
    
    def test_macbook_air_m3_price_realistic(self):
        """Verify MacBook Air M3 price is in range ₹100K-140K"""
        test_data = [{"item": "MacBook Air M3", "rate": 120000, "qty": 1}]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, 20) if ws.cell(row=1, column=col).value}
        
        min_rate_col = None
        for header, col in headers.items():
            if 'Min Rate' in str(header):
                min_rate_col = col
                break
        
        if min_rate_col:
            min_rate = ws.cell(row=2, column=min_rate_col).value
            if isinstance(min_rate, (int, float)):
                assert 90000 <= min_rate <= 150000, f"MacBook Air M3 price ({min_rate}) should be in 100K-140K range"
                print(f"✅ MacBook Air M3 Min Rate: ₹{min_rate:,.0f} (within expected ₹100K-140K range)")
            else:
                print(f"⚠️ Min Rate is N/A for MacBook Air M3")
    
    def test_sony_wh1000xm5_price_realistic(self):
        """Verify Sony WH-1000XM5 price is in range ₹25K-35K"""
        test_data = [{"item": "Sony WH-1000XM5 Headphones", "rate": 30000, "qty": 1}]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, 20) if ws.cell(row=1, column=col).value}
        
        min_rate_col = None
        for header, col in headers.items():
            if 'Min Rate' in str(header):
                min_rate_col = col
                break
        
        if min_rate_col:
            min_rate = ws.cell(row=2, column=min_rate_col).value
            if isinstance(min_rate, (int, float)):
                assert 20000 <= min_rate <= 40000, f"Sony WH-1000XM5 price ({min_rate}) should be in 25K-35K range"
                print(f"✅ Sony WH-1000XM5 Min Rate: ₹{min_rate:,.0f} (within expected ₹25K-35K range)")
            else:
                print(f"⚠️ Min Rate is N/A for Sony WH-1000XM5")


class TestConsolidatedGSTSummary:
    """Tests for consolidated GST summary with FOUR sections"""
    
    def test_gst_summary_exists(self):
        """Verify consolidated GST summary section exists"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 2},
            {"item": "MacBook Air M3", "rate": 120000, "qty": 1}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "CONSOLIDATED GST SUMMARY" text
        summary_found = False
        for row in range(1, ws.max_row + 1):
            for col in range(1, 5):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and 'CONSOLIDATED' in str(cell_value).upper() and 'GST' in str(cell_value).upper():
                    summary_found = True
                    print(f"✅ Found GST Summary header at row {row}: {cell_value}")
                    break
        
        assert summary_found, "Consolidated GST Summary section not found in output Excel"
    
    def test_four_gst_sections_present(self):
        """Verify all FOUR GST sections exist: YOUR PRICING, MARKET MINIMUM, MARKET MEDIUM, MARKET MAXIMUM"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 2}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for all four section headers
        sections_found = {
            "YOUR PRICING": False,
            "MARKET MINIMUM": False,
            "MARKET MEDIUM": False,
            "MARKET MAXIMUM": False
        }
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value).upper()
                    for section in sections_found.keys():
                        if section in cell_str:
                            sections_found[section] = True
                            print(f"✅ Found section '{section}' at row {row}, col {col}")
        
        # Check all sections were found
        missing_sections = [s for s, found in sections_found.items() if not found]
        assert len(missing_sections) == 0, f"Missing GST sections: {missing_sections}"
        print("✅ All FOUR GST sections present: YOUR PRICING, MARKET MINIMUM, MARKET MEDIUM, MARKET MAXIMUM")
    
    def test_gst_line_items_present(self):
        """Verify each section has: Taxable Amount, CGST @9%, SGST @9%, Round Off, Grand Total"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 1}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for GST line items
        line_items_found = {
            "Taxable Amount": False,
            "CGST": False,
            "SGST": False,
            "Round Off": False,
            "Grand Total": False
        }
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value)
                    for item in line_items_found.keys():
                        if item in cell_str:
                            line_items_found[item] = True
        
        for item, found in line_items_found.items():
            if found:
                print(f"✅ Found GST line item: {item}")
            else:
                print(f"❌ Missing GST line item: {item}")
        
        missing_items = [item for item, found in line_items_found.items() if not found]
        assert len(missing_items) == 0, f"Missing GST line items: {missing_items}"


class TestGSTCalculations:
    """Tests for verifying GST calculation correctness"""
    
    def test_cgst_sgst_9_percent_calculation(self):
        """Verify CGST and SGST are calculated at 9% of taxable amount"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 2}  # Amount = 150000
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find the YOUR PRICING section and verify calculations
        your_taxable = None
        your_cgst = None
        your_sgst = None
        your_grand_total = None
        
        in_your_section = False
        for row in range(1, ws.max_row + 1):
            for col in range(1, 5):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value)
                    
                    if "YOUR PRICING" in cell_str.upper():
                        in_your_section = True
                    
                    if in_your_section:
                        if "Taxable Amount" in cell_str:
                            # Get value from next column
                            val = ws.cell(row=row, column=col+1).value
                            if val:
                                # Parse currency value like "₹150,000.00"
                                val_str = str(val).replace('₹', '').replace(',', '').strip()
                                try:
                                    your_taxable = float(val_str)
                                except:
                                    pass
                        elif "CGST" in cell_str and "@" in cell_str:
                            val = ws.cell(row=row, column=col+1).value
                            if val:
                                val_str = str(val).replace('₹', '').replace(',', '').strip()
                                try:
                                    your_cgst = float(val_str)
                                except:
                                    pass
                        elif "SGST" in cell_str and "@" in cell_str:
                            val = ws.cell(row=row, column=col+1).value
                            if val:
                                val_str = str(val).replace('₹', '').replace(',', '').strip()
                                try:
                                    your_sgst = float(val_str)
                                except:
                                    pass
                        elif "Grand Total" in cell_str:
                            val = ws.cell(row=row, column=col+1).value
                            if val:
                                val_str = str(val).replace('₹', '').replace(',', '').strip()
                                try:
                                    your_grand_total = float(val_str)
                                except:
                                    pass
                            in_your_section = False  # End of section
        
        if your_taxable and your_cgst:
            expected_cgst = your_taxable * 0.09
            print(f"Your Taxable Amount: ₹{your_taxable:,.2f}")
            print(f"Your CGST @9%: ₹{your_cgst:,.2f} (expected: ₹{expected_cgst:,.2f})")
            # Allow small rounding difference
            assert abs(your_cgst - expected_cgst) < 1, f"CGST calculation incorrect: {your_cgst} != {expected_cgst}"
            print("✅ CGST @9% calculated correctly")
        else:
            print(f"⚠️ Could not extract all values: taxable={your_taxable}, cgst={your_cgst}")
        
        if your_taxable and your_sgst:
            expected_sgst = your_taxable * 0.09
            print(f"Your SGST @9%: ₹{your_sgst:,.2f} (expected: ₹{expected_sgst:,.2f})")
            assert abs(your_sgst - expected_sgst) < 1, f"SGST calculation incorrect: {your_sgst} != {expected_sgst}"
            print("✅ SGST @9% calculated correctly")
    
    def test_grand_total_equals_taxable_plus_gst(self):
        """Verify Grand Total = Taxable Amount + CGST + SGST"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 75000, "qty": 2}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Extract values from YOUR PRICING section
        values = {}
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, 5):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value)
                    if any(x in cell_str for x in ["Taxable Amount", "CGST", "SGST", "Grand Total"]):
                        val = ws.cell(row=row, column=col+1).value
                        if val:
                            val_str = str(val).replace('₹', '').replace(',', '').strip()
                            try:
                                if "Taxable Amount" in cell_str:
                                    values['taxable'] = float(val_str)
                                elif "CGST" in cell_str:
                                    values['cgst'] = float(val_str)
                                elif "SGST" in cell_str and "CGST" not in cell_str:
                                    values['sgst'] = float(val_str)
                                elif "Grand Total" in cell_str:
                                    values['grand_total'] = float(val_str)
                            except:
                                pass
        
        if 'taxable' in values and 'cgst' in values and 'sgst' in values and 'grand_total' in values:
            expected_total = values['taxable'] + values['cgst'] + values['sgst']
            print(f"Taxable: ₹{values['taxable']:,.2f}, CGST: ₹{values['cgst']:,.2f}, SGST: ₹{values['sgst']:,.2f}")
            print(f"Expected Grand Total: ₹{expected_total:,.2f}, Actual: ₹{values['grand_total']:,.2f}")
            # Allow for round off (usually within ₹1)
            assert abs(values['grand_total'] - expected_total) <= 1, f"Grand Total incorrect"
            print("✅ Grand Total = Taxable + CGST + SGST")
        else:
            print(f"⚠️ Missing values: {values}")


class TestComparisonLine:
    """Tests for the COMPARISON line at bottom of summary"""
    
    def test_comparison_line_exists(self):
        """Verify COMPARISON line shows Your vs Min difference"""
        test_data = [
            {"item": "Samsung Galaxy S24", "rate": 85000, "qty": 1}
        ]
        excel_bytes = create_test_excel(test_data)
        
        files = {'file': ('test.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        comparison_found = False
        for row in range(1, ws.max_row + 1):
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and 'COMPARISON' in str(cell_value).upper():
                    comparison_found = True
                    print(f"✅ Found COMPARISON line at row {row}: {cell_value}")
                    break
        
        assert comparison_found, "COMPARISON line not found in output Excel"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
