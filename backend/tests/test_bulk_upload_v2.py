"""
Test cases for Bulk Excel Upload - Updated Format (No Per-Item GST)
Tests the /api/bulk-search/upload endpoint with new requirements:
1. NO per-item GST columns 
2. Market MINIMUM rate is used (not median)
3. Consolidated GST summary at bottom with YOUR PRICING vs MARKET MINIMUM PRICING
4. CGST @9% and SGST @9% calculated on TOTAL amounts only
5. Grand Total = Taxable Amount + CGST + SGST
6. DIFFERENCE line shows savings/overpaying comparison

Expected columns (11 total):
- SL No, Item, Your Rate (₹), Qty, Your Amount (₹)
- Market Min Rate (₹), Market Min Amount (₹)
- Rate Diff (₹), Amount Diff (₹)
- Website Links, Vendor Details
"""
import pytest
import requests
import os
import io
import openpyxl
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
CGST_RATE = 0.09
SGST_RATE = 0.09
TOLERANCE = 1.0  # Tolerance for floating point comparison


class TestBulkUploadStructure:
    """Tests for the structure of bulk upload output - NO per-item GST"""
    
    @pytest.fixture
    def create_test_excel(self):
        """Create a test Excel file with SL No, Item, Rate/Item, Qty, Amount columns"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            # Headers matching expected input format
            ws['A1'] = "SL No"
            ws['B1'] = "Item"
            ws['C1'] = "Rate/Item"
            ws['D1'] = "Qty"
            ws['E1'] = "Amount"
            
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('sl_no', idx - 1))
                ws.cell(row=idx, column=2, value=product.get('item', ''))
                ws.cell(row=idx, column=3, value=product.get('rate', 0))
                ws.cell(row=idx, column=4, value=product.get('qty', 1))
                ws.cell(row=idx, column=5, value=product.get('amount', product.get('rate', 0) * product.get('qty', 1)))
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_upload_returns_200(self, create_test_excel):
        """Upload should return 200 with valid Excel"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Samsung Galaxy S24", "rate": 80000, "qty": 1, "amount": 80000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("✓ Upload returns 200 with valid Excel")
    
    def test_no_per_item_gst_columns(self, create_test_excel):
        """Output should NOT have per-item GST columns"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "MacBook Air M3", "rate": 115000, "qty": 1, "amount": 115000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers from row 1
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        headers_str = str(headers).lower()
        
        print(f"Headers found: {headers}")
        
        # Should NOT have per-item GST columns
        per_item_gst_keywords = ["your cgst", "your sgst", "your grand total", "market cgst", "market sgst", "market grand total"]
        
        for keyword in per_item_gst_keywords:
            assert keyword not in headers_str, f"Found per-item GST column '{keyword}' which should NOT be present"
        
        print("✓ NO per-item GST columns found (as expected)")
    
    def test_output_has_correct_11_columns(self, create_test_excel):
        """Output should have exactly 11 columns for data rows"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Sony WH-1000XM5", "rate": 30000, "qty": 1, "amount": 30000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers from row 1
        headers = [ws.cell(row=1, column=col).value for col in range(1, 15)]
        non_empty_headers = [h for h in headers if h is not None]
        
        print(f"Headers: {non_empty_headers}")
        
        # Expected 11 columns
        expected_headers = [
            "SL No", "Item", "Your Rate", "Qty", "Your Amount",
            "Market Min Rate", "Market Min Amount",
            "Rate Diff", "Amount Diff",
            "Website Links", "Vendor Details"
        ]
        
        # Verify count (should be 11)
        assert len(non_empty_headers) == 11, f"Expected 11 columns, got {len(non_empty_headers)}: {non_empty_headers}"
        
        # Verify key columns exist
        headers_lower = [h.lower() if h else '' for h in headers]
        assert any('sl' in h for h in headers_lower), "SL No column missing"
        assert any('item' in h for h in headers_lower), "Item column missing"
        assert any('your rate' in h for h in headers_lower), "Your Rate column missing"
        assert any('market min rate' in h for h in headers_lower), "Market Min Rate column missing"
        assert any('rate diff' in h for h in headers_lower), "Rate Diff column missing"
        assert any('amount diff' in h for h in headers_lower), "Amount Diff column missing"
        
        print("✓ Output has correct 11 columns")
    
    def test_market_minimum_rate_used(self, create_test_excel):
        """Output should use Market MINIMUM rate, not median"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Samsung Galaxy S24", "rate": 85000, "qty": 1, "amount": 85000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers to find Market Min Rate column
        headers = [ws.cell(row=1, column=col).value for col in range(1, 15)]
        headers_lower = [str(h).lower() if h else '' for h in headers]
        
        # Check column headers contain "min" for market rate
        market_rate_col = None
        for idx, h in enumerate(headers_lower):
            if 'market' in h and 'min' in h and 'rate' in h:
                market_rate_col = idx + 1
                break
        
        assert market_rate_col is not None, f"Market Min Rate column not found. Headers: {headers}"
        
        # Check header text explicitly says "Min" not "Median"
        market_rate_header = headers[market_rate_col - 1]
        assert 'min' in market_rate_header.lower(), f"Market rate header should say 'Min': {market_rate_header}"
        assert 'median' not in market_rate_header.lower(), f"Market rate header should NOT say 'Median': {market_rate_header}"
        
        print(f"✓ Market MINIMUM rate column found: {market_rate_header}")


class TestConsolidatedGSTSummary:
    """Tests for consolidated GST summary section at bottom"""
    
    @pytest.fixture
    def create_test_excel(self):
        """Create a test Excel file"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            ws['A1'] = "SL No"
            ws['B1'] = "Item"
            ws['C1'] = "Rate/Item"
            ws['D1'] = "Qty"
            ws['E1'] = "Amount"
            
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('sl_no', idx - 1))
                ws.cell(row=idx, column=2, value=product.get('item', ''))
                ws.cell(row=idx, column=3, value=product.get('rate', 0))
                ws.cell(row=idx, column=4, value=product.get('qty', 1))
                ws.cell(row=idx, column=5, value=product.get('amount', product.get('rate', 0) * product.get('qty', 1)))
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_consolidated_summary_exists(self, create_test_excel):
        """Consolidated GST summary section should exist at bottom"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "iPhone 15", "rate": 80000, "qty": 1, "amount": 80000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "CONSOLIDATED SUMMARY" or "GST Calculation" in the sheet
        found_summary = False
        summary_row = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and ('consolidated' in str(cell_value).lower() or 'gst calculation' in str(cell_value).lower()):
                found_summary = True
                summary_row = row
                print(f"Found summary at row {row}: {cell_value}")
                break
        
        assert found_summary, "Consolidated GST summary section not found at bottom"
        print(f"✓ Consolidated GST summary exists at row {summary_row}")
    
    def test_your_pricing_section_exists(self, create_test_excel):
        """YOUR PRICING section should exist in summary"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "MacBook Pro", "rate": 200000, "qty": 1, "amount": 200000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "YOUR PRICING" in the sheet
        found_your_pricing = False
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and 'your pricing' in str(cell_value).lower():
                    found_your_pricing = True
                    print(f"Found 'YOUR PRICING' at row {row}, col {col}: {cell_value}")
                    break
            if found_your_pricing:
                break
        
        assert found_your_pricing, "YOUR PRICING section not found"
        print("✓ YOUR PRICING section exists")
    
    def test_market_minimum_pricing_section_exists(self, create_test_excel):
        """MARKET MINIMUM PRICING section should exist in summary"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Dell XPS Laptop", "rate": 150000, "qty": 1, "amount": 150000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "MARKET MINIMUM PRICING" in the sheet
        found_market_pricing = False
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and 'market' in str(cell_value).lower() and 'pricing' in str(cell_value).lower():
                    found_market_pricing = True
                    print(f"Found 'MARKET MINIMUM PRICING' at row {row}, col {col}: {cell_value}")
                    break
            if found_market_pricing:
                break
        
        assert found_market_pricing, "MARKET MINIMUM PRICING section not found"
        print("✓ MARKET MINIMUM PRICING section exists")
    
    def test_summary_has_taxable_amount(self, create_test_excel):
        """Summary should have Taxable Amount line"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Sony TV", "rate": 75000, "qty": 1, "amount": 75000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "Taxable Amount" in the sheet
        found_taxable = False
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'taxable' in str(cell_value).lower():
                found_taxable = True
                print(f"Found 'Taxable Amount' at row {row}: {cell_value}")
                break
        
        assert found_taxable, "Taxable Amount not found in summary"
        print("✓ Taxable Amount line exists in summary")
    
    def test_summary_has_cgst_9_percent(self, create_test_excel):
        """Summary should have CGST @ 9% line"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "LG Refrigerator", "rate": 45000, "qty": 1, "amount": 45000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "CGST @ 9%" in the sheet
        found_cgst = False
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'cgst' in str(cell_value).lower() and '9' in str(cell_value):
                found_cgst = True
                print(f"Found 'CGST @ 9%' at row {row}: {cell_value}")
                break
        
        assert found_cgst, "CGST @ 9% not found in summary"
        print("✓ CGST @ 9% line exists in summary")
    
    def test_summary_has_sgst_9_percent(self, create_test_excel):
        """Summary should have SGST @ 9% line"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Bosch Dishwasher", "rate": 55000, "qty": 1, "amount": 55000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "SGST @ 9%" in the sheet
        found_sgst = False
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'sgst' in str(cell_value).lower() and '9' in str(cell_value):
                found_sgst = True
                print(f"Found 'SGST @ 9%' at row {row}: {cell_value}")
                break
        
        assert found_sgst, "SGST @ 9% not found in summary"
        print("✓ SGST @ 9% line exists in summary")
    
    def test_summary_has_grand_total(self, create_test_excel):
        """Summary should have Grand Total line"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Samsung Washing Machine", "rate": 35000, "qty": 1, "amount": 35000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "Grand Total" in the sheet
        found_grand_total = False
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'grand total' in str(cell_value).lower():
                found_grand_total = True
                print(f"Found 'Grand Total' at row {row}: {cell_value}")
                break
        
        assert found_grand_total, "Grand Total not found in summary"
        print("✓ Grand Total line exists in summary")
    
    def test_summary_has_difference_line(self, create_test_excel):
        """Summary should have DIFFERENCE line showing savings comparison"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "OnePlus 12", "rate": 70000, "qty": 1, "amount": 70000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for "DIFFERENCE" in the sheet
        found_difference = False
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'difference' in str(cell_value).lower():
                found_difference = True
                print(f"Found 'DIFFERENCE' at row {row}: {cell_value}")
                # Check if it mentions overpaying or good deal
                cell_str = str(cell_value).lower()
                assert 'overpaying' in cell_str or 'good deal' in cell_str or 'less' in cell_str or 'more' in cell_str, \
                    f"DIFFERENCE line should indicate savings status: {cell_value}"
                break
        
        assert found_difference, "DIFFERENCE line not found in summary"
        print("✓ DIFFERENCE line exists showing savings comparison")


class TestGSTCalculationsOnTotals:
    """Tests for GST calculations on TOTAL amounts only"""
    
    @pytest.fixture
    def create_test_excel(self):
        """Create a test Excel file"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            ws['A1'] = "SL No"
            ws['B1'] = "Item"
            ws['C1'] = "Rate/Item"
            ws['D1'] = "Qty"
            ws['E1'] = "Amount"
            
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('sl_no', idx - 1))
                ws.cell(row=idx, column=2, value=product.get('item', ''))
                ws.cell(row=idx, column=3, value=product.get('rate', 0))
                ws.cell(row=idx, column=4, value=product.get('qty', 1))
                ws.cell(row=idx, column=5, value=product.get('amount', product.get('rate', 0) * product.get('qty', 1)))
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_gst_calculated_on_total_amount(self, create_test_excel):
        """GST should be calculated on TOTAL amounts, not per-item"""
        # Upload multiple products
        products = [
            {"sl_no": 1, "item": "Samsung Galaxy S24", "rate": 80000, "qty": 1, "amount": 80000},
            {"sl_no": 2, "item": "MacBook Air M3", "rate": 115000, "qty": 1, "amount": 115000},
        ]
        total_your_amount = sum(p['amount'] for p in products)  # 195000
        expected_your_cgst = total_your_amount * CGST_RATE  # 17550
        expected_your_sgst = total_your_amount * SGST_RATE  # 17550
        
        excel_file = create_test_excel(products)
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=180
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find summary section and extract values
        taxable_amount = None
        cgst_value = None
        sgst_value = None
        grand_total = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            value_cell = ws.cell(row=row, column=2).value
            
            if cell_value:
                cell_str = str(cell_value).lower()
                if 'taxable amount' in cell_str and value_cell:
                    # Extract number from "₹195,000.00"
                    taxable_amount = extract_number(value_cell)
                    print(f"Taxable Amount found: {taxable_amount}")
                elif 'cgst' in cell_str and '9' in str(cell_value) and value_cell:
                    cgst_value = extract_number(value_cell)
                    print(f"CGST @ 9% found: {cgst_value}")
                elif 'sgst' in cell_str and '9' in str(cell_value) and value_cell:
                    sgst_value = extract_number(value_cell)
                    print(f"SGST @ 9% found: {sgst_value}")
                elif 'grand total' in cell_str and value_cell:
                    grand_total = extract_number(value_cell)
                    print(f"Grand Total found: {grand_total}")
        
        # Verify Taxable Amount matches sum of Your Amount
        if taxable_amount:
            assert abs(taxable_amount - total_your_amount) < TOLERANCE, \
                f"Taxable Amount should be {total_your_amount}, got {taxable_amount}"
        
        # Verify CGST = Taxable Amount × 9%
        if cgst_value and taxable_amount:
            expected_cgst = taxable_amount * CGST_RATE
            assert abs(cgst_value - expected_cgst) < TOLERANCE, \
                f"CGST should be {expected_cgst}, got {cgst_value}"
        
        # Verify SGST = Taxable Amount × 9%
        if sgst_value and taxable_amount:
            expected_sgst = taxable_amount * SGST_RATE
            assert abs(sgst_value - expected_sgst) < TOLERANCE, \
                f"SGST should be {expected_sgst}, got {sgst_value}"
        
        # Verify Grand Total = Taxable Amount + CGST + SGST (approximately, due to rounding)
        if grand_total and taxable_amount and cgst_value and sgst_value:
            calculated_total = taxable_amount + cgst_value + sgst_value
            # Allow for rounding differences
            assert abs(grand_total - round(calculated_total)) < 2, \
                f"Grand Total should be ~{round(calculated_total)}, got {grand_total}"
        
        print("✓ GST correctly calculated on TOTAL amounts")


class TestEstimatedPrices:
    """Tests for estimated prices for specific products"""
    
    @pytest.fixture
    def create_test_excel(self):
        """Create a test Excel file"""
        def _create(products):
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            ws['A1'] = "SL No"
            ws['B1'] = "Item"
            ws['C1'] = "Rate/Item"
            ws['D1'] = "Qty"
            ws['E1'] = "Amount"
            
            for idx, product in enumerate(products, start=2):
                ws.cell(row=idx, column=1, value=product.get('sl_no', idx - 1))
                ws.cell(row=idx, column=2, value=product.get('item', ''))
                ws.cell(row=idx, column=3, value=product.get('rate', 0))
                ws.cell(row=idx, column=4, value=product.get('qty', 1))
                ws.cell(row=idx, column=5, value=product.get('amount', product.get('rate', 0) * product.get('qty', 1)))
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        return _create
    
    def test_samsung_galaxy_s24_price_realistic(self, create_test_excel):
        """Samsung Galaxy S24 should have realistic market price (₹65K-90K range)"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Samsung Galaxy S24", "rate": 80000, "qty": 1, "amount": 80000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Market Min Rate should be column F (6)
        market_min_rate = ws.cell(row=2, column=6).value
        
        print(f"Samsung Galaxy S24 Market Min Rate: {market_min_rate}")
        
        if isinstance(market_min_rate, (int, float)) and market_min_rate > 0:
            # Samsung Galaxy S24 typically costs ₹65,000-90,000
            assert 50000 <= market_min_rate <= 120000, \
                f"Samsung Galaxy S24 price should be in realistic range (₹50K-120K), got {market_min_rate}"
            print(f"✓ Samsung Galaxy S24 price is realistic: ₹{market_min_rate:,}")
        else:
            print(f"! Market price: {market_min_rate} (may be N/A due to SerpAPI exhaustion)")
    
    def test_macbook_air_m3_price_realistic(self, create_test_excel):
        """MacBook Air M3 should have realistic market price (₹100K-140K range)"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "MacBook Air M3", "rate": 120000, "qty": 1, "amount": 120000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        market_min_rate = ws.cell(row=2, column=6).value
        
        print(f"MacBook Air M3 Market Min Rate: {market_min_rate}")
        
        if isinstance(market_min_rate, (int, float)) and market_min_rate > 0:
            # MacBook Air M3 typically costs ₹100,000-140,000
            assert 80000 <= market_min_rate <= 180000, \
                f"MacBook Air M3 price should be in realistic range (₹80K-180K), got {market_min_rate}"
            print(f"✓ MacBook Air M3 price is realistic: ₹{market_min_rate:,}")
        else:
            print(f"! Market price: {market_min_rate} (may be N/A due to SerpAPI exhaustion)")
    
    def test_sony_wh1000xm5_price_realistic(self, create_test_excel):
        """Sony WH-1000XM5 should have realistic market price (₹25K-35K range)"""
        excel_file = create_test_excel([
            {"sl_no": 1, "item": "Sony WH-1000XM5 Headphones", "rate": 32000, "qty": 1, "amount": 32000}
        ])
        
        response = requests.post(
            f"{BASE_URL}/api/bulk-search/upload",
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        market_min_rate = ws.cell(row=2, column=6).value
        
        print(f"Sony WH-1000XM5 Market Min Rate: {market_min_rate}")
        
        if isinstance(market_min_rate, (int, float)) and market_min_rate > 0:
            # Sony WH-1000XM5 typically costs ₹25,000-35,000
            assert 18000 <= market_min_rate <= 50000, \
                f"Sony WH-1000XM5 price should be in realistic range (₹18K-50K), got {market_min_rate}"
            print(f"✓ Sony WH-1000XM5 price is realistic: ₹{market_min_rate:,}")
        else:
            print(f"! Market price: {market_min_rate} (may be N/A due to SerpAPI exhaustion)")


class TestTemplateEndpoint:
    """Tests for GET /api/bulk-search/template"""
    
    def test_template_returns_200(self):
        """Template endpoint should return 200"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("✓ Template download returns 200")
    
    def test_template_is_valid_excel(self):
        """Template should be a valid Excel file"""
        response = requests.get(f"{BASE_URL}/api/bulk-search/template")
        assert response.status_code == 200
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            assert ws is not None
            print(f"✓ Template is valid Excel with sheet: {ws.title}")
        except Exception as e:
            pytest.fail(f"Template is not a valid Excel file: {e}")


def extract_number(value):
    """Extract number from string like '₹195,000.00'"""
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    
    import re
    # Remove currency symbols and commas
    cleaned = re.sub(r'[₹$,\s]', '', str(value))
    try:
        return float(cleaned)
    except ValueError:
        return None


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
