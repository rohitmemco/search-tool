"""
Test Bulk Excel Upload with Rate Diff and Amount Diff columns for ALL rate types (Min, Med, Max).

Requirements tested:
1. POST /api/bulk-search/upload returns Excel with 19 columns
2. Column names include: Rate Diff (Min), Amount Diff (Min), Rate Diff (Med), Amount Diff (Med), Rate Diff (Max), Amount Diff (Max)
3. Verify Rate Diff (Min) = Your Rate - Min Rate
4. Verify Amount Diff (Min) = Your Amount - Min Amount  
5. Verify Rate Diff (Med) = Your Rate - Med Rate
6. Verify Amount Diff (Med) = Your Amount - Med Amount
7. Verify Rate Diff (Max) = Your Rate - Max Rate
8. Verify Amount Diff (Max) = Your Amount - Max Amount
9. Verify positive diffs are highlighted red (overpaying)
10. Verify negative diffs are highlighted green (good deal)
11. Verify consolidated GST summary has FOUR sections
"""
import pytest
import requests
import os
import io
import openpyxl
from openpyxl.styles import PatternFill

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestBulkUploadAllDiffs:
    """Test bulk upload with Rate Diff and Amount Diff for ALL rate types"""
    
    @pytest.fixture(scope="class")
    def test_excel_file(self):
        """Create test Excel file with 3 products"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers matching expected format
        ws.append(["SL No", "Item", "Rate/Item", "Qty", "Amount"])
        
        # Test data - Samsung Galaxy S24 (Your Rate 75K)
        ws.append([1, "Samsung Galaxy S24", 75000, 1, 75000])
        
        # MacBook Air M3 (Your Rate 120K)
        ws.append([2, "MacBook Air M3", 120000, 1, 120000])
        
        # Sony WH-1000XM5 (Your Rate 25K)
        ws.append([3, "Sony WH-1000XM5", 25000, 1, 25000])
        
        # Save to BytesIO
        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        return file_buffer
    
    @pytest.fixture(scope="class")
    def upload_response(self, test_excel_file):
        """Upload Excel and get response - runs once for all tests"""
        files = {
            'file': ('test_products.xlsx', test_excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        response = requests.post(f"{BASE_URL}/api/bulk-search/upload", files=files, timeout=120)
        return response
    
    @pytest.fixture(scope="class")
    def output_workbook(self, upload_response):
        """Parse the response Excel file"""
        if upload_response.status_code != 200:
            pytest.skip(f"Upload failed with status {upload_response.status_code}")
        
        return openpyxl.load_workbook(io.BytesIO(upload_response.content))
    
    def test_upload_returns_200(self, upload_response):
        """Test that upload endpoint returns 200 OK"""
        assert upload_response.status_code == 200, f"Expected 200, got {upload_response.status_code}"
        print("✅ POST /api/bulk-search/upload returns 200 OK")
    
    def test_output_has_19_columns(self, output_workbook):
        """Verify output Excel has 19 columns"""
        ws = output_workbook.active
        headers = [cell.value for cell in ws[1]]
        
        # Count non-None headers
        column_count = len([h for h in headers if h is not None])
        
        print(f"Column count: {column_count}")
        print(f"Headers found: {headers[:20]}")
        
        assert column_count == 19, f"Expected 19 columns, got {column_count}. Headers: {headers}"
        print("✅ Output Excel has 19 columns")
    
    def test_column_names_correct(self, output_workbook):
        """Verify all column names are correct including Rate Diff and Amount Diff for Min/Med/Max"""
        ws = output_workbook.active
        headers = [cell.value for cell in ws[1]]
        
        expected_headers = [
            "SL No", "Item", 
            "Your Rate (₹)", "Qty", "Your Amount (₹)",
            "Min Rate (₹)", "Min Amount (₹)", "Rate Diff (Min) (₹)", "Amount Diff (Min) (₹)",
            "Med Rate (₹)", "Med Amount (₹)", "Rate Diff (Med) (₹)", "Amount Diff (Med) (₹)",
            "Max Rate (₹)", "Max Amount (₹)", "Rate Diff (Max) (₹)", "Amount Diff (Max) (₹)",
            "Website Links", "Vendor Details"
        ]
        
        for i, expected in enumerate(expected_headers):
            actual = headers[i] if i < len(headers) else None
            assert actual == expected, f"Column {i+1}: Expected '{expected}', got '{actual}'"
        
        print("✅ All column names correct:")
        print("  - Rate Diff (Min) (₹) at column H")
        print("  - Amount Diff (Min) (₹) at column I")
        print("  - Rate Diff (Med) (₹) at column L")
        print("  - Amount Diff (Med) (₹) at column M")
        print("  - Rate Diff (Max) (₹) at column P")
        print("  - Amount Diff (Max) (₹) at column Q")
    
    def test_rate_diff_min_calculation(self, output_workbook):
        """Verify Rate Diff (Min) = Your Rate - Min Rate"""
        ws = output_workbook.active
        
        # Check first data row (row 2)
        row = 2
        your_rate = ws.cell(row=row, column=3).value  # Column C
        min_rate = ws.cell(row=row, column=6).value    # Column F
        rate_diff_min = ws.cell(row=row, column=8).value  # Column H
        
        print(f"Row {row}: Your Rate={your_rate}, Min Rate={min_rate}, Rate Diff (Min)={rate_diff_min}")
        
        if isinstance(your_rate, (int, float)) and isinstance(min_rate, (int, float)):
            expected_diff = your_rate - min_rate
            assert abs(rate_diff_min - expected_diff) < 0.01, f"Rate Diff (Min) mismatch: {rate_diff_min} != {expected_diff}"
            print(f"✅ Rate Diff (Min) = Your Rate - Min Rate: {your_rate} - {min_rate} = {rate_diff_min}")
        else:
            print(f"⚠️ Skipping calculation check - values not numeric")
    
    def test_amount_diff_min_calculation(self, output_workbook):
        """Verify Amount Diff (Min) = Your Amount - Min Amount"""
        ws = output_workbook.active
        
        row = 2
        your_amount = ws.cell(row=row, column=5).value  # Column E
        min_amount = ws.cell(row=row, column=7).value    # Column G
        amount_diff_min = ws.cell(row=row, column=9).value  # Column I
        
        print(f"Row {row}: Your Amount={your_amount}, Min Amount={min_amount}, Amount Diff (Min)={amount_diff_min}")
        
        if isinstance(your_amount, (int, float)) and isinstance(min_amount, (int, float)):
            expected_diff = your_amount - min_amount
            assert abs(amount_diff_min - expected_diff) < 0.01, f"Amount Diff (Min) mismatch: {amount_diff_min} != {expected_diff}"
            print(f"✅ Amount Diff (Min) = Your Amount - Min Amount: {your_amount} - {min_amount} = {amount_diff_min}")
        else:
            print(f"⚠️ Skipping calculation check - values not numeric")
    
    def test_rate_diff_med_calculation(self, output_workbook):
        """Verify Rate Diff (Med) = Your Rate - Med Rate"""
        ws = output_workbook.active
        
        row = 2
        your_rate = ws.cell(row=row, column=3).value  # Column C
        med_rate = ws.cell(row=row, column=10).value   # Column J
        rate_diff_med = ws.cell(row=row, column=12).value  # Column L
        
        print(f"Row {row}: Your Rate={your_rate}, Med Rate={med_rate}, Rate Diff (Med)={rate_diff_med}")
        
        if isinstance(your_rate, (int, float)) and isinstance(med_rate, (int, float)):
            expected_diff = your_rate - med_rate
            assert abs(rate_diff_med - expected_diff) < 0.01, f"Rate Diff (Med) mismatch: {rate_diff_med} != {expected_diff}"
            print(f"✅ Rate Diff (Med) = Your Rate - Med Rate: {your_rate} - {med_rate} = {rate_diff_med}")
        else:
            print(f"⚠️ Skipping calculation check - values not numeric")
    
    def test_amount_diff_med_calculation(self, output_workbook):
        """Verify Amount Diff (Med) = Your Amount - Med Amount"""
        ws = output_workbook.active
        
        row = 2
        your_amount = ws.cell(row=row, column=5).value  # Column E
        med_amount = ws.cell(row=row, column=11).value   # Column K
        amount_diff_med = ws.cell(row=row, column=13).value  # Column M
        
        print(f"Row {row}: Your Amount={your_amount}, Med Amount={med_amount}, Amount Diff (Med)={amount_diff_med}")
        
        if isinstance(your_amount, (int, float)) and isinstance(med_amount, (int, float)):
            expected_diff = your_amount - med_amount
            assert abs(amount_diff_med - expected_diff) < 0.01, f"Amount Diff (Med) mismatch: {amount_diff_med} != {expected_diff}"
            print(f"✅ Amount Diff (Med) = Your Amount - Med Amount: {your_amount} - {med_amount} = {amount_diff_med}")
        else:
            print(f"⚠️ Skipping calculation check - values not numeric")
    
    def test_rate_diff_max_calculation(self, output_workbook):
        """Verify Rate Diff (Max) = Your Rate - Max Rate"""
        ws = output_workbook.active
        
        row = 2
        your_rate = ws.cell(row=row, column=3).value  # Column C
        max_rate = ws.cell(row=row, column=14).value   # Column N
        rate_diff_max = ws.cell(row=row, column=16).value  # Column P
        
        print(f"Row {row}: Your Rate={your_rate}, Max Rate={max_rate}, Rate Diff (Max)={rate_diff_max}")
        
        if isinstance(your_rate, (int, float)) and isinstance(max_rate, (int, float)):
            expected_diff = your_rate - max_rate
            assert abs(rate_diff_max - expected_diff) < 0.01, f"Rate Diff (Max) mismatch: {rate_diff_max} != {expected_diff}"
            print(f"✅ Rate Diff (Max) = Your Rate - Max Rate: {your_rate} - {max_rate} = {rate_diff_max}")
        else:
            print(f"⚠️ Skipping calculation check - values not numeric")
    
    def test_amount_diff_max_calculation(self, output_workbook):
        """Verify Amount Diff (Max) = Your Amount - Max Amount"""
        ws = output_workbook.active
        
        row = 2
        your_amount = ws.cell(row=row, column=5).value  # Column E
        max_amount = ws.cell(row=row, column=15).value   # Column O
        amount_diff_max = ws.cell(row=row, column=17).value  # Column Q
        
        print(f"Row {row}: Your Amount={your_amount}, Max Amount={max_amount}, Amount Diff (Max)={amount_diff_max}")
        
        if isinstance(your_amount, (int, float)) and isinstance(max_amount, (int, float)):
            expected_diff = your_amount - max_amount
            assert abs(amount_diff_max - expected_diff) < 0.01, f"Amount Diff (Max) mismatch: {amount_diff_max} != {expected_diff}"
            print(f"✅ Amount Diff (Max) = Your Amount - Max Amount: {your_amount} - {max_amount} = {amount_diff_max}")
        else:
            print(f"⚠️ Skipping calculation check - values not numeric")
    
    def test_positive_diff_highlighted_red(self, output_workbook):
        """Verify positive diffs (overpaying) are highlighted red"""
        ws = output_workbook.active
        
        # Red fill color (FFC7CE)
        red_fill_color = "FFC7CE"
        
        # Check all difference columns: H(8), I(9), L(12), M(13), P(16), Q(17)
        diff_columns = [8, 9, 12, 13, 16, 17]
        diff_column_names = ["Rate Diff (Min)", "Amount Diff (Min)", "Rate Diff (Med)", "Amount Diff (Med)", "Rate Diff (Max)", "Amount Diff (Max)"]
        
        positive_diff_found = False
        positive_with_red = 0
        
        for row in range(2, min(5, ws.max_row + 1)):  # Check first 3 data rows
            for col_idx, col_name in zip(diff_columns, diff_column_names):
                cell = ws.cell(row=row, column=col_idx)
                value = cell.value
                
                if isinstance(value, (int, float)) and value > 0:
                    positive_diff_found = True
                    fill = cell.fill
                    if fill and hasattr(fill, 'fgColor') and fill.fgColor:
                        fill_color = str(fill.fgColor.rgb)[-6:] if fill.fgColor.rgb else ""
                        if fill_color.upper() == red_fill_color:
                            positive_with_red += 1
                            print(f"✅ Row {row}, {col_name}: +{value} highlighted RED")
        
        if positive_diff_found:
            assert positive_with_red > 0, "Positive diffs found but none highlighted red"
            print(f"✅ Positive diffs (overpaying) highlighted red: {positive_with_red} cells")
        else:
            print("⚠️ No positive diffs found in data - color highlighting cannot be verified")
    
    def test_negative_diff_highlighted_green(self, output_workbook):
        """Verify negative diffs (good deal) are highlighted green"""
        ws = output_workbook.active
        
        # Green fill color (C6EFCE)
        green_fill_color = "C6EFCE"
        
        # Check all difference columns: H(8), I(9), L(12), M(13), P(16), Q(17)
        diff_columns = [8, 9, 12, 13, 16, 17]
        diff_column_names = ["Rate Diff (Min)", "Amount Diff (Min)", "Rate Diff (Med)", "Amount Diff (Med)", "Rate Diff (Max)", "Amount Diff (Max)"]
        
        negative_diff_found = False
        negative_with_green = 0
        
        for row in range(2, min(5, ws.max_row + 1)):  # Check first 3 data rows
            for col_idx, col_name in zip(diff_columns, diff_column_names):
                cell = ws.cell(row=row, column=col_idx)
                value = cell.value
                
                if isinstance(value, (int, float)) and value < 0:
                    negative_diff_found = True
                    fill = cell.fill
                    if fill and hasattr(fill, 'fgColor') and fill.fgColor:
                        fill_color = str(fill.fgColor.rgb)[-6:] if fill.fgColor.rgb else ""
                        if fill_color.upper() == green_fill_color:
                            negative_with_green += 1
                            print(f"✅ Row {row}, {col_name}: {value} highlighted GREEN")
        
        if negative_diff_found:
            assert negative_with_green > 0, "Negative diffs found but none highlighted green"
            print(f"✅ Negative diffs (good deal) highlighted green: {negative_with_green} cells")
        else:
            print("⚠️ No negative diffs found in data - color highlighting cannot be verified")
    
    def test_gst_summary_has_four_sections(self, output_workbook):
        """Verify consolidated GST summary has FOUR sections"""
        ws = output_workbook.active
        
        # Search for GST summary sections
        sections_found = []
        required_sections = ["YOUR PRICING", "MARKET MINIMUM", "MARKET MEDIUM", "MARKET MAXIMUM"]
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell)
            for section in required_sections:
                if section in row_text and section not in sections_found:
                    sections_found.append(section)
                    print(f"✅ Found GST section: {section}")
        
        missing_sections = [s for s in required_sections if s not in sections_found]
        assert len(missing_sections) == 0, f"Missing GST sections: {missing_sections}"
        
        print(f"✅ Consolidated GST Summary has all FOUR sections: {required_sections}")
    
    def test_gst_summary_header_exists(self, output_workbook):
        """Verify 'CONSOLIDATED GST SUMMARY' header exists"""
        ws = output_workbook.active
        
        header_found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell)
            if "CONSOLIDATED GST SUMMARY" in row_text:
                header_found = True
                print("✅ Found 'CONSOLIDATED GST SUMMARY' header")
                break
        
        assert header_found, "CONSOLIDATED GST SUMMARY header not found"
    
    def test_gst_line_items_present(self, output_workbook):
        """Verify GST line items present: Taxable Amount, CGST, SGST, Round Off, Grand Total"""
        ws = output_workbook.active
        
        required_items = ["Taxable Amount", "CGST", "SGST", "Round Off", "Grand Total"]
        items_found = []
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell)
            for item in required_items:
                if item in row_text and item not in items_found:
                    items_found.append(item)
        
        missing_items = [i for i in required_items if i not in items_found]
        assert len(missing_items) == 0, f"Missing GST line items: {missing_items}"
        
        print(f"✅ All GST line items present: {required_items}")
    
    def test_all_diff_calculations_for_all_rows(self, output_workbook):
        """Verify all diff calculations for all data rows"""
        ws = output_workbook.active
        
        all_correct = True
        errors = []
        
        for row in range(2, min(5, ws.max_row + 1)):  # Check all data rows
            your_rate = ws.cell(row=row, column=3).value
            your_amount = ws.cell(row=row, column=5).value
            min_rate = ws.cell(row=row, column=6).value
            min_amount = ws.cell(row=row, column=7).value
            med_rate = ws.cell(row=row, column=10).value
            med_amount = ws.cell(row=row, column=11).value
            max_rate = ws.cell(row=row, column=14).value
            max_amount = ws.cell(row=row, column=15).value
            
            rate_diff_min = ws.cell(row=row, column=8).value
            amount_diff_min = ws.cell(row=row, column=9).value
            rate_diff_med = ws.cell(row=row, column=12).value
            amount_diff_med = ws.cell(row=row, column=13).value
            rate_diff_max = ws.cell(row=row, column=16).value
            amount_diff_max = ws.cell(row=row, column=17).value
            
            item = ws.cell(row=row, column=2).value
            
            # Check calculations
            if isinstance(your_rate, (int, float)) and isinstance(min_rate, (int, float)):
                expected = your_rate - min_rate
                if isinstance(rate_diff_min, (int, float)) and abs(rate_diff_min - expected) > 0.01:
                    errors.append(f"Row {row} ({item}): Rate Diff (Min) {rate_diff_min} != {expected}")
                    all_correct = False
            
            if isinstance(your_amount, (int, float)) and isinstance(min_amount, (int, float)):
                expected = your_amount - min_amount
                if isinstance(amount_diff_min, (int, float)) and abs(amount_diff_min - expected) > 0.01:
                    errors.append(f"Row {row} ({item}): Amount Diff (Min) {amount_diff_min} != {expected}")
                    all_correct = False
            
            if isinstance(your_rate, (int, float)) and isinstance(med_rate, (int, float)):
                expected = your_rate - med_rate
                if isinstance(rate_diff_med, (int, float)) and abs(rate_diff_med - expected) > 0.01:
                    errors.append(f"Row {row} ({item}): Rate Diff (Med) {rate_diff_med} != {expected}")
                    all_correct = False
            
            if isinstance(your_amount, (int, float)) and isinstance(med_amount, (int, float)):
                expected = your_amount - med_amount
                if isinstance(amount_diff_med, (int, float)) and abs(amount_diff_med - expected) > 0.01:
                    errors.append(f"Row {row} ({item}): Amount Diff (Med) {amount_diff_med} != {expected}")
                    all_correct = False
            
            if isinstance(your_rate, (int, float)) and isinstance(max_rate, (int, float)):
                expected = your_rate - max_rate
                if isinstance(rate_diff_max, (int, float)) and abs(rate_diff_max - expected) > 0.01:
                    errors.append(f"Row {row} ({item}): Rate Diff (Max) {rate_diff_max} != {expected}")
                    all_correct = False
            
            if isinstance(your_amount, (int, float)) and isinstance(max_amount, (int, float)):
                expected = your_amount - max_amount
                if isinstance(amount_diff_max, (int, float)) and abs(amount_diff_max - expected) > 0.01:
                    errors.append(f"Row {row} ({item}): Amount Diff (Max) {amount_diff_max} != {expected}")
                    all_correct = False
            
            if all_correct:
                print(f"✅ Row {row} ({item}): All diff calculations correct")
        
        if errors:
            print("❌ Calculation errors found:")
            for e in errors:
                print(f"  - {e}")
        
        assert all_correct, f"Diff calculation errors: {errors}"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
